import tkinter as tk
from tkinter import messagebox, Entry, Button, Text, Listbox, MULTIPLE, Toplevel, Checkbutton, IntVar, StringVar, ttk, Radiobutton, Label
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
from datetime import datetime
import os
from getpass import getuser
import random
import qrcode
from io import BytesIO
import traceback
from tkcalendar import DateEntry
import pyperclip
import time
import requests  # Teams投稿用
import sqlite3  # SQLite用
import json  # データをJSON形式で保存するため
import glob  # ファイル検索用
import subprocess  # ファイルを開くため
import shutil  # ファイルコピー用
import tempfile  # 一時ファイル用
import PIL.Image
import PIL.ImageTk
# インポート部分に以下を追加
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib import colors
from reportlab.lib.units import mm, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PyPDF2 import PdfWriter, PdfReader
import io
import platform
import sys
import ctypes

#250616：依頼報告書の明細画面でのスクロールの改善、入力内容の記憶、関連出荷指示番号の全表示、統合での分母エラー表示を調査中

# ========== バージョン情報 ==========
APP_VERSION = "1.0.8"  # Current application version
APP_NAME = "事前梱包依頼書管理アプリ"


def minimize_console_window():
    """
    Windows実行時のみ、起動したコンソールウィンドウを最小化します。
    """
    try:
        if sys.platform != "win32":
            return
        hwnd = ctypes.windll.kernel32.GetConsoleWindow()
        if hwnd:
            SW_MINIMIZE = 6
            ctypes.windll.user32.ShowWindow(hwnd, SW_MINIMIZE)
    except Exception:
        # コンソール最小化に失敗してもアプリ本体は継続起動する
        pass


def get_update_folder_path():
    """
    アップデートフォルダのパスを取得します。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\業務用pythonアプリ最新版\\事前梱包依頼書作成\\update",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\業務用pythonアプリ最新版\\事前梱包依頼書作成\\update",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\業務用pythonアプリ最新版\\事前梱包依頼書作成\\update"
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    return None


def check_for_updates():
    """
    アップデートがあるかどうかをチェックします。
    Returns: (has_update, latest_version, update_notes) or (False, None, None) if check fails
    """
    try:
        update_folder = get_update_folder_path()
        if not update_folder:
            print("アップデートフォルダが見つかりません。")
            return False, None, None

        version_file = os.path.join(update_folder, "version.txt")
        if not os.path.exists(version_file):
            print("version.txtが見つかりません。")
            return False, None, None

        with open(version_file, 'r', encoding='utf-8') as f:
            content = f.read().strip()

        # Parse version.txt (format: first line = version, rest = update notes)
        lines = content.split('\n')
        latest_version = lines[0].strip()
        update_notes = '\n'.join(lines[1:]).strip() if len(lines) > 1 else ""

        # Compare versions
        def parse_version(v):
            return tuple(map(int, v.split('.')))

        try:
            if parse_version(latest_version) > parse_version(APP_VERSION):
                return True, latest_version, update_notes
        except:
            pass

        return False, latest_version, update_notes

    except Exception as e:
        print(f"バージョンチェックエラー: {e}")
        return False, None, None


def perform_update():
    """
    アップデートを実行します。
    最新版のファイルを現在のファイルに上書きコピーします。
    """
    try:
        update_folder = get_update_folder_path()
        if not update_folder:
            return False, "アップデートフォルダが見つかりません。"

        # Find the latest .py file in update folder
        update_file = os.path.join(update_folder, "事前梱包依頼書作成.py")
        if not os.path.exists(update_file):
            return False, "アップデートファイルが見つかりません。"

        # Get current file path
        current_file = os.path.abspath(__file__)

        # Create backup in a user-writable location (Program Files is often read-only)
        backup_dir = os.path.join(tempfile.gettempdir(), "事前梱包依頼書作成_backup")
        os.makedirs(backup_dir, exist_ok=True)
        backup_name = f"事前梱包依頼書作成_{datetime.now().strftime('%Y%m%d_%H%M%S')}.py.backup"
        backup_file = os.path.join(backup_dir, backup_name)
        shutil.copy2(current_file, backup_file)

        # Copy new file to current location
        try:
            shutil.copy2(update_file, current_file)
        except PermissionError:
            return False, (
                "アップデートに失敗しました: 現在の配置先に書き込み権限がありません。\n"
                f"配置先: {current_file}\n\n"
                "管理者として実行するか、書き込み可能なフォルダへアプリを配置して再実行してください。"
            )

        return True, f"アップデートが完了しました。\nアプリケーションを再起動してください。\n\nバックアップファイル: {backup_file}"

    except Exception as e:
        return False, f"アップデートに失敗しました: {str(e)}"


def show_update_dialog(parent, latest_version, update_notes):
    """
    アップデートダイアログを表示します。
    """
    result = messagebox.askyesno(
        "アップデートのお知らせ",
        f"新しいバージョンが利用可能です。\n\n"
        f"現在のバージョン: {APP_VERSION}\n"
        f"最新バージョン: {latest_version}\n\n"
        f"{update_notes}\n\n"
        f"今すぐアップデートしますか？"
    )

    if result:
        success, message = perform_update()
        if success:
            messagebox.showinfo("アップデート完了", message)
            # Close the application
            parent.quit()
            parent.destroy()
            import sys
            sys.exit(0)
        else:
            messagebox.showerror("アップデートエラー", message)


def check_and_prompt_update(parent):
    """
    アプリ起動時にアップデートをチェックし、必要に応じてダイアログを表示します。
    """
    has_update, latest_version, update_notes = check_for_updates()
    if has_update:
        # Use after to show dialog after main window is displayed
        parent.after(500, lambda: show_update_dialog(parent, latest_version, update_notes))


# 各梱包担当者ごとに異なるTeamsチャネルのIncoming Webhook URLを指定してください。
TEAMS_WEBHOOK_URLS = {
    "11_細田": "https://defaultaaa1bbd1f29c4c62aa2d1d9a90d71e.cf.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/8d49f16cbbed43b998adbf9cf83ceabf/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=3zNrDxU4tV0sru5Y6lFrKCVTSz3FYcuwDyYokQLFpu0",   # 11_細田用のWebhook URL
    "12_平松": "https://defaultaaa1bbd1f29c4c62aa2d1d9a90d71e.cf.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/eb25b3986e674006b7d5f2b7fca1d7a2/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=cw0fl4gBMwtQYZZmxg_93TiRo65KcQeMpx7MwnDg_ZM",  # 12_平松用のWebhook URL
    "13_坂上": "https://defaultaaa1bbd1f29c4c62aa2d1d9a90d71e.cf.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/45f4d0fb27ed414488525128cc4b2212/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=SXECKoRxDV832wLOrJoZrcvqb3gwOdoW5cDBOGbAA8M",   # 13_坂上用のWebhook URL
    "16_土田": "https://defaultaaa1bbd1f29c4c62aa2d1d9a90d71e.cf.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/90bf16d2ea734953ab910222ae2f5bf7/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=ugCLmnrlw0Vn-zFDmHXFoUclKv_rv4x6APurEqbt-f4"    # 16_土田用のWebhook URL
}

def send_adaptive_card_to_teams(header_info):
    """
    ヘッダー情報（dict）を元に、Adaptive Card形式のメッセージをTeamsチャネルへ投稿します。
    梱包担当者に対応したWebhook URLへ送信するように変更済みです。
    """
    adaptive_card = {
        "$schema": "http://adaptivecards.io/schemas/adaptive-card.json",
        "type": "AdaptiveCard",
        "version": "1.2",
        "body": [
            # 追加：赤文字で注意文言を表示
            {
                "type": "TextBlock",
                "text": "※印刷が完了次第、必ずリアクションで「済マーク」を返すこと。",
                "color": "Attention",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": header_info.get("title", "事前：梱包依頼書"),
                "weight": "Bolder",
                "size": "Large"
            },
            {
                "type": "RichTextBlock",
                "inlines": [
                    {
                        "type": "TextRun",
                        "text": "事前梱包依頼番号",
                        "weight": "Bolder"
                    },
                    {
                        "type": "TextRun",
                        "text": "で該当のExcelファイルを検索・印刷し、梱包を行って下さい"
                    }
                ]
            },
            {
                "type": "TextBlock",
                "text": f"事前梱包依頼番号: {header_info.get('unique_number', '')}",
                "weight": "Bolder",
                "size": "Large",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"梱包期限日: {header_info.get('deadline', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"見積管理番号: {header_info.get('estimate_no', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"船名: {header_info.get('ship_name', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"受注番号: {header_info.get('order_no', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"客注番号: {header_info.get('customer_order_no', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"発注番号: {header_info.get('order_numbers', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"得意先名: {header_info.get('customer_name', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"受渡場所名: {header_info.get('delivery_location', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"営業担当者: {header_info.get('salesperson', '')}",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"梱包担当者: {header_info.get('packing_person', '')}",
                "weight": "Bolder",
                "size": "Large",
                "wrap": True
            },
            {
                "type": "TextBlock",
                "text": f"梱包依頼摘要: {header_info.get('packaging_note', '')}",
                "wrap": True
            }
        ]
    }
    
    if "exclude_inos" in header_info and header_info["exclude_inos"]:
        adaptive_card["body"].append({
            "type": "TextBlock",
            "text": f"出力除外I/no.: {', '.join(map(str, header_info['exclude_inos']))}",
            "wrap": True
        })
    
    payload = {
        "type": "message",
        "attachments": [
            {
                "contentType": "application/vnd.microsoft.card.adaptive",
                "content": adaptive_card
            }
        ]
    }
    
    # header_infoから梱包担当者を取得し、対応するWebhook URLを使用
    packing_person = header_info.get("packing_person")
    webhook_url = TEAMS_WEBHOOK_URLS.get(packing_person)
    if not webhook_url:
        raise ValueError(f"無効な梱包担当者です。packing_person: {packing_person}")
    
    try:
        response = requests.post(webhook_url, json=payload)
        if response.status_code == 200:
            print("Adaptive Card successfully posted to Teams.")
        else:
            print(f"Failed to post Adaptive Card to Teams: {response.status_code} {response.text}")
    except Exception as e:
        print(f"Error posting Adaptive Card to Teams: {e}")

def get_db_path():
    """
    DB配置先を取得します。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\事前梱包依頼関連",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\事前梱包依頼関連",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\事前梱包依頼関連"
    ]
    for folder in candidate_paths:
        if os.path.exists(folder):
            return os.path.join(folder, "packing_request.db")
    raise FileNotFoundError("適切なDB保存先フォルダが見つかりません。候補: " + ", ".join(candidate_paths))

def get_generated_numbers_db_path():
    """
    生成済み番号DBの配置先を取得します。
    メインDBとは別ファイルにすることでロック競合を防止。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\事前梱包依頼関連",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\事前梱包依頼関連",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\事前梱包依頼関連"
    ]
    for folder in candidate_paths:
        if os.path.exists(folder):
            return os.path.join(folder, "generated_numbers.db")
    raise FileNotFoundError("適切なDB保存先フォルダが見つかりません。")


def get_order_db_path():
    """
    受注データDBのパスを取得します。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\受注データ\\order_data.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\受注データ\\order_data.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\受注データ\\order_data.db"
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("受注データDBが見つかりません。")


def get_purchase_order_db_path():
    """
    発注データDBのパスを取得します。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\発注データ\\purchase_order_data.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\発注データ\\purchase_order_data.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\発注データ\\purchase_order_data.db"
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("発注データDBが見つかりません。")


def get_arrival_db_path():
    """
    入荷データDBのパスを取得します。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\入荷データ\\arrival_data.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\入荷データ\\arrival_data.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\入荷データ\\arrival_data.db"
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("入荷データDBが見つかりません。")


def get_inventory_db_path():
    """
    在庫一覧DBのパスを取得します。
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\社内在庫ロケーション一覧\\TYT01-02在庫一覧.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\社内在庫ロケーション一覧\\TYT01-02在庫一覧.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\DB\\社内在庫ロケーション一覧\\TYT01-02在庫一覧.db"
    ]
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("在庫一覧DBが見つかりません。")


def load_inventory_data_from_db():
    """
    在庫一覧DBから必要なカラムのみを読み込みます。
    """
    db_path = get_inventory_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        # Only select required columns for better performance
        query = 'SELECT "商品コード", "ロット番号", "棚番１" FROM TYT01_02在庫一覧'
        df = pd.read_sql_query(query, conn, dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


def load_order_data_from_db():
    """
    受注データDBから必要なカラムのみを読み込みます。
    """
    db_path = get_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        # Only select required columns for better performance
        columns = [
            '受注番号', '客注番号', '得意先名', '得意先', '受渡場所名', '社員名',
            '明細_倉庫コード', '明細_共通項目2', '明細_商品コード', '明細_商品受注名',
            '明細_発注引当仕入数量', '明細_受注数量', '受注件名', '取込伝票番号',
            '明細_出荷売上数量', '明細_自社在庫引当数量', '明細_直接売上数量',
            '明細_自社出荷数量', '明細_受注金額', '明細_共通項目3'
        ]
        columns_str = ', '.join([f'"{col}"' for col in columns])
        query = f"SELECT {columns_str} FROM order_data"
        df = pd.read_sql_query(query, conn, dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


def load_purchase_order_data_from_db():
    """
    発注データDBから必要なカラムのみを読み込みます。
    """
    db_path = get_purchase_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        # Only select required columns for better performance
        columns = ['発注番号', '明細_商品コード', '明細_共通項目2', '受注番号']
        columns_str = ', '.join([f'"{col}"' for col in columns])
        query = f"SELECT {columns_str} FROM purchase_order_data"
        df = pd.read_sql_query(query, conn, dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


def load_arrival_data_from_db():
    """
    入荷データDBから必要なカラムのみを読み込みます。
    """
    db_path = get_arrival_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        # Only select required columns for better performance
        columns = [
            '発注番号', '明細_商品コード', '明細_共通項目2', '明細_ロット番号',
            '明細_共通項目3', '明細_商品略名'
        ]
        columns_str = ', '.join([f'"{col}"' for col in columns])
        query = f"SELECT {columns_str} FROM arrival_data"
        df = pd.read_sql_query(query, conn, dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


# ========== データキャッシュ機構 ==========
# グローバルキャッシュ変数
_data_cache = {
    'order_data': None,           # 受注データ
    'purchase_order_data': None,  # 発注データ
    'arrival_data': None,         # 入荷データ
    'inventory_data': None,       # 在庫一覧データ
    'merged_data': None,          # マージ済みデータ
    'estimate_os_case': {},       # 見積管理番号ごとのOS案件判定キャッシュ
    'last_update': None           # 最終更新時刻
}


def clear_data_cache():
    """
    データキャッシュをクリアします。
    新しいデータを読み込む必要がある場合に呼び出してください。
    """
    global _data_cache
    _data_cache = {
        'order_data': None,
        'purchase_order_data': None,
        'arrival_data': None,
        'inventory_data': None,
        'merged_data': None,
        'estimate_os_case': {},
        'last_update': None
    }
    print("データキャッシュをクリアしました。")


def is_os_case_by_estimate_no(estimate_no):
    """
    見積管理番号に紐づく受注番号にOS始まりが含まれるかを軽量SQLで判定します。
    """
    global _data_cache
    estimate_no = clean_input(estimate_no)
    if not estimate_no:
        return False

    cache_map = _data_cache.get('estimate_os_case', {})
    if estimate_no in cache_map:
        return cache_map[estimate_no]

    db_path = get_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        cursor = conn.cursor()
        cursor.execute(
            'SELECT 1 FROM order_data WHERE "取込伝票番号" = ? AND UPPER("受注番号") LIKE \'OS%\' LIMIT 1',
            (estimate_no,)
        )
        is_os_case = cursor.fetchone() is not None
        cache_map[estimate_no] = is_os_case
        _data_cache['estimate_os_case'] = cache_map
        return is_os_case
    finally:
        conn.close()


def get_cached_order_data():
    """
    キャッシュから受注データを取得します。
    キャッシュがない場合はDBから読み込みます。
    """
    global _data_cache
    if _data_cache['order_data'] is None:
        print("受注データをDBから読み込み中...")
        _data_cache['order_data'] = load_order_data_from_db()
        print(f"受注データ読み込み完了: {len(_data_cache['order_data'])}件")
    return _data_cache['order_data']


def get_cached_purchase_order_data():
    """
    キャッシュから発注データを取得します。
    キャッシュがない場合はDBから読み込みます。
    """
    global _data_cache
    if _data_cache['purchase_order_data'] is None:
        print("発注データをDBから読み込み中...")
        _data_cache['purchase_order_data'] = load_purchase_order_data_from_db()
        print(f"発注データ読み込み完了: {len(_data_cache['purchase_order_data'])}件")
    return _data_cache['purchase_order_data']


def get_cached_arrival_data():
    """
    キャッシュから入荷データを取得します。
    キャッシュがない場合はDBから読み込みます。
    """
    global _data_cache
    if _data_cache['arrival_data'] is None:
        print("入荷データをDBから読み込み中...")
        _data_cache['arrival_data'] = load_arrival_data_from_db()
        print(f"入荷データ読み込み完了: {len(_data_cache['arrival_data'])}件")
    return _data_cache['arrival_data']


def get_cached_inventory_data():
    """
    キャッシュから在庫一覧データを取得します。
    キャッシュがない場合はDBから読み込みます。
    """
    global _data_cache
    if _data_cache['inventory_data'] is None:
        print("在庫一覧データをDBから読み込み中...")
        _data_cache['inventory_data'] = load_inventory_data_from_db()
        print(f"在庫一覧データ読み込み完了: {len(_data_cache['inventory_data'])}件")
    return _data_cache['inventory_data']


def get_merged_data_for_packing():
    """
    梱包依頼用のマージ済みデータを取得します。
    キャッシュがある場合はキャッシュから返します。
    """
    global _data_cache

    if _data_cache['merged_data'] is not None:
        print("マージ済みデータをキャッシュから取得")
        return _data_cache['merged_data']

    print("マージ済みデータを作成中...")

    # 発注データを取得
    df_order = get_cached_purchase_order_data()
    df1 = df_order[['発注番号','明細_商品コード','明細_共通項目2','受注番号']].copy()
    df1['A'] = df1['発注番号'] + df1['明細_商品コード'] + df1['明細_共通項目2']
    df1['B'] = df1['明細_共通項目2']
    df1['C'] = df1['受注番号']
    df1['D'] = df1['受注番号'] + df1['明細_共通項目2']
    df1['E'] = df1['発注番号']
    df1 = df1[['A','B','C','D','E']]

    # 入荷データを取得
    df_arrival = get_cached_arrival_data()
    df2 = df_arrival[['発注番号','明細_商品コード','明細_共通項目2','明細_ロット番号']].copy()
    df2['key'] = df2['発注番号'] + df2['明細_商品コード'] + df2['明細_共通項目2']
    df3 = pd.merge(df1[['A','C','B']],
                   df2[['key','明細_ロット番号']],
                   left_on='A', right_on='key', how='inner')
    df3.drop('key', axis=1, inplace=True)
    df3.rename(columns={'C':'受注番号','B':'明細_共通項目2'}, inplace=True)

    # 受注データを取得
    df_order_list = get_cached_order_data()
    excluded_codes = ['888888-88888','777777-77777']

    columns_needed = [
        '受注番号','客注番号','得意先名','得意先','受渡場所名','社員名','明細_倉庫コード',
        '明細_共通項目2','明細_商品コード','明細_商品受注名',
        '明細_発注引当仕入数量','明細_受注数量','受注件名','取込伝票番号',
        '明細_出荷売上数量','明細_自社在庫引当数量','明細_直接売上数量',
        '明細_自社出荷数量','明細_受注金額'
    ]
    df4 = df_order_list[~df_order_list['明細_商品コード'].isin(excluded_codes)][columns_needed].copy()
    df4 = pd.merge(df4,
                   df3[['受注番号','明細_共通項目2','明細_ロット番号']],
                   on=['受注番号','明細_共通項目2'],
                   how='left')
    df4['key'] = df4['受注番号'] + df4['明細_共通項目2']
    df4 = pd.merge(df4, df1[['D','E']], left_on='key', right_on='D', how='left')
    df4.rename(columns={'E':'発注番号'}, inplace=True)
    df4.drop(['key','D'], axis=1, inplace=True)

    for col in df4.columns:
        if col in ['明細_共通項目2','明細_発注引当仕入数量','明細_受注数量','明細_出荷売上数量','明細_自社在庫引当数量','明細_直接売上数量']:
            df4[col] = pd.to_numeric(df4[col], errors='coerce').fillna(0).astype(int)
        elif col == '明細_受注金額':
            df4[col] = pd.to_numeric(df4[col], errors='coerce').fillna(0.0)
        elif col == '明細_ロット番号':
            df4[col] = df4[col].fillna('')
        else:
            df4[col] = df4[col].replace('nan','').fillna('')

    df4['受注残数'] = df4.apply(
        lambda row: row['明細_受注数量'] - row['明細_直接売上数量']
                    if row['明細_倉庫コード'] == '99999'
                    else row['明細_受注数量'] - row['明細_出荷売上数量'],
        axis=1
    )
    df4.drop_duplicates(subset=['受注番号','明細_共通項目2'], keep='first', inplace=True)

    _data_cache['merged_data'] = df4
    _data_cache['last_update'] = datetime.now()
    print(f"マージ済みデータ作成完了: {len(df4)}件")

    return df4


def query_order_by_number(order_number):
    """
    受注番号で絞り込んだデータを取得します。
    SQLクエリで直接絞り込むため高速です。
    """
    db_path = get_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        query = "SELECT * FROM order_data WHERE 受注番号 = ?"
        df = pd.read_sql_query(query, conn, params=(order_number,), dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


def query_order_by_estimate_no(estimate_no):
    """
    見積管理番号（取込伝票番号）で絞り込んだデータを取得します。
    SQLクエリで直接絞り込むため高速です。
    """
    db_path = get_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        query = "SELECT * FROM order_data WHERE 取込伝票番号 = ?"
        df = pd.read_sql_query(query, conn, params=(estimate_no,), dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


COMMON_PERSON_TO_PACKING_PERSON = {
    '細田宗之介': '11_細田',
    '坂上敦士': '13_坂上',
    '平松良太': '12_平松',
    '土田周平': '16_土田'
}


def normalize_person_name(name):
    """氏名の空白（半角/全角）を除去して比較しやすい形式にします。"""
    if name is None:
        return ''
    return ''.join(str(name).replace('\u3000', ' ').split())


def query_common_person_name(input_value, search_method):
    """
    入力番号（受注番号 or 見積管理番号）に対応する共通項目2名を取得します。
    一致データが複数ある場合は、最初に見つかった非空の値を返します。
    """
    normalized_input = clean_input(str(input_value)).strip()
    if not normalized_input:
        return ''

    if search_method == '受注番号':
        where_clause = '"受注番号" = ?'
    elif search_method == '見積管理番号':
        where_clause = '"取込伝票番号" = ?'
    else:
        return ''

    db_path = get_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        query = f'''
            SELECT "共通項目2名"
            FROM order_data
            WHERE {where_clause}
              AND TRIM(COALESCE("共通項目2名", '')) <> ''
            LIMIT 1
        '''
        cursor = conn.cursor()
        cursor.execute(query, (normalized_input,))
        row = cursor.fetchone()
        if not row or row[0] is None:
            return ''
        return str(row[0]).strip()
    finally:
        conn.close()


def query_purchase_order_by_order_numbers(order_numbers):
    """
    受注番号リストで発注データを絞り込んで取得します。
    """
    if not order_numbers:
        return pd.DataFrame()

    db_path = get_purchase_order_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        placeholders = ','.join(['?' for _ in order_numbers])
        query = f"SELECT * FROM purchase_order_data WHERE 受注番号 IN ({placeholders})"
        df = pd.read_sql_query(query, conn, params=order_numbers, dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


def query_arrival_by_order_numbers(order_numbers):
    """
    発注番号リストで入荷データを絞り込んで取得します。
    """
    if not order_numbers:
        return pd.DataFrame()

    db_path = get_arrival_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    try:
        placeholders = ','.join(['?' for _ in order_numbers])
        query = f"SELECT * FROM arrival_data WHERE 発注番号 IN ({placeholders})"
        df = pd.read_sql_query(query, conn, params=order_numbers, dtype=str)
        df = df.fillna('')
        return df
    finally:
        conn.close()


class BatchDBConnection:
    """
    バッチ処理用のDB接続管理クラス。
    複数件の処理を実行する際に、1つのDB接続を使い回すことで
    OneDrive同期との競合による破損を防止します。
    """
    _instance = None
    _conn = None
    _generated_numbers_conn = None
    _same_db = False  # メインDBと生成済み番号DBが同じかどうか

    @classmethod
    def get_instance(cls):
        if cls._instance is None:
            cls._instance = cls()
        return cls._instance

    def start_batch(self):
        """バッチ処理開始時にDB接続を確立"""
        try:
            db_path = get_db_path()
            gen_db_path = get_generated_numbers_db_path()

            # メインDB接続
            self._conn = sqlite3.connect(db_path, timeout=60.0)
            self._conn.execute("PRAGMA journal_mode=WAL")

            # 生成済み番号DBが同じパスかどうかを確認
            if db_path == gen_db_path:
                # 同じDBの場合は同じ接続を共有
                self._generated_numbers_conn = self._conn
                self._same_db = True
                print("バッチDB接続を確立しました（単一DB）")
            else:
                # 異なるDBの場合は別接続
                self._generated_numbers_conn = sqlite3.connect(gen_db_path, timeout=60.0)
                self._generated_numbers_conn.execute("PRAGMA journal_mode=WAL")
                self._same_db = False
                print("バッチDB接続を確立しました（複数DB）")

            return True
        except Exception as e:
            print(f"バッチDB接続の確立に失敗: {e}")
            return False

    def end_batch(self):
        """バッチ処理終了時にDB接続をクローズ"""
        try:
            if self._conn:
                self._conn.commit()
                self._conn.close()
                self._conn = None

            # 別DBの場合のみ生成済み番号DB接続をクローズ
            if not self._same_db and self._generated_numbers_conn:
                self._generated_numbers_conn.commit()
                self._generated_numbers_conn.close()

            self._generated_numbers_conn = None
            self._same_db = False
            print("バッチDB接続をクローズしました")
        except Exception as e:
            print(f"バッチDB接続のクローズでエラー: {e}")

    def get_main_connection(self):
        """メインDBの接続を取得"""
        return self._conn

    def get_generated_numbers_connection(self):
        """生成済み番号DBの接続を取得"""
        return self._generated_numbers_conn

    def is_batch_active(self):
        """バッチ処理が有効かどうか"""
        return self._conn is not None


def init_database():
    """
    DBを初期化し、必要なテーブルを作成します。
    """
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # 梱包依頼書情報テーブル
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS packing_requests (
        unique_number TEXT PRIMARY KEY,
        order_number TEXT,
        deadline TEXT,
        estimate_no TEXT,
        ship_name TEXT,
        customer_order_no TEXT,
        order_numbers TEXT,
        customer_name TEXT,
        customer_code TEXT,
        delivery_location TEXT,
        salesperson TEXT,
        packing_person TEXT,
        packaging_note TEXT,
        exclude_inos TEXT,
        created_at TEXT,
        details TEXT,
        item_count INTEGER,
        order_amount REAL,
        packing_detail INTEGER DEFAULT 0,
        output_file_path TEXT,
        is_deleted INTEGER DEFAULT 0,
        deleted_at TEXT,
        deleted_by TEXT,
        delete_reason TEXT
    )
    """)

    # 既存のテーブルにitem_countカラムが存在しない場合は追加
    cursor.execute("PRAGMA table_info(packing_requests)")
    columns = [column[1] for column in cursor.fetchall()]
    if 'item_count' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN item_count INTEGER")

    # 既存のテーブルにorder_amountカラムが存在しない場合は追加
    if 'order_amount' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN order_amount REAL")
    # 既存のテーブルにcustomer_codeカラムが存在しない場合は追加
    if 'customer_code' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN customer_code TEXT")

    # 既存のテーブルにweightカラムが存在しない場合は追加
    if 'weight' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN weight REAL")
    # 既存のテーブルにpacking_detailカラムが存在しない場合は追加
    if 'packing_detail' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN packing_detail INTEGER DEFAULT 0")
    # 既存のテーブルにoutput_file_pathカラムが存在しない場合は追加
    if 'output_file_path' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN output_file_path TEXT")
    # 既存のテーブルにis_deletedカラムが存在しない場合は追加
    if 'is_deleted' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN is_deleted INTEGER DEFAULT 0")
    # 既存のテーブルにdeleted_atカラムが存在しない場合は追加
    if 'deleted_at' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN deleted_at TEXT")
    # 既存のテーブルにdeleted_byカラムが存在しない場合は追加
    if 'deleted_by' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN deleted_by TEXT")
    # 既存のテーブルにdelete_reasonカラムが存在しない場合は追加
    if 'delete_reason' not in columns:
        cursor.execute("ALTER TABLE packing_requests ADD COLUMN delete_reason TEXT")

    conn.commit()

    # ===== 生成済み番号DBの初期化と移行 =====
    gen_db_path = get_generated_numbers_db_path()
    gen_conn = sqlite3.connect(gen_db_path)
    gen_cursor = gen_conn.cursor()

    # 新しいDBにテーブル作成
    gen_cursor.execute("""
    CREATE TABLE IF NOT EXISTS generated_numbers (
        number TEXT PRIMARY KEY
    )
    """)

    # 旧DBからのデータ移行（メインDBにgenerated_numbersテーブルが存在する場合）
    try:
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='generated_numbers'")
        if cursor.fetchone():
            # 旧DBから番号を取得
            cursor.execute("SELECT number FROM generated_numbers")
            old_numbers = cursor.fetchall()

            if old_numbers:
                # 新DBに既存データがあるか確認
                gen_cursor.execute("SELECT COUNT(*) FROM generated_numbers")
                new_count = gen_cursor.fetchone()[0]

                if new_count == 0:
                    # 新DBが空の場合のみ移行
                    print(f"旧DBから{len(old_numbers)}件の生成済み番号を移行中...")
                    for (number,) in old_numbers:
                        try:
                            gen_cursor.execute("INSERT OR IGNORE INTO generated_numbers (number) VALUES (?)", (number,))
                        except:
                            pass
                    gen_conn.commit()
                    print(f"移行完了: {len(old_numbers)}件")
                else:
                    print(f"新DBに既にデータが存在するため移行をスキップ（既存: {new_count}件）")
    except Exception as e:
        print(f"データ移行中にエラー（無視して続行）: {e}")

    gen_conn.commit()
    gen_conn.close()
    conn.close()

def load_generated_numbers():
    """
    SQLite DBから既に登録済みの事前梱包依頼番号をsetで返します。
    """
    db_path = get_generated_numbers_db_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS generated_numbers (number TEXT PRIMARY KEY)")
    cursor.execute("SELECT number FROM generated_numbers")
    rows = cursor.fetchall()
    conn.close()
    return set(row[0] for row in rows)

def save_generated_number(number):
    """
    SQLite DBに生成済みの事前梱包依頼番号を登録します。
    """
    db_path = get_generated_numbers_db_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("CREATE TABLE IF NOT EXISTS generated_numbers (number TEXT PRIMARY KEY)")
    cursor.execute("INSERT INTO generated_numbers (number) VALUES (?)", (number,))
    conn.commit()
    conn.close()

def create_progress_dialog(message):
    """
    進行状況ダイアログを作成
    """
    progress_dialog = Toplevel()
    progress_dialog.title("処理中")
    progress_dialog.geometry("350x120")
    progress_dialog.resizable(False, False)
    
    # ダイアログを中央に配置
    progress_dialog.update_idletasks()
    x = (progress_dialog.winfo_screenwidth() // 2) - (progress_dialog.winfo_width() // 2)
    y = (progress_dialog.winfo_screenheight() // 2) - (progress_dialog.winfo_height() // 2)
    progress_dialog.geometry(f"+{x}+{y}")
    
    # メッセージラベル
    msg_label = Label(progress_dialog, text=message, font=("Arial", 10))
    msg_label.pack(pady=15)
    
    # プログレスバー（不確定）
    progress_bar = ttk.Progressbar(progress_dialog, mode='indeterminate', length=250)
    progress_bar.pack(pady=10, padx=20)
    progress_bar.start()
    
    # ダイアログを最前面に
    progress_dialog.transient()
    progress_dialog.grab_set()
    progress_dialog.update()
    
    return progress_dialog, msg_label

def get_db_connection_with_progress(db_path, max_retries=5, progress_label=None):
    """
    進行状況を表示しながらリトライ機能付きのDB接続
    """
    for attempt in range(max_retries):
        try:
            if progress_label:
                progress_label.config(text=f"データベース接続中... ({attempt + 1}/{max_retries})")
                progress_label.update()
            
            conn = sqlite3.connect(db_path, timeout=30.0)
            conn.execute("PRAGMA journal_mode=WAL")
            return conn
            
        except sqlite3.OperationalError as e:
            print(f"DB接続試行 {attempt + 1}/{max_retries}: {e}")
            
            if attempt < max_retries - 1:
                if progress_label:
                    progress_label.config(text=f"接続待機中... ({attempt + 1}/{max_retries})")
                    progress_label.update()
                
                wait_time = random.uniform(0.1, 0.5)
                time.sleep(wait_time)
                continue
            else:
                raise e

def save_packing_request_with_detailed_feedback(header_info, detail_df, use_batch_connection=False):
    """
    詳細な進行状況を表示しながらDB保存
    use_batch_connection: Trueの場合、BatchDBConnectionを使用
    """
    progress_dialog = None
    use_external = False
    conn = None

    try:
        # バッチ接続を使用するかどうか
        if use_batch_connection:
            batch_db = BatchDBConnection.get_instance()
            if batch_db.is_batch_active():
                conn = batch_db.get_main_connection()
                use_external = True

        if not use_external:
            # 進行状況ダイアログを作成（バッチ時は表示しない）
            progress_dialog, progress_label = create_progress_dialog("処理を開始しています...")

            # DB接続（リトライ機能付き）
            db_path = get_db_path()
            conn = get_db_connection_with_progress(db_path, progress_label=progress_label)

            # 保存処理の進行状況を更新
            progress_label.config(text="データを保存中...")
            progress_label.update()

        cursor = conn.cursor()

        # 詳細情報をJSON形式に変換
        details_json = detail_df.to_json(orient='records')
        exclude_inos_json = json.dumps(header_info.get('exclude_inos', []))
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # アイテム数を計算（梱包可能数が0より大きいレコードのみをカウント）
        item_count = len(detail_df[detail_df['梱包可能数'] > 0])

        # ★★★ ここから追加 ★★★
        # 総重量を計算
        total_weight = calculate_total_weight(detail_df)
        # ★★★ ここまで追加 ★★★

        # データ挿入（order_amountとweightを追加）
        cursor.execute("""
        INSERT OR REPLACE INTO packing_requests
        (unique_number, order_number, deadline, estimate_no, ship_name, customer_order_no,
         order_numbers, customer_name, customer_code, delivery_location, salesperson, packing_person,
         packaging_note, exclude_inos, created_at, details, item_count, order_amount, weight, packing_detail,
         output_file_path, is_deleted, deleted_at, deleted_by, delete_reason)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            header_info['unique_number'],
            header_info['order_no'],
            header_info['deadline'],
            header_info['estimate_no'],
            header_info['ship_name'],
            header_info['customer_order_no'],
            header_info['order_numbers'],
            header_info['customer_name'],
            header_info.get('customer_code', ''),
            header_info['delivery_location'],
            header_info['salesperson'],
            header_info['packing_person'],
            header_info['packaging_note'],
            exclude_inos_json,
            created_at,
            details_json,
            item_count,
            header_info.get('order_amount', 0.0),
            total_weight,  # ★追加
            int(header_info.get('packing_detail', 0)),
            header_info.get('output_file_path', ''),
            0,
            None,
            None,
            None
        ))

        conn.commit()

        if not use_external:
            conn.close()

        # 成功時
        if progress_dialog:
            progress_dialog.destroy()
        return True

    except sqlite3.OperationalError as e:
        if progress_dialog:
            progress_dialog.destroy()
        messagebox.showerror("データベースエラー",
            f"データベースへのアクセスに失敗しました。\n"
            f"他のユーザーが同時に処理を行っている可能性があります。\n"
            f"しばらく時間をおいてから再度お試しください。\n\n"
            f"詳細: {str(e)}")
        return False
    except Exception as e:
        if progress_dialog:
            progress_dialog.destroy()
        messagebox.showerror("エラー", f"保存中にエラーが発生しました: {str(e)}")
        return False

def load_generated_numbers_with_retry(external_conn=None):
    """
    リトライ機能付きの生成済み番号読み込み
    external_conn: バッチ処理時に外部から渡される接続（省略時は新規接続）
    """
    try:
        use_external = external_conn is not None
        if use_external:
            conn = external_conn
        else:
            db_path = get_generated_numbers_db_path()
            conn = get_db_connection_with_progress(db_path, max_retries=3)

        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS generated_numbers (number TEXT PRIMARY KEY)")
        cursor.execute("SELECT number FROM generated_numbers")
        rows = cursor.fetchall()

        if not use_external:
            conn.close()

        return set(row[0] for row in rows)
    except Exception as e:
        print(f"生成済み番号の読み込みでエラー: {e}")
        return set()  # エラー時は空のセットを返す

def save_generated_number_with_retry(number, external_conn=None):
    """
    リトライ機能付きの生成済み番号保存
    external_conn: バッチ処理時に外部から渡される接続（省略時は新規接続）
    """
    try:
        use_external = external_conn is not None
        if use_external:
            conn = external_conn
        else:
            db_path = get_generated_numbers_db_path()
            conn = get_db_connection_with_progress(db_path, max_retries=3)

        cursor = conn.cursor()
        cursor.execute("CREATE TABLE IF NOT EXISTS generated_numbers (number TEXT PRIMARY KEY)")
        cursor.execute("INSERT INTO generated_numbers (number) VALUES (?)", (number,))
        conn.commit()

        if not use_external:
            conn.close()

        return True
    except Exception as e:
        print(f"生成済み番号の保存でエラー: {e}")
        return False

def save_packing_request(header_info, detail_df):
    """
    梱包依頼書情報をDBに保存します。
    """
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 詳細情報をJSON形式に変換
    details_json = detail_df.to_json(orient='records')
    
    # 除外I/noをJSONに変換
    exclude_inos_json = json.dumps(header_info.get('exclude_inos', []))
    
    # 現在の日時
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # アイテム数を計算（梱包可能数が0より大きいレコードのみをカウント）
    item_count = len(detail_df[detail_df['梱包可能数'] > 0])
    
    # ★★★ ここから追加 ★★★
    # 総重量を計算
    total_weight = calculate_total_weight(detail_df)
    # ★★★ ここまで追加 ★★★
    
    # データ挿入
    cursor.execute("""
    INSERT OR REPLACE INTO packing_requests
    (unique_number, order_number, deadline, estimate_no, ship_name, customer_order_no, 
     order_numbers, customer_name, customer_code, delivery_location, salesperson, packing_person, 
     packaging_note, exclude_inos, created_at, details, item_count, weight, packing_detail,
     output_file_path, is_deleted, deleted_at, deleted_by, delete_reason)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        header_info['unique_number'],
        header_info['order_no'],
        header_info['deadline'],
        header_info['estimate_no'],
        header_info['ship_name'],
        header_info['customer_order_no'],
        header_info['order_numbers'],
        header_info['customer_name'],
        header_info.get('customer_code', ''),
        header_info['delivery_location'],
        header_info['salesperson'],
        header_info['packing_person'],
        header_info['packaging_note'],
        exclude_inos_json,
        created_at,
        details_json,
        item_count,
        total_weight,  # ★追加
        int(header_info.get('packing_detail', 0)),
        header_info.get('output_file_path', ''),
        0,
        None,
        None,
        None
    ))
    
    conn.commit()
    conn.close()

def load_packing_request(unique_number, include_deleted=False):
    """
    指定された事前梱包依頼番号の情報をDBから取得します。
    """
    db_path = get_db_path()
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    query = "SELECT * FROM packing_requests WHERE unique_number = ?"
    params = [unique_number]
    if not include_deleted:
        query += " AND IFNULL(is_deleted, 0) = 0"
    cursor.execute(query, params)
    
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return None
    
    # カラム名を取得
    cursor.execute("PRAGMA table_info(packing_requests)")
    columns = [col[1] for col in cursor.fetchall()]
    conn.close()
    
    # 結果を辞書に変換
    result = dict(zip(columns, row))
    
    # 詳細情報をJSONからDataFrameに変換
    result['details'] = pd.read_json(result['details'], orient='records')
    
    # 除外I/noをJSONからリストに変換
    result['exclude_inos'] = json.loads(result['exclude_inos'])
    
    return result

def soft_delete_packing_request(unique_number, delete_reason=""):
    """
    指定の梱包依頼を論理削除し、対応Excelファイルを削除済みフォルダへ退避します。
    """
    db_path = get_db_path()
    conn = sqlite3.connect(db_path, timeout=30.0)
    conn.execute("PRAGMA journal_mode=WAL")
    cursor = conn.cursor()

    try:
        cursor.execute("""
        SELECT unique_number, output_file_path, packing_person, IFNULL(is_deleted, 0)
        FROM packing_requests
        WHERE unique_number = ?
        """, (unique_number,))
        row = cursor.fetchone()

        if not row:
            return False, f"事前梱包依頼番号 [{unique_number}] は見つかりません。"

        _, output_file_path, packing_person, is_deleted = row
        if int(is_deleted) == 1:
            return False, f"事前梱包依頼番号 [{unique_number}] は既に削除済みです。"

        resolved_output_file_path = resolve_output_file_path_for_current_pc(output_file_path, packing_person)
        moved_path = ""
        file_status_note = "関連ファイルなし"

        # 出力ファイルが存在する場合は「削除済み」フォルダへ退避
        if resolved_output_file_path and os.path.exists(resolved_output_file_path):
            output_dir = os.path.dirname(resolved_output_file_path)
            archive_dir = os.path.join(output_dir, "削除済み")
            os.makedirs(archive_dir, exist_ok=True)

            src_name = os.path.basename(resolved_output_file_path)
            dst_path = os.path.join(archive_dir, src_name)
            if os.path.exists(dst_path):
                base, ext = os.path.splitext(src_name)
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                dst_path = os.path.join(archive_dir, f"{base}_deleted_{stamp}{ext}")

            shutil.move(resolved_output_file_path, dst_path)
            moved_path = dst_path
            file_status_note = f"ファイル退避先: {dst_path}"
        elif resolved_output_file_path:
            file_status_note = f"元ファイル未検出: {resolved_output_file_path}"

        deleted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        deleted_by = getuser()
        reason = (delete_reason or "").strip()

        cursor.execute("""
        UPDATE packing_requests
        SET is_deleted = 1,
            deleted_at = ?,
            deleted_by = ?,
            delete_reason = ?,
            output_file_path = ?
        WHERE unique_number = ?
        """, (deleted_at, deleted_by, reason, moved_path or resolved_output_file_path, unique_number))

        conn.commit()
        return True, f"事前梱包依頼番号 [{unique_number}] を削除しました。\n{file_status_note}"
    except Exception as e:
        conn.rollback()
        return False, f"削除処理中にエラーが発生しました: {str(e)}"
    finally:
        conn.close()

# 新機能：過去の梱包依頼から梱包可能数を検索する関数
def get_previous_packing_quantities(order_number, i_no):
    """
    指定された受注番号とI/no.に対応する過去の梱包可能数を取得します。
    """
    try:
        db_path = get_db_path()
        conn = get_db_connection_with_progress(db_path, max_retries=3)
        cursor = conn.cursor()
        
        cursor.execute("""
        SELECT details FROM packing_requests
        WHERE order_number = ?
          AND IFNULL(is_deleted, 0) = 0
        """, (order_number,))
        
        rows = cursor.fetchall()
        conn.close()
        
        total_previous_quantity = 0
        
        for row in rows:
            details_json = row[0]
            details_df = pd.read_json(details_json, orient='records')
            
            # 該当するI/no.のレコードを検索
            matching_records = details_df[details_df['明細_共通項目2'] == i_no]
            if not matching_records.empty:
                for _, record in matching_records.iterrows():
                    total_previous_quantity += record.get('梱包可能数', 0)
        
        return total_previous_quantity
        
    except Exception as e:
        print(f"過去の梱包依頼検索でエラー: {e}")
        return 0

def get_packing_list_db_path():
    """
    packing_list.dbファイルのパスを取得
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\梱包明細\\packing_list.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\梱包明細\\packing_list.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\梱包明細\\packing_list.db"
    ]
    
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    
    raise FileNotFoundError("packing_list.dbファイルが見つかりません。候補: " + ", ".join(candidate_paths))

def get_shipment_numbers_by_order(order_number):
    """
    受注番号に対応するshipment_numberのリストを取得
    """
    try:
        db_path = get_packing_list_db_path()
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
        SELECT DISTINCT shipment_number FROM packing_details
        WHERE order_number = ?
        ORDER BY shipment_number
        """, (order_number,))
        
        rows = cursor.fetchall()
        conn.close()
        
        return [row[0] for row in rows]
        
    except Exception as e:
        print(f"shipment_number取得エラー: {e}")
        return []

def get_packing_details_by_shipment(shipment_number):
    """
    shipment_numberに対応する梱包明細データを取得
    """
    try:
        db_path = get_packing_list_db_path()
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
        SELECT case_number, length, width, height, weight, item_details, packing_style
        FROM packing_details
        WHERE shipment_number = ?
        ORDER BY case_number
        """, (shipment_number,))
        
        rows = cursor.fetchall()
        conn.close()
        
        # 辞書形式で返却
        packing_data = []
        for row in rows:
            packing_data.append({
                'case_number': row[0],
                'length': row[1],
                'width': row[2],
                'height': row[3],
                'weight': row[4],
                'item_details': row[5],
                'packing_style': row[6]
            })
        
        return packing_data
        
    except Exception as e:
        print(f"梱包明細取得エラー: {e}")
        return []

def get_all_packing_details_by_order(order_number):
    """
    受注番号に関連するすべてのshipment_numberの梱包明細を取得
    """
    try:
        # まず、受注番号に対応するすべてのshipment_numberを取得
        shipment_numbers = get_shipment_numbers_by_order(order_number)
        
        all_packing_data = []
        for shipment_number in shipment_numbers:
            packing_data = get_packing_details_by_shipment(shipment_number)
            for item in packing_data:
                item['shipment_number'] = shipment_number  # shipment_numberを追加
                all_packing_data.append(item)
        
        return all_packing_data
        
    except Exception as e:
        print(f"受注番号の梱包明細取得エラー: {e}")
        return []

class PackingQuantityDialog:
    """
    梱包可能数を手動で変更するためのダイアログクラス
    """
    def __init__(self, parent, items_data, target_number, search_method):
        self.parent = parent
        self.items_data = items_data.sort_values('明細_共通項目2')  # i/no.の昇順でソート
        self.result = None
        
        # ダイアログウィンドウを作成
        self.dialog = Toplevel(parent)
        search_type = "受注番号" if search_method == "order_number" else "見積管理番号"
        self.dialog.title(f"梱包可能数の変更 - {search_type}: {target_number}")
        self.dialog.geometry("800x600")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # メインフレーム
        main_frame = tk.Frame(self.dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # 説明ラベル（対象番号を表示）
        info_label = tk.Label(main_frame, 
                             text=f"【{search_type}: {target_number}】\n梱包可能数を変更したいアイテムの数量を入力してください。\n※過去の依頼分は受注数量から自動で差し引かれています。",
                             font=("Arial", 11, "bold"), fg="blue")
        info_label.pack(pady=(0, 10))
        
        # スクロール可能なフレーム
        canvas = tk.Canvas(main_frame)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # ヘッダー
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(header_frame, text="I/no.", width=8, relief=tk.RIDGE).pack(side=tk.LEFT)
        tk.Label(header_frame, text="商品コード", width=15, relief=tk.RIDGE).pack(side=tk.LEFT)
        tk.Label(header_frame, text="商品名", width=30, relief=tk.RIDGE).pack(side=tk.LEFT)
        tk.Label(header_frame, text="受注数量", width=10, relief=tk.RIDGE).pack(side=tk.LEFT)
        tk.Label(header_frame, text="元の梱包可能数", width=15, relief=tk.RIDGE).pack(side=tk.LEFT)
        tk.Label(header_frame, text="新しい梱包可能数", width=15, relief=tk.RIDGE).pack(side=tk.LEFT)
        
        # 入力エントリの辞書
        self.entries = {}
        
        # 各アイテムの入力行を作成
        for _, item in self.items_data.iterrows():
            item_frame = tk.Frame(scrollable_frame)
            item_frame.pack(fill=tk.X, pady=1)
            
            i_no = item['明細_共通項目2']
            order_number = item['受注番号']
            
            # 過去の梱包依頼数を取得
            previous_quantity = get_previous_packing_quantities(order_number, i_no)
            
            # デフォルト値を計算（明細_受注数量 - 過去の梱包依頼数）
            order_quantity = item['明細_受注数量']
            default_value = max(0, order_quantity - previous_quantity)
            
            tk.Label(item_frame, text=str(i_no), width=8, relief=tk.RIDGE).pack(side=tk.LEFT)
            tk.Label(item_frame, text=item['明細_商品コード'], width=15, relief=tk.RIDGE).pack(side=tk.LEFT)
            tk.Label(item_frame, text=item['明細_商品受注名'][:30], width=30, relief=tk.RIDGE).pack(side=tk.LEFT)
            tk.Label(item_frame, text=str(order_quantity), width=10, relief=tk.RIDGE).pack(side=tk.LEFT)
            tk.Label(item_frame, text=str(item['梱包可能数']), width=15, relief=tk.RIDGE).pack(side=tk.LEFT)
            
            # 入力エントリ
            entry = tk.Entry(item_frame, width=15)
            entry.pack(side=tk.LEFT, padx=5)
            entry.insert(0, str(default_value))
            
            self.entries[i_no] = entry
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # ボタンフレーム
        button_frame = tk.Frame(self.dialog)
        button_frame.pack(pady=10)
        
        ok_button = tk.Button(button_frame, text="OK", command=self.on_ok)
        ok_button.pack(side=tk.LEFT, padx=10)
        
        cancel_button = tk.Button(button_frame, text="キャンセル", command=self.on_cancel)
        cancel_button.pack(side=tk.LEFT, padx=10)
        
        # ダイアログを中央に配置
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
        
        # PackingRowDialog.__init__ の末尾あたりに追記
        self.note_text.bind("<KeyRelease-Return>", self._normalize_note_text)
        self.note_text.bind("<FocusOut>", self._normalize_note_text)

        def _normalize_note_text(self, event=None):
            # 全角ハイフン/長音/マイナスを半角に正規化
            s = self.note_text.get("1.0", "end-1c")
            table = str.maketrans({
                "－": "-",  # U+FF0D FULLWIDTH HYPHEN-MINUS
                "―": "-",  # U+2015 HORIZONTAL BAR
                "ー": "-",  # U+30FC KATAKANA-HIRAGANA PROLONGED SOUND MARK
                "−": "-",  # U+2212 MINUS SIGN
            })
            t = s.translate(table)
            if s != t:
                cur = self.note_text.index(tk.INSERT)
                self.note_text.delete("1.0", "end")
                self.note_text.insert("1.0", t)
                self.note_text.mark_set(tk.INSERT, cur)

    
    def on_ok(self):
        """OKボタンが押された時の処理"""
        try:
            result = {}
            for i_no, entry in self.entries.items():
                value = entry.get().strip()
                if value:
                    result[i_no] = int(value)
                else:
                    result[i_no] = 0
            
            self.result = result
            self.dialog.destroy()
            
        except ValueError:
            messagebox.showerror("エラー", "数値を正しく入力してください。")
    
    def on_cancel(self):
        """キャンセルボタンが押された時の処理"""
        self.result = None
        self.dialog.destroy()

def parse_item_details_with_quantities(item_details_str):
    """
    アイテム明細から数値と数量情報を抽出する
    例: "1～4、5（7/10）、6～20" → {1: (1,1), 2: (1,1), 3: (1,1), 4: (1,1), 5: (7,10), 6: (1,1), ...}
    """
    import re
    item_quantities = {}
    
    if not item_details_str:
        return item_quantities
    
    # 「、」や「,」で分割
    parts = re.split(r'[、,]', str(item_details_str))
    
    for part in parts:
        part = part.strip()
        
        # ★修正：複数の数量付きパターンに対応★
        # 日本語括弧パターン「5（7/10）」
        quantity_match = re.search(r'(\d+)（(\d+)/(\d+)）', part)
        if not quantity_match:
            # 英語括弧パターン「5(7/10)」
            quantity_match = re.search(r'(\d+)\((\d+)/(\d+)\)', part)
        if not quantity_match:
            # コロン付き括弧パターン「5:(7/10)」
            quantity_match = re.search(r'(\d+):\((\d+)/(\d+)\)', part)
        
        if quantity_match:
            item_num = int(quantity_match.group(1))
            packed_qty = int(quantity_match.group(2))
            total_qty = int(quantity_match.group(3))
            item_quantities[item_num] = (packed_qty, total_qty)
            continue
        
        # 範囲パターン「1～4」「1-4」「1〜4」
        range_match = re.search(r'(\d+)[～\-〜]+(\d+)', part)
        if range_match:
            start = int(range_match.group(1))
            end = int(range_match.group(2))
            for num in range(start, end + 1):
                item_quantities[num] = (1, 1)  # 完全梱包として扱う
            continue
        
        # 単一数値パターン「5」
        single_match = re.search(r'(\d+)', part)
        if single_match:
            item_num = int(single_match.group(1))
            item_quantities[item_num] = (1, 1)  # 完全梱包として扱う
    
    return item_quantities

def format_consolidated_range(item_quantities):
    """
    統合された数量情報を文字列に変換
    例: {1:(1,1), 2:(1,1), 3:(1,1), 4:(1,1), 5:(10,10), 6:(1,1), ...} → "1～20"
    """
    if not item_quantities:
        return ""
    
    # 完全梱包されたアイテムのみを抽出
    complete_items = []
    partial_items = []
    
    for item_num, (packed_qty, total_qty) in sorted(item_quantities.items()):
        if packed_qty >= total_qty:
            complete_items.append(item_num)
        else:
            partial_items.append(f"{item_num}（{packed_qty}/{total_qty}）")
    
    # 完全梱包アイテムを範囲表示に変換
    complete_ranges = []
    if complete_items:
        complete_ranges.append(format_number_range(complete_items))
    
    # 結果を結合
    result_parts = []
    if complete_ranges and complete_ranges[0]:
        result_parts.extend(complete_ranges)
    if partial_items:
        result_parts.extend(partial_items)
    
    return "、".join(result_parts)

def format_number_range(sorted_numbers):
    """
    ソートされた数値リストを「1～20」形式の文字列に変換
    """
    if not sorted_numbers:
        return ""
    
    if len(sorted_numbers) == 1:
        return str(sorted_numbers[0])
    
    ranges = []
    start = sorted_numbers[0]
    end = sorted_numbers[0]
    
    for i in range(1, len(sorted_numbers)):
        if sorted_numbers[i] == end + 1:
            end = sorted_numbers[i]
        else:
            if start == end:
                ranges.append(str(start))
            else:
                ranges.append(f"{start}～{end}")
            start = sorted_numbers[i]
            end = sorted_numbers[i]
    
    # 最後の範囲を追加
    if start == end:
        ranges.append(str(start))
    else:
        ranges.append(f"{start}～{end}")
    
    return "、".join(ranges)

class PackingConsolidationDialog:
    """
    梱包明細統合ダイアログクラス
    """
    def __init__(self, parent, order_number):
        self.parent = parent
        self.order_number = order_number
        self.result = None
        self.selected_shipments = []
        self.consolidation_rules = []  # [(target_case, source_shipment, source_case), ...]
        
        # ダイアログウィンドウを作成
        self.dialog = Toplevel(parent)
        self.dialog.title("梱包明細統合設定")
        self.dialog.geometry("1400x800")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # メインフレーム
        self.main_frame = tk.Frame(self.dialog)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.create_widgets()
        self.load_shipment_data()
        
        # ダイアログを中央に配置
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
        
        # ★追加：初期状態でも処理モードの状態を設定★
        self.on_mode_changed()
    
    def create_widgets(self):
        """ウィジェットを作成"""
        # タイトル
        title_label = tk.Label(self.main_frame, text=f"梱包明細統合設定 - 受注番号: {self.order_number}", 
                            font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 20))
        
        # メインコンテンツエリア
        content_frame = tk.Frame(self.main_frame)
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # 左側：出荷依頼番号選択
        left_frame = tk.Frame(content_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        # 出荷依頼番号選択セクション
        shipment_label = tk.Label(left_frame, text="1. 統合対象の出荷依頼番号を選択", 
                                font=("Arial", 12, "bold"))
        shipment_label.pack(anchor='w', pady=(0, 10))
        
        # 出荷依頼番号リスト（チェックボックス付き）
        shipment_list_frame = tk.Frame(left_frame)
        shipment_list_frame.pack(fill=tk.BOTH, expand=True)
        
        # スクロール可能なフレーム
        shipment_canvas = tk.Canvas(shipment_list_frame, height=200)
        shipment_scrollbar = tk.Scrollbar(shipment_list_frame, orient="vertical", command=shipment_canvas.yview)
        self.shipment_scrollable_frame = tk.Frame(shipment_canvas)
        
        self.shipment_scrollable_frame.bind(
            "<Configure>",
            lambda e: shipment_canvas.configure(scrollregion=shipment_canvas.bbox("all"))
        )
        
        shipment_canvas.create_window((0, 0), window=self.shipment_scrollable_frame, anchor="nw")
        shipment_canvas.configure(yscrollcommand=shipment_scrollbar.set)
        
        shipment_canvas.pack(side="left", fill="both", expand=True)
        shipment_scrollbar.pack(side="right", fill="y")
        
        # 統合設定セクション
        consolidation_label = tk.Label(left_frame, text="2. 統合設定", 
                                    font=("Arial", 12, "bold"))
        consolidation_label.pack(anchor='w', pady=(20, 10))
        
        # ★★★ 結合/統合モード選択を追加 ★★★
        mode_frame = tk.Frame(left_frame)
        mode_frame.pack(fill=tk.X, pady=(0, 10))

        self.consolidation_mode = StringVar(value="combine")
        tk.Label(mode_frame, text="処理モード:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=(0, 10))

        combine_radio = tk.Radiobutton(mode_frame, text="結合（ケース番号そのまま）", 
                                    variable=self.consolidation_mode, value="combine",
                                    command=self.on_mode_changed)
        combine_radio.pack(side=tk.LEFT, padx=(0, 15))

        merge_radio = tk.Radiobutton(mode_frame, text="統合（ケース番号を統合）", 
                                    variable=self.consolidation_mode, value="merge",
                                    command=self.on_mode_changed)
        merge_radio.pack(side=tk.LEFT)

        # ★修正：統合ルール設定フレームを`self.rule_settings_frame`として作成★
        self.rule_settings_frame = tk.Frame(left_frame)
        self.rule_settings_frame.pack(fill=tk.X, pady=(10, 0))
        
        # 統合ルール追加フレーム（上段）
        add_rule_frame1 = tk.Frame(self.rule_settings_frame)
        add_rule_frame1.pack(fill=tk.X, pady=(0, 5))

        tk.Label(add_rule_frame1, text="統合先出荷依頼:").pack(side=tk.LEFT, padx=(0, 5))
        self.target_shipment_var = StringVar()
        self.target_shipment_combo = ttk.Combobox(add_rule_frame1, textvariable=self.target_shipment_var, 
                                                width=15, state="readonly")
        self.target_shipment_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.target_shipment_combo.bind("<<ComboboxSelected>>", self.on_target_shipment_changed)

        tk.Label(add_rule_frame1, text="統合先ケース番号:").pack(side=tk.LEFT, padx=(0, 5))
        self.target_case_var = StringVar()
        self.target_case_combo = ttk.Combobox(add_rule_frame1, textvariable=self.target_case_var, 
                                            width=15, state="readonly")
        self.target_case_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.target_case_combo.bind("<<ComboboxSelected>>", self.on_target_case_changed)

        # 統合ルール追加フレーム（下段）
        add_rule_frame2 = tk.Frame(self.rule_settings_frame)
        add_rule_frame2.pack(fill=tk.X, pady=(0, 10))

        tk.Label(add_rule_frame2, text="統合元出荷依頼:").pack(side=tk.LEFT, padx=(0, 5))
        self.source_shipment_var = StringVar()
        self.source_shipment_combo = ttk.Combobox(add_rule_frame2, textvariable=self.source_shipment_var, 
                                                width=15, state="readonly")
        self.source_shipment_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.source_shipment_combo.bind("<<ComboboxSelected>>", self.on_source_shipment_changed)

        tk.Label(add_rule_frame2, text="統合元ケース番号:").pack(side=tk.LEFT, padx=(0, 5))
        self.source_case_var = StringVar()
        self.source_case_combo = ttk.Combobox(add_rule_frame2, textvariable=self.source_case_var, 
                                            width=15)
        self.source_case_combo.pack(side=tk.LEFT, padx=(0, 10))

        self.add_rule_button = tk.Button(add_rule_frame2, text="統合ルール追加", command=self.add_consolidation_rule)
        self.add_rule_button.pack(side=tk.LEFT, padx=(10, 0))
        
        # 統合ルール一覧
        rules_label = tk.Label(left_frame, text="設定済み統合ルール:", font=("Arial", 10, "bold"))
        rules_label.pack(anchor='w', pady=(10, 5))
        
        # 統合ルールリスト
        rules_list_frame = tk.Frame(left_frame)
        rules_list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.rules_listbox = tk.Listbox(rules_list_frame, height=6)
        rules_scrollbar = tk.Scrollbar(rules_list_frame, orient="vertical", command=self.rules_listbox.yview)
        self.rules_listbox.configure(yscrollcommand=rules_scrollbar.set)
        
        self.rules_listbox.pack(side="left", fill="both", expand=True)
        rules_scrollbar.pack(side="right", fill="y")
        
        # 統合ルール削除ボタン
        delete_rule_button = tk.Button(left_frame, text="選択したルールを削除", command=self.delete_consolidation_rule)
        delete_rule_button.pack(pady=(5, 0))
        
        # 右側：プレビュー表示
        right_frame = tk.Frame(content_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        preview_label = tk.Label(right_frame, text="3. ケース明細プレビュー", 
                                font=("Arial", 12, "bold"))
        preview_label.pack(anchor='w', pady=(0, 10))
        
        # プレビュー表示エリア
        preview_frame = tk.Frame(right_frame)
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        # プレビュー用のスクロール可能フレーム
        preview_canvas = tk.Canvas(preview_frame)
        preview_scrollbar = tk.Scrollbar(preview_frame, orient="vertical", command=preview_canvas.yview)
        self.preview_scrollable_frame = tk.Frame(preview_canvas)
        
        self.preview_scrollable_frame.bind(
            "<Configure>",
            lambda e: preview_canvas.configure(scrollregion=preview_canvas.bbox("all"))
        )
        
        preview_canvas.create_window((0, 0), window=self.preview_scrollable_frame, anchor="nw")
        preview_canvas.configure(yscrollcommand=preview_scrollbar.set)
        
        preview_canvas.pack(side="left", fill="both", expand=True)
        preview_scrollbar.pack(side="right", fill="y")
        
        # ボタンフレーム
        button_frame = tk.Frame(self.main_frame)
        button_frame.pack(pady=20)
        
        preview_result_button = tk.Button(button_frame, text="統合結果プレビュー", command=self.preview_consolidation_result)
        preview_result_button.pack(side=tk.LEFT, padx=10)
        
        ok_button = tk.Button(button_frame, text="OK", command=self.on_ok)
        ok_button.pack(side=tk.LEFT, padx=10)
        
        cancel_button = tk.Button(button_frame, text="キャンセル", command=self.on_cancel)
        cancel_button.pack(side=tk.LEFT, padx=10)
    
    def load_shipment_data(self):
        """出荷依頼番号データを読み込み"""
        try:
            # 受注番号に対応するshipment_numberを取得
            self.shipment_numbers = get_shipment_numbers_by_order(self.order_number)
            self.shipment_data = {}
            
            # 各shipment_numberの梱包明細を取得
            for shipment_number in self.shipment_numbers:
                packing_data = get_packing_details_by_shipment(shipment_number)
                self.shipment_data[shipment_number] = packing_data
            
            # 出荷依頼番号チェックボックスを作成
            self.shipment_vars = {}
            for shipment_number in self.shipment_numbers:
                var = tk.IntVar()
                checkbox = tk.Checkbutton(
                    self.shipment_scrollable_frame, 
                    text=f"{shipment_number} ({len(self.shipment_data[shipment_number])}ケース)",
                    variable=var,
                    command=self.on_shipment_selection_changed
                )
                checkbox.pack(anchor='w', pady=2)
                self.shipment_vars[shipment_number] = var
            
        except Exception as e:
            messagebox.showerror("エラー", f"データ読み込みエラー: {str(e)}")
    
    def on_shipment_selection_changed(self):
        """出荷依頼番号選択変更時の処理"""
        # 選択された出荷依頼番号を更新
        self.selected_shipments = [
            shipment for shipment, var in self.shipment_vars.items() 
            if var.get()
        ]
        
        # プルダウンの選択肢を更新
        self.update_combo_options()

    def on_mode_changed(self):
        """処理モード変更時の処理"""
        mode = self.consolidation_mode.get()
        
        # 既存の説明ラベルがあれば削除
        if hasattr(self, 'mode_info_label'):
            self.mode_info_label.destroy()
        
        if mode == "combine":
            # 結合モード：統合ルール設定を無効化
            self.target_shipment_combo.config(state='disabled')
            self.target_case_combo.config(state='disabled')
            self.source_shipment_combo.config(state='disabled')
            self.source_case_combo.config(state='disabled')
            self.add_rule_button.config(state='disabled')
            
            # ★修正：結合モードの説明を表示★
            self.mode_info_label = tk.Label(self.rule_settings_frame, 
                                        text="結合モード：選択した出荷依頼の梱包明細をそのまま結合します",
                                        fg="blue", font=("Arial", 9))
            self.mode_info_label.pack(anchor='w', pady=5)
            
        else:
            # 統合モード：統合ルール設定を有効化
            self.target_shipment_combo.config(state='readonly')
            self.target_case_combo.config(state='readonly')
            self.source_shipment_combo.config(state='readonly')
            self.source_case_combo.config(state='normal')  # 入力可能
            self.add_rule_button.config(state='normal')
            
            # ★修正：統合モードの説明を詳細化★
            self.mode_info_label = tk.Label(self.rule_settings_frame, 
                                        text="統合モード：指定したケース番号のアイテムを統合して新しい梱包明細を作成します\n※重量は統合先の重量に統一されます",
                                        fg="green", font=("Arial", 9), justify=tk.LEFT)
            self.mode_info_label.pack(anchor='w', pady=5)

    def update_combo_options(self):
        """プルダウンの選択肢を更新"""
        # 統合先出荷依頼番号の選択肢
        self.target_shipment_combo['values'] = self.selected_shipments
        
        # 統合元出荷依頼番号の選択肢
        self.source_shipment_combo['values'] = self.selected_shipments
        
        # 統合先ケース番号の選択肢は統合先出荷依頼選択後に更新される
    
    def natural_sort_key(self, text):
        """自然順序ソート用のキー関数"""
        import re
        return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', text)]

    def on_target_shipment_changed(self, event=None):
        """統合先出荷依頼番号変更時の処理"""
        target_shipment = self.target_shipment_var.get()
        if target_shipment and target_shipment in self.shipment_data:
            # 統合先ケース番号の選択肢を更新
            cases = [item['case_number'] for item in self.shipment_data[target_shipment]]
            sorted_cases = sorted(set(cases), key=lambda x: self.natural_sort_key(str(x)))
            self.target_case_combo['values'] = sorted_cases
            
            # 統合先ケース番号をクリア
            self.target_case_var.set("")
            
            # プレビューを更新
            self.update_preview()
    
    def on_target_case_changed(self, event=None):
        """統合先ケース番号変更時の処理"""
        target_case = self.target_case_var.get()
        if target_case:
            # 統合元ケース番号も同じ値に自動設定
            self.source_case_var.set(target_case)
            
            # プレビューを更新
            self.update_preview()
    
    def on_source_shipment_changed(self, event=None):
        """統合元出荷依頼番号変更時の処理"""
        source_shipment = self.source_shipment_var.get()
        if source_shipment and source_shipment in self.shipment_data:
            # 統合元ケース番号の選択肢を更新
            cases = [item['case_number'] for item in self.shipment_data[source_shipment]]
            sorted_cases = sorted(set(cases), key=lambda x: self.natural_sort_key(str(x)))
            self.source_case_combo['values'] = sorted_cases
            
            # プレビューを更新
            self.update_preview()
    
    def update_preview(self):
        """プレビュー表示を更新"""
        # プレビューエリアをクリア
        for widget in self.preview_scrollable_frame.winfo_children():
            widget.destroy()
        
        # 統合先の明細を表示
        target_shipment = self.target_shipment_var.get()
        target_case = self.target_case_var.get()
        source_shipment = self.source_shipment_var.get()
        source_case = self.source_case_var.get()
        
        if target_shipment and target_case:
            # 統合先の明細表示
            target_label = tk.Label(self.preview_scrollable_frame, 
                                text=f"統合先: {target_shipment} - ケース番号: {target_case}", 
                                font=("Arial", 11, "bold"), fg="blue")
            target_label.pack(anchor='w', pady=(0, 10))
            
            # 統合先の明細を表示
            if target_shipment in self.shipment_data:
                matching_items = [item for item in self.shipment_data[target_shipment] 
                                if item['case_number'] == target_case]
                
                if matching_items:
                    target_frame = tk.Frame(self.preview_scrollable_frame, relief=tk.GROOVE, borderwidth=1)
                    target_frame.pack(fill=tk.X, pady=2, padx=5)
                    
                    for item in matching_items:
                        detail_text = f"    {item['length']}×{item['width']}×{item['height']}cm, {item['weight']}kg"
                        detail_text += f"\n    明細: {item['item_details']}"
                        detail_text += f"\n    荷姿: {item['packing_style']}"
                        
                        detail_label = tk.Label(target_frame, text=detail_text, 
                                            font=("Arial", 9), justify=tk.LEFT)
                        detail_label.pack(anchor='w', padx=10, pady=2)
        
        if source_shipment and source_case:
            # 統合元の明細表示
            separator = tk.Label(self.preview_scrollable_frame, text="─" * 50, fg="gray")
            separator.pack(pady=10)
            
            source_label = tk.Label(self.preview_scrollable_frame, 
                                text=f"統合元: {source_shipment} - ケース番号: {source_case}", 
                                font=("Arial", 11, "bold"), fg="red")
            source_label.pack(anchor='w', pady=(0, 10))
            
            if source_shipment in self.shipment_data:
                matching_items = [item for item in self.shipment_data[source_shipment] 
                                if item['case_number'] == source_case]
                
                if matching_items:
                    source_frame = tk.Frame(self.preview_scrollable_frame, relief=tk.GROOVE, borderwidth=1)
                    source_frame.pack(fill=tk.X, pady=2, padx=5)
                    
                    for item in matching_items:
                        detail_text = f"    {item['length']}×{item['width']}×{item['height']}cm, {item['weight']}kg"
                        detail_text += f"\n    明細: {item['item_details']}"
                        detail_text += f"\n    荷姿: {item['packing_style']}"
                        
                        detail_label = tk.Label(source_frame, text=detail_text, 
                                            font=("Arial", 9), justify=tk.LEFT, fg="red")
                        detail_label.pack(anchor='w', padx=10, pady=2)
    
    def add_consolidation_rule(self):
        """統合ルールを追加"""
        target_shipment = self.target_shipment_var.get()
        target_case = self.target_case_var.get()
        source_shipment = self.source_shipment_var.get()
        source_case = self.source_case_var.get()
        
        if not all([target_shipment, target_case, source_shipment, source_case]):
            messagebox.showwarning("警告", "統合先出荷依頼、統合先ケース番号、統合元出荷依頼、統合元ケース番号をすべて選択してください。")
            return
        
        # 重複チェック
        rule = (target_shipment, target_case, source_shipment, source_case)
        if rule in self.consolidation_rules:
            messagebox.showwarning("警告", "同じ統合ルールが既に存在します。")
            return
        
        # ルールを追加
        self.consolidation_rules.append(rule)
        
        # リストボックスに表示
        rule_text = f"{target_shipment}:{target_case} ← {source_shipment}:{source_case}"
        self.rules_listbox.insert(tk.END, rule_text)
        
        # 入力欄をクリア
        self.target_shipment_var.set("")
        self.target_case_var.set("")
        self.source_shipment_var.set("")
        self.source_case_var.set("")
        
        messagebox.showinfo("情報", "統合ルールを追加しました。")
    
    def delete_consolidation_rule(self):
        """選択された統合ルールを削除"""
        selected_indices = self.rules_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("警告", "削除するルールを選択してください。")
            return
        
        # 後ろから削除（インデックスずれを防ぐため）
        for index in reversed(selected_indices):
            del self.consolidation_rules[index]
            self.rules_listbox.delete(index)
        
        messagebox.showinfo("情報", "選択したルールを削除しました。")
    
    def preview_consolidation_result(self):
        """統合結果をプレビュー表示"""
        mode = self.consolidation_mode.get()
        
        if mode == "merge" and not self.consolidation_rules:
            messagebox.showwarning("警告", "統合ルールが設定されていません。")
            return
        
        # ★追加：警告チェック★
        if mode == "merge" and self.consolidation_rules:
            warnings = self.validate_consolidation_rules()
            if warnings:
                if not self.show_consolidation_warnings(warnings):
                    return  # ユーザーが「いいえ」を選択した場合は処理中止
        
        try:
            # 統合処理を実行
            if mode == "merge":
                consolidated_data = self.perform_consolidation()
            else:
                consolidated_data = self.perform_combination()
            
            # 結果表示ダイアログを作成
            self.show_consolidation_result_dialog(consolidated_data)
            
        except Exception as e:
            messagebox.showerror("エラー", f"統合処理中にエラーが発生しました: {str(e)}")
    
    def perform_consolidation(self):
        """統合処理を実行"""
        # 選択された出荷依頼番号のみのデータを取得
        all_packing_data = []
        selected_shipments_only = self.selected_shipments  # 選択された出荷依頼番号のみ
        
        for shipment in selected_shipments_only:
            if shipment in self.shipment_data:
                for item in self.shipment_data[shipment]:
                    item_copy = item.copy()
                    item_copy['shipment_number'] = shipment
                    all_packing_data.append(item_copy)
        
        # 統合ルールを適用
        consolidated_data = self.apply_consolidation_rules(all_packing_data)
        
        # ★修正：統合結果に選択された全出荷依頼番号を記録★
        if consolidated_data:
            selected_shipments_str = ', '.join(sorted(selected_shipments_only))
            for item in consolidated_data:
                item['source_shipments'] = selected_shipments_str
        
        return consolidated_data
    
    def perform_combination(self):
        """結合処理を実行（ケース番号重複対応）"""
        combined_data = []
        case_counter = {}  # ケース番号の重複カウント
        selected_shipments_only = self.selected_shipments  # 選択された出荷依頼番号のみ
        
        for shipment in selected_shipments_only:
            if shipment in self.shipment_data:
                for item in self.shipment_data[shipment]:
                    item_copy = item.copy()
                    
                    # ケース番号の重複チェック
                    original_case = item_copy['case_number']
                    
                    # 既に同じケース番号が存在する場合は番号を調整
                    if original_case in case_counter:
                        case_counter[original_case] += 1
                        # 重複時は出荷依頼番号をプレフィックスとして追加
                        adjusted_case = f"{shipment}-{original_case}"
                        item_copy['case_number'] = adjusted_case
                        item_copy['original_case_number'] = original_case
                    else:
                        case_counter[original_case] = 1
                        item_copy['original_case_number'] = original_case
                    
                    # ★修正：選択された全出荷依頼番号を記録★
                    selected_shipments_str = ', '.join(sorted(selected_shipments_only))
                    item_copy['source_shipments'] = selected_shipments_str
                    combined_data.append(item_copy)
        
        return combined_data

    def apply_consolidation_rules(self, all_packing_data):
        """統合ルールを適用して梱包明細を統合"""
        case_consolidation = {}
        processed_items = set()  # 処理済みアイテムを追跡
        
        # 統合ルールを適用
        for target_shipment, target_case, source_shipment, source_case in self.consolidation_rules:
            consolidation_key = f"{target_shipment}_{target_case}"
            
            if consolidation_key not in case_consolidation:
                case_consolidation[consolidation_key] = {
                    'case_number': target_case,
                    'length': 0,
                    'width': 0,
                    'height': 0,
                    'weight': 0,
                    'target_weight_set': False,
                    'packing_style': "",
                    'item_quantities': {},
                    'source_shipments': set()
                }
            
            # 統合対象のアイテムを検索（選択された出荷依頼番号のみ）
            for item in all_packing_data:
                # 選択された出荷依頼番号のみを処理
                if item['shipment_number'] not in self.selected_shipments:
                    continue
                    
                item_id = f"{item['shipment_number']}_{item['case_number']}"
                
                # 統合先または統合元に該当するアイテムを処理
                is_target = (item['case_number'] == target_case and 
                        item['shipment_number'] == target_shipment)
                is_source = (item['case_number'] == source_case and 
                        item['shipment_number'] == source_shipment)
                
                if (is_target or is_source) and item_id not in processed_items:
                    # 既存の処理を継続...
                    case_data = case_consolidation[consolidation_key]
                    case_data['length'] = max(case_data['length'], float(item['length'] or 0))
                    case_data['width'] = max(case_data['width'], float(item['width'] or 0))
                    case_data['height'] = max(case_data['height'], float(item['height'] or 0))
                    
                    # 重量は統合先の重量に統一
                    if is_target and not case_data['target_weight_set']:
                        case_data['weight'] = float(item['weight'] or 0)
                        case_data['target_weight_set'] = True
                    elif not case_data['target_weight_set']:
                        case_data['weight'] = float(item['weight'] or 0)
                    
                    if not case_data['packing_style']:
                        case_data['packing_style'] = item['packing_style']
                    
                    case_data['source_shipments'].add(item['shipment_number'])
                    
                    # アイテム明細から数量情報を抽出して統合
                    item_quantities = parse_item_details_with_quantities(item['item_details'])
                    
                    for item_num, (packed_qty, total_qty) in item_quantities.items():
                        if item_num in case_data['item_quantities']:
                            existing_packed, existing_total = case_data['item_quantities'][item_num]
                            new_packed = existing_packed + packed_qty
                            new_total = max(existing_total, total_qty)
                            case_data['item_quantities'][item_num] = (new_packed, new_total)
                        else:
                            case_data['item_quantities'][item_num] = (packed_qty, total_qty)
                    
                    processed_items.add(item_id)
        
        # 統合されなかったアイテムも追加（選択された出荷依頼番号のみ）
        for item in all_packing_data:
            if item['shipment_number'] not in self.selected_shipments:
                continue
                
            item_id = f"{item['shipment_number']}_{item['case_number']}"
            if item_id not in processed_items:
                consolidation_key = f"{item['shipment_number']}_{item['case_number']}"
                
                case_consolidation[consolidation_key] = {
                    'case_number': item['case_number'],
                    'length': float(item['length'] or 0),
                    'width': float(item['width'] or 0),
                    'height': float(item['height'] or 0),
                    'weight': float(item['weight'] or 0),
                    'packing_style': item['packing_style'],
                    'item_quantities': parse_item_details_with_quantities(item['item_details']),
                    'source_shipments': {item['shipment_number']}
                }
        
        # 統合結果を整理
        consolidated_data = []
        for consolidation_key, data in case_consolidation.items():
            if 'target_weight_set' in data:
                del data['target_weight_set']
                
            consolidated_range = format_consolidated_range(data['item_quantities'])
            
            # ★修正：選択された全出荷依頼番号を記録★
            selected_source_shipments = [s for s in data['source_shipments'] if s in self.selected_shipments]
            
            consolidated_item = {
                'case_number': data['case_number'],
                'length': str(data['length']),
                'width': str(data['width']),
                'height': str(data['height']),
                'weight': str(round(data['weight'], 2)),
                'item_details': consolidated_range,
                'packing_style': data['packing_style'],
                'source_shipments': ', '.join(sorted(selected_source_shipments))
            }
            
            consolidated_data.append(consolidated_item)
            print(f"統合アイテム作成: {consolidated_item}")  # デバッグ

        print(f"統合結果: {len(consolidated_data)}件")  # デバッグ
        return consolidated_data
    
    def show_consolidation_result_dialog(self, consolidated_data):
        """統合結果表示ダイアログ"""
        result_dialog = Toplevel(self.dialog)
        result_dialog.title("統合結果プレビュー")
        result_dialog.geometry("1000x600")
        result_dialog.transient(self.dialog)
        result_dialog.grab_set()
        
        # メインフレーム
        main_frame = tk.Frame(result_dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # タイトル
        title_label = tk.Label(main_frame, text="梱包明細統合結果", font=("Arial", 14, "bold"))
        title_label.pack(pady=(0, 10))
        
        # スクロール可能なフレーム
        canvas = tk.Canvas(main_frame)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # ヘッダー
        headers = ["ケース番号", "縦(cm)", "横(cm)", "高(cm)", "重量(kg)", "アイテム明細", "荷姿", "元出荷依頼"]
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        col_widths = [12, 8, 8, 8, 10, 25, 15, 20]
        for i, (header, width) in enumerate(zip(headers, col_widths)):
            tk.Label(header_frame, text=header, width=width, relief=tk.RIDGE, 
                    font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        
        # データ行
        for data in consolidated_data:
            data_frame = tk.Frame(scrollable_frame)
            data_frame.pack(fill=tk.X, pady=1)
            
            values = [
                data['case_number'],
                data['length'],
                data['width'],
                data['height'],
                data['weight'],
                data['item_details'],
                data['packing_style'],
                data['source_shipments']
            ]
            
            for value, width in zip(values, col_widths):
                tk.Label(data_frame, text=str(value), width=width, relief=tk.RIDGE, 
                        font=("Arial", 9)).pack(side=tk.LEFT)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # 閉じるボタン
        close_button = tk.Button(result_dialog, text="閉じる", command=result_dialog.destroy)
        close_button.pack(pady=10)
    
    def on_ok(self):
        """OKボタンが押された時の処理"""
        mode = self.consolidation_mode.get()
        
        if mode == "merge" and not self.consolidation_rules:
            if not messagebox.askyesno("確認", "統合ルールが設定されていません。統合なしで続行しますか？"):
                return
        
        # ★追加：警告チェック★
        if mode == "merge" and self.consolidation_rules:
            warnings = self.validate_consolidation_rules()
            if warnings:
                if not self.show_consolidation_warnings(warnings):
                    return  # ユーザーが「いいえ」を選択した場合は処理中止
        
        try:
            # 統合処理を実行
            if mode == "merge" and self.consolidation_rules:
                self.result = self.perform_consolidation()
                print(f"統合処理完了: {len(self.result)}件のケース")  # デバッグ
            else:
                # 結合または統合なしの場合
                self.result = self.perform_combination()
                print(f"結合処理完了: {len(self.result)}件のケース")  # デバッグ
            
            # デバッグ: 結果の内容を表示
            for i, item in enumerate(self.result):
                print(f"結果[{i}]: ケース={item.get('case_number')}, 明細={item.get('item_details')}")
            
            self.dialog.destroy()
            
        except Exception as e:
            print(f"統合処理エラー: {str(e)}")  # デバッグ
            messagebox.showerror("エラー", f"統合処理中にエラーが発生しました: {str(e)}")
    
    def on_cancel(self):
        """キャンセルボタンが押された時の処理"""
        self.result = None
        self.dialog.destroy()

    def validate_consolidation_rules(self):
        """統合ルールの安全性をチェックし、必要に応じて警告を表示"""
        warnings = []
        
        for target_shipment, target_case, source_shipment, source_case in self.consolidation_rules:
            # 統合対象のアイテムを取得
            target_items = [item for item in self.shipment_data.get(target_shipment, []) 
                        if item['case_number'] == target_case]
            source_items = [item for item in self.shipment_data.get(source_shipment, []) 
                        if item['case_number'] == source_case]
            
            if not target_items or not source_items:
                continue
                
            target_item = target_items[0]
            source_item = source_items[0]
            
            # ★デバッグ用コードを追加★
            print(f"=== デバッグ情報 ===")
            print(f"統合先: {target_shipment}:{target_case}")
            print(f"統合先明細文字列: '{target_item['item_details']}'")
            print(f"統合元: {source_shipment}:{source_case}")
            print(f"統合元明細文字列: '{source_item['item_details']}'")
            
            # 3. 部分梱包の数量合算チェック
            target_quantities = parse_item_details_with_quantities(target_item['item_details'])
            source_quantities = parse_item_details_with_quantities(source_item['item_details'])
            
            # ★デバッグ用：解析結果を表示★
            print(f"統合先解析結果: {target_quantities}")
            print(f"統合元解析結果: {source_quantities}")
            
            for item_num in set(target_quantities.keys()) | set(source_quantities.keys()):
                target_packed, target_total = target_quantities.get(item_num, (0, 0))
                source_packed, source_total = source_quantities.get(item_num, (0, 0))
                
                # ★デバッグ用：各アイテムの計算過程を表示★
                print(f"アイテム{item_num}: 統合先({target_packed}/{target_total}) + 統合元({source_packed}/{source_total})")
                
                if target_total > 0 and source_total > 0:
                    # 両方に部分梱包がある場合の合算チェック
                    combined_packed = target_packed + source_packed
                    max_total = max(target_total, source_total)
                    
                    print(f"  → 合算: {combined_packed}, 母数: {max_total}")
                    
                    if combined_packed > max_total:
                        print(f"  → 警告発生: {combined_packed} > {max_total}")
                        warnings.append(f"アイテム番号{item_num}の梱包数量合算（{combined_packed}）が母数（{max_total}）を超えています。")
            
            print(f"==================")
        
        return warnings
    
    def show_consolidation_warnings(self, warnings):
        """統合警告ダイアログを表示"""
        if not warnings:
            return True  # 警告がない場合は続行
        
        warning_dialog = Toplevel(self.dialog)
        warning_dialog.title("統合時の注意事項")
        warning_dialog.geometry("700x500")  # ★高さを400→500に、幅も600→700に拡大
        warning_dialog.minsize(600, 450)   # ★最小サイズを設定
        warning_dialog.transient(self.dialog)
        warning_dialog.grab_set()
        
        # メインフレーム
        main_frame = tk.Frame(warning_dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)  # ★パディングを調整
        
        # タイトル
        title_label = tk.Label(main_frame, text="以下の注意事項が検出されました", 
                              font=("Arial", 12, "bold"), fg="orange")
        title_label.pack(pady=(0, 15))
        
        # 警告リスト表示
        warning_frame = tk.Frame(main_frame)
        warning_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))  # ★下部マージンを調整
        
        # スクロール可能なテキスト
        text_widget = tk.Text(warning_frame, wrap=tk.WORD, font=("Arial", 10), height=12)  # ★高さを明示的に指定
        scrollbar = tk.Scrollbar(warning_frame, orient="vertical", command=text_widget.yview)
        text_widget.configure(yscrollcommand=scrollbar.set)
        
        for i, warning in enumerate(warnings, 1):
            text_widget.insert(tk.END, f"{i}. {warning}\n\n")
        
        text_widget.config(state=tk.DISABLED)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 確認メッセージ
        confirm_label = tk.Label(main_frame, text="このまま統合処理を続行しますか？", 
                                font=("Arial", 11, "bold"))
        confirm_label.pack(pady=(15, 10))  # ★上下のマージンを調整
        
        # 結果を格納する変数
        result = [False]
        
        def on_yes():
            result[0] = True
            warning_dialog.destroy()
        
        def on_no():
            result[0] = False
            warning_dialog.destroy()
        
        # ボタンフレーム
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=(10, 0))  # ★上部マージンを追加
        
        yes_button = tk.Button(button_frame, text="はい（続行）", command=on_yes, 
                            font=("Arial", 11), bg="lightgreen", padx=25, pady=8)  # ★サイズとフォントを調整
        yes_button.pack(side=tk.LEFT, padx=15)  # ★間隔を広げる
        
        no_button = tk.Button(button_frame, text="いいえ（中止）", command=on_no, 
                            font=("Arial", 11), bg="lightcoral", padx=25, pady=8)  # ★サイズとフォントを調整
        no_button.pack(side=tk.LEFT, padx=15)  # ★間隔を広げる
        
        # ダイアログを中央に配置
        warning_dialog.update_idletasks()
        x = (warning_dialog.winfo_screenwidth() // 2) - (warning_dialog.winfo_width() // 2)
        y = (warning_dialog.winfo_screenheight() // 2) - (warning_dialog.winfo_height() // 2)
        warning_dialog.geometry(f"+{x}+{y}")
        
        # モーダルダイアログとして待機
        warning_dialog.wait_window()
        
        return result[0]

class PackingRowDialog:
    """梱包依頼行の編集ダイアログ"""
    def __init__(self, parent, edit_data=None):
        self.parent = parent
        self.result = None
        self.edit_data = edit_data
        
        # ダイアログウィンドウを作成
        self.dialog = Toplevel(parent)
        self.dialog.title("梱包依頼行の編集" if edit_data else "梱包依頼行の追加")
        self.dialog.geometry("550x550")
        self.dialog.minsize(500, 500)
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # メインフレーム
        main_frame = tk.Frame(self.dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 検索方法 - ラジオボタンに変更
        search_method_frame = tk.Frame(main_frame)
        search_method_frame.pack(fill='x', pady=8)
        tk.Label(search_method_frame, text="検索方法:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        
        # ラジオボタン用の変数
        self.search_method_var = StringVar(value=edit_data.get('検索方法', '受注番号') if edit_data else '受注番号')
        
        order_number_radio = Radiobutton(search_method_frame, text="受注番号", 
                                        variable=self.search_method_var, value="受注番号",
                                        font=("Arial", 10))
        order_number_radio.pack(side=tk.LEFT, padx=15)
        
        estimate_no_radio = Radiobutton(search_method_frame, text="見積管理番号", 
                                       variable=self.search_method_var, value="見積管理番号",
                                       font=("Arial", 10))
        estimate_no_radio.pack(side=tk.LEFT, padx=15)
        
        # 番号
        number_frame = tk.Frame(main_frame)
        number_frame.pack(fill='x', pady=5)
        tk.Label(number_frame, text="番号:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        self.number_var = StringVar(value=edit_data.get('番号', '') if edit_data else '')
        number_entry = Entry(number_frame, textvariable=self.number_var, width=25, font=("Arial", 10))
        number_entry.pack(side=tk.LEFT, padx=5)
        
        # 梱包期限日
        deadline_frame = tk.Frame(main_frame)
        deadline_frame.pack(fill='x', pady=5)
        tk.Label(deadline_frame, text="梱包期限日:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        deadline_value = edit_data.get('梱包期限日', '') if edit_data else ''
        if deadline_value:
            try:
                deadline_date = datetime.strptime(deadline_value, '%Y/%m/%d').date()
            except:
                deadline_date = datetime.now().date()
        else:
            deadline_date = datetime.now().date()
        self.deadline_cal = DateEntry(deadline_frame, width=12, background='darkblue', 
                                     foreground='white', borderwidth=2, locale='ja_JP')
        self.deadline_cal.set_date(deadline_date)
        self.deadline_cal.pack(side=tk.LEFT, padx=5)
        
        # 梱包依頼摘要
        note_frame = tk.Frame(main_frame)
        note_frame.pack(fill='x', pady=8)
        tk.Label(note_frame, text="梱包依頼摘要:", font=("Arial", 10, "bold")).pack(anchor='w', padx=5)
        self.note_text = Text(note_frame, height=3, width=50, font=("Arial", 10))
        if edit_data and edit_data.get('梱包依頼摘要'):
            self.note_text.insert("1.0", edit_data['梱包依頼摘要'])
        self.note_text.pack(fill='x', padx=5, pady=5)
        
        # 梱包担当者
        person_frame = tk.Frame(main_frame)
        person_frame.pack(fill='x', pady=5)
        tk.Label(person_frame, text="梱包担当者:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        self.person_var = StringVar(value=edit_data.get('梱包担当者', '11_細田') if edit_data else '11_細田')
        person_combo = ttk.Combobox(person_frame, textvariable=self.person_var, 
                                   values=['11_細田', '12_平松', '13_坂上', '16_土田'], 
                                   state="readonly", width=15, font=("Arial", 10))
        person_combo.pack(side=tk.LEFT, padx=5)

        # 梱包明細依頼（メイン選択項目）
        packing_detail_frame = tk.Frame(main_frame)
        packing_detail_frame.pack(fill='x', pady=5)
        tk.Label(packing_detail_frame, text="梱包明細:", font=("Arial", 10, "bold")).pack(side=tk.LEFT, padx=5)
        self.packing_detail_var = IntVar(value=1 if edit_data and edit_data.get('梱包明細') == '梱包明細を依頼する案件' else 0)
        packing_detail_check = Checkbutton(
            packing_detail_frame,
            text="梱包明細を依頼する",
            variable=self.packing_detail_var,
            font=("Arial", 10)
        )
        packing_detail_check.pack(side=tk.LEFT, padx=5)
        self.packing_detail_check = packing_detail_check

        self.os_warning_label = tk.Label(
            main_frame,
            text="※海外案件（受注番号がOS始まり）は自動でONになるためチェック不要です",
            font=("Arial", 10, "bold"),
            fg="#d32f2f"
        )
        self.os_warning_label.pack(fill='x', pady=(0, 5), padx=5)
        
        # オプション設定
        options_frame = tk.LabelFrame(main_frame, text="オプション設定", padx=10, pady=10,
                                     font=("Arial", 10, "bold"))
        options_frame.pack(fill='x', pady=10)
        
        self.exclude_var = IntVar(value=1 if edit_data and edit_data.get('出力除外選択') == 'はい' else 0)
        exclude_check = Checkbutton(options_frame, text="出力するアイテムを選択する", 
                                   variable=self.exclude_var, font=("Arial", 10))
        exclude_check.pack(anchor='w', pady=3)
        
        self.zero_packing_var = IntVar(value=1 if not edit_data or edit_data.get('受注残0出力') == 'はい' else 0)
        zero_check = Checkbutton(options_frame, text="受注残が0、かつ梱包可能数が0のアイテムも出力する", 
                                variable=self.zero_packing_var, font=("Arial", 10))
        zero_check.pack(anchor='w', pady=3)
        
        self.change_quantity_var = IntVar(value=1 if edit_data and edit_data.get('梱包可能数変更') == 'はい' else 0)
        quantity_check = Checkbutton(options_frame, text="梱包可能数を変更する", 
                                    variable=self.change_quantity_var, font=("Arial", 10))
        quantity_check.pack(anchor='w', pady=3)
        
        # ボタンフレーム - 下部に固定配置
        button_frame = tk.Frame(main_frame)
        button_frame.pack(side=tk.BOTTOM, pady=15)
        
        ok_button = tk.Button(button_frame, text="OK", command=self.on_ok, 
                             font=("Arial", 11), padx=20, pady=5)
        ok_button.pack(side=tk.LEFT, padx=10)
        
        cancel_button = tk.Button(button_frame, text="キャンセル", command=self.on_cancel,
                                 font=("Arial", 11), padx=20, pady=5)
        cancel_button.pack(side=tk.LEFT, padx=10)

        self.number_var.trace_add("write", self._enforce_os_packing_detail)
        self.search_method_var.trace_add("write", self._enforce_os_packing_detail)
        self.number_var.trace_add("write", self._auto_select_packing_person)
        self.search_method_var.trace_add("write", self._auto_select_packing_person)
        self._enforce_os_packing_detail()
        self._auto_select_packing_person()
        
        # ダイアログを中央に配置
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")

    def _enforce_os_packing_detail(self, *args):
        """受注番号がOS始まりの場合は梱包明細依頼を強制ONにする。"""
        number_text = clean_input(self.number_var.get())
        is_os_case = self.search_method_var.get() == '受注番号' and str(number_text).upper().startswith("OS")
        if is_os_case:
            self.packing_detail_var.set(1)
            self.packing_detail_check.config(state='disabled')
        else:
            self.packing_detail_check.config(state='normal')

    def _auto_select_packing_person(self, *args):
        """番号入力時にorder_data.dbの共通項目2名から梱包担当者を自動選択します。"""
        try:
            input_value = self.number_var.get().strip()
            if not input_value:
                return

            common_person = query_common_person_name(input_value, self.search_method_var.get())
            if not common_person:
                return

            mapped_person = COMMON_PERSON_TO_PACKING_PERSON.get(normalize_person_name(common_person), '')
            if mapped_person:
                self.person_var.set(mapped_person)
        except Exception as e:
            print(f"梱包担当者の自動選択でエラー: {e}")
    
    def on_ok(self):
        """OKボタンが押された時の処理"""
        if not self.number_var.get().strip():
            messagebox.showerror("エラー", "番号を入力してください。")
            return

        is_os_case = self.search_method_var.get() == '受注番号' and str(clean_input(self.number_var.get())).upper().startswith("OS")
        if is_os_case:
            self.packing_detail_var.set(1)
        
        self.result = {
            '検索方法': self.search_method_var.get(),
            '番号': self.number_var.get().strip(),
            '梱包期限日': self.deadline_cal.get_date().strftime('%Y/%m/%d'),
            '梱包依頼摘要': self.note_text.get("1.0", tk.END).strip(),
            '梱包担当者': self.person_var.get(),
            '出力除外選択': 'はい' if self.exclude_var.get() else 'いいえ',
            '受注残0出力': 'はい' if self.zero_packing_var.get() else 'いいえ',
            '梱包可能数変更': 'はい' if self.change_quantity_var.get() else 'いいえ',
            '梱包明細': '梱包明細を依頼する案件' if self.packing_detail_var.get() else ''
        }
        
        self.dialog.destroy()
    
    def on_cancel(self):
        """キャンセルボタンが押された時の処理"""
        self.result = None
        self.dialog.destroy()

def enforce_os_packing_detail_status(row_data):
    """OS案件は梱包明細を依頼する案件としてGUI行データに即時反映する。"""
    try:
        search_method = row_data.get('検索方法', '')
        input_value = clean_input(row_data.get('番号', ''))
        is_os_case = False

        if search_method == '受注番号':
            is_os_case = str(input_value).upper().startswith("OS")
        elif search_method == '見積管理番号':
            is_os_case = is_os_case_by_estimate_no(input_value)

        if is_os_case:
            row_data['梱包明細'] = '梱包明細を依頼する案件'
    except Exception as e:
        print(f"OS案件の梱包明細自動反映でエラー: {e}")

    return row_data

def add_packing_row():
    """梱包依頼行を追加"""
    dialog = PackingRowDialog(root)
    root.wait_window(dialog.dialog)
    
    if dialog.result:
        dialog.result = enforce_os_packing_detail_status(dialog.result)
        # Treeviewに行を追加 - packing_columnsを使用
        values = tuple(dialog.result[col] for col in packing_columns)
        packing_tree.insert('', 'end', values=values)
        
        # データリストにも追加
        packing_data_list.append(dialog.result)


def edit_packing_row():
    """選択された梱包依頼行を編集"""
    selected_items = packing_tree.selection()
    if not selected_items:
        messagebox.showwarning("警告", "編集する行を選択してください。")
        return
    
    item = selected_items[0]
    # 現在の値を取得
    current_values = packing_tree.item(item, 'values')
    current_data = dict(zip(packing_columns, current_values))
    
    dialog = PackingRowDialog(root, current_data)
    root.wait_window(dialog.dialog)
    
    if dialog.result:
        dialog.result = enforce_os_packing_detail_status(dialog.result)
        # Treeviewを更新
        values = tuple(dialog.result[col] for col in packing_columns)
        packing_tree.item(item, values=values)
        
        # データリストも更新
        row_index = packing_tree.index(item)
        packing_data_list[row_index] = dialog.result

def on_packing_tree_double_click(event):
    """梱包依頼一覧の行をダブルクリックした際に編集ダイアログを開く"""
    clicked_item = packing_tree.identify_row(event.y)
    if not clicked_item:
        return
    packing_tree.selection_set(clicked_item)
    packing_tree.focus(clicked_item)
    edit_packing_row()


def delete_packing_row():
    """選択された梱包依頼行を削除"""
    selected_items = packing_tree.selection()
    if not selected_items:
        messagebox.showwarning("警告", "削除する行を選択してください。")
        return
    
    if messagebox.askyesno("確認", "選択した行を削除しますか？"):
        # 逆順で削除してインデックスのずれを防ぐ
        for item in reversed(selected_items):
            row_index = packing_tree.index(item)
            packing_tree.delete(item)
            # データリストからも削除
            del packing_data_list[row_index]


def execute_selected_rows():
    """選択された行を実行"""
    selected_items = packing_tree.selection()
    if not selected_items:
        messagebox.showwarning("警告", "実行する行を選択してください。")
        return
    
    if messagebox.askyesno("確認", f"{len(selected_items)}行の梱包依頼書を生成しますか？"):
        execute_rows([packing_tree.index(item) for item in selected_items])


def execute_all_rows():
    """全ての行を実行"""
    if not packing_data_list:
        messagebox.showwarning("警告", "実行する行がありません。")
        return
    
    if messagebox.askyesno("確認", f"{len(packing_data_list)}行の梱包依頼書を生成しますか？"):
        execute_rows(list(range(len(packing_data_list))))


def execute_rows(row_indices):
    """指定された行インデックスの梱包依頼書を生成"""
    success_count = 0
    error_count = 0
    error_details = []
    success_files = []

    # 複数件処理の場合はバッチ接続を使用（OneDrive同期との競合防止）
    use_batch = len(row_indices) > 1
    batch_db = None

    if use_batch:
        batch_db = BatchDBConnection.get_instance()
        if not batch_db.start_batch():
            messagebox.showerror("エラー", "データベース接続の確立に失敗しました。")
            return

    try:
        for row_index in row_indices:
            try:
                row_data = packing_data_list[row_index]
                result = process_single_packing_request(row_data, use_batch_connection=use_batch)
                if result:
                    tree_items = packing_tree.get_children()
                    if 0 <= row_index < len(tree_items):
                        item_id = tree_items[row_index]
                        values = tuple(row_data.get(col, '') for col in packing_columns)
                        packing_tree.item(item_id, values=values)
                    success_count += 1
                    success_files.append(f"行{row_index + 1}: {row_data['番号']}")
                else:
                    error_count += 1
                    error_details.append(f"行{row_index + 1}: 処理がキャンセルされました")
            except Exception as e:
                error_count += 1
                error_details.append(f"行{row_index + 1}: {str(e)}")
    finally:
        # バッチ接続を必ずクローズ
        if use_batch and batch_db:
            batch_db.end_batch()

    # 結果表示
    result_message = f"処理完了\n成功: {success_count}件\n失敗: {error_count}件"

    if success_files:
        result_message += f"\n\n成功した処理:\n" + "\n".join(success_files[:5])
        if len(success_files) > 5:
            result_message += f"\n...他{len(success_files) - 5}件"

    if error_details:
        result_message += f"\n\nエラー詳細:\n" + "\n".join(error_details[:3])  # エラーは3件まで表示
        if len(error_details) > 3:
            result_message += f"\n...他{len(error_details) - 3}件"

    messagebox.showinfo("実行結果", result_message)

def generate_unique_number(use_batch_connection=False):
    """
    JKxxXxxxxx形式（例: JK01X00012）のユニーク番号を生成し、
    既に登録されていなければDBに登録して返します（最大1000回試行）。

    use_batch_connection: Trueの場合、BatchDBConnectionを使用
    """
    # バッチ接続を使用するかどうか
    batch_conn = None
    if use_batch_connection:
        batch_db = BatchDBConnection.get_instance()
        if batch_db.is_batch_active():
            batch_conn = batch_db.get_generated_numbers_connection()

    generated_numbers = load_generated_numbers_with_retry(external_conn=batch_conn)
    print(f"読み込まれた生成済み番号: {len(generated_numbers)}個")
    attempts = 0
    while attempts < 1000:
        number = f"JK{random.randint(0, 99):02d}X{random.randint(0, 99999):05d}"
        if number not in generated_numbers:
            if save_generated_number_with_retry(number, external_conn=batch_conn):
                print(f"新しい番号を生成: {number}")
                return number
            else:
                print(f"番号の保存に失敗: {number}")
        attempts += 1
    raise ValueError("1000回の試行後も一意の番号を生成できませんでした")

def generate_qr_code(data, box_size=10):
    """
    QRコード（PNG）をバイナリとして返します。失敗時はNoneを返します。
    
    Args:
        data: QRコードにエンコードするデータ
        box_size: 各ボックスのピクセルサイズ（デフォルト: 10）
    """
    try:
        qr = qrcode.QRCode(version=1, box_size=box_size, border=4)
        qr.add_data(data)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, format="PNG")
        return buffer.getvalue()
    except Exception as e:
        print(f"QRコード生成エラー: {e}")
        return None

def clean_input(input_string):
    """入力文字列の末尾改行等を除去します。"""
    print(f"元の入力: {repr(input_string)}")
    cleaned = input_string.rstrip('\n\r')
    print(f"クリーニング後の入力: {repr(cleaned)}")
    return cleaned

def select_order_number_dialog(parent, order_numbers, df):
    """
    見積管理番号で検索した場合に複数の受注番号がヒットした場合、
    どの受注番号のデータを使用するか選択するダイアログを表示します。
    """
    dialog = Toplevel(parent)
    dialog.title("受注番号選択")
    dialog.geometry("800x400")
    
    label = tk.Label(dialog, text="この見積管理番号には複数の受注番号が紐づいています。\n使用する受注番号を選択してください:")
    label.pack(pady=10)
    
    # 受注番号リストをリストボックスに表示
    list_frame = tk.Frame(dialog)
    list_frame.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
    
    listbox = Listbox(list_frame, width=90, height=15)
    scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=listbox.yview)
    listbox.configure(yscrollcommand=scrollbar.set)
    
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    # 各受注番号の詳細情報を表示
    for order_num in order_numbers:
        order_items = df[df['受注番号'] == order_num]
        sample_row = order_items.iloc[0]
        customer_name = sample_row['得意先名']
        item_count = len(order_items)
        display_text = f"{order_num} - {customer_name} ({item_count}件の商品)"
        listbox.insert(tk.END, display_text)
    
    selected_order = [None]  # リスト内に変数を格納して参照を共有
    
    def on_select():
        selected_indices = listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("警告", "受注番号を選択してください。")
            return
        
        selected_text = listbox.get(selected_indices[0])
        # 表示テキストから受注番号部分を抽出
        selected_order[0] = selected_text.split(" - ")[0]
        dialog.destroy()
    
    def on_cancel():
        dialog.destroy()
    
    # ボタンフレーム
    btn_frame = tk.Frame(dialog)
    btn_frame.pack(pady=10)
    
    select_btn = Button(btn_frame, text="選択", command=on_select)
    select_btn.pack(side=tk.LEFT, padx=10)
    
    cancel_btn = Button(btn_frame, text="キャンセル", command=on_cancel)
    cancel_btn.pack(side=tk.LEFT, padx=10)
    
    # モーダルダイアログとして表示
    dialog.transient(parent)
    dialog.grab_set()
    parent.wait_window(dialog)
    
    return selected_order[0]

class SelectOutputItemsDialog:
    """
    出力するI/no.をユーザーがチェックボックスで選択できるダイアログです。
    """
    def __init__(self, parent, items, target_number, search_method):
        self.top = tk.Toplevel(parent)
        
        # タイトルに対象番号を追加
        search_type = "受注番号" if search_method == "order_number" else "見積管理番号"
        self.top.title(f"出力するI/no.を選択 - {search_type}: {target_number}")
        self.top.geometry("600x500")
        
        self.items = items
        self.selected_items = []
        self.item_vars = []
        
        # メインフレーム
        main_frame = tk.Frame(self.top, padx=10, pady=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # 説明ラベル（対象番号を表示）
        info_label = tk.Label(main_frame, 
                             text=f"【{search_type}: {target_number}】\n出力するアイテムを選択してください",
                             font=("Arial", 12, "bold"), fg="blue")
        info_label.pack(pady=(0, 10))
        
        control_frame = tk.Frame(main_frame)
        control_frame.pack(fill=tk.X, pady=(0, 8))

        select_all_button = Button(control_frame, text="全選択", command=self.select_all)
        select_all_button.pack(side=tk.LEFT, padx=(0, 5))

        clear_all_button = Button(control_frame, text="全解除", command=self.clear_all)
        clear_all_button.pack(side=tk.LEFT)

        # チェックボックス一覧（スクロール対応）
        list_container = tk.Frame(main_frame)
        list_container.pack(fill=tk.BOTH, expand=True)

        canvas = tk.Canvas(list_container)
        scrollbar = tk.Scrollbar(list_container, orient="vertical", command=canvas.yview)
        self.checkbox_frame = tk.Frame(canvas)

        self.checkbox_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.checkbox_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        for item in items:
            var = IntVar(value=1)
            self.item_vars.append(var)
            text = f"{item['I/no']} - {item['商品コード']} - {item['商品名']}"
            chk = Checkbutton(self.checkbox_frame, text=text, variable=var, anchor='w', justify='left')
            chk.pack(fill=tk.X, anchor='w')
        
        # ボタンフレーム
        button_frame = tk.Frame(main_frame)
        button_frame.pack(pady=10)
        
        select_button = Button(button_frame, text="選択", command=self.on_select)
        select_button.pack(side=tk.LEFT, padx=10)
        
        cancel_button = Button(button_frame, text="キャンセル", command=self.on_cancel)
        cancel_button.pack(side=tk.LEFT, padx=10)
        
        # ダイアログを中央に配置
        self.top.transient(parent)
        self.top.grab_set()
        self.top.update_idletasks()
        x = (self.top.winfo_screenwidth() // 2) - (self.top.winfo_width() // 2)
        y = (self.top.winfo_screenheight() // 2) - (self.top.winfo_height() // 2)
        self.top.geometry(f"+{x}+{y}")

    def select_all(self):
        for var in self.item_vars:
            var.set(1)

    def clear_all(self):
        for var in self.item_vars:
            var.set(0)

    def on_select(self):
        self.selected_items = [self.items[i] for i, var in enumerate(self.item_vars) if var.get()]
        if not self.selected_items:
            messagebox.showwarning("警告", "出力するアイテムを1件以上選択してください。")
            return
        self.top.destroy()
    
    def on_cancel(self):
        self.selected_items = []
        self.top.destroy()

def select_items_to_output(df, order_number, target_number, search_method):
    """
    受注番号に紐づくI/no.リストを表示し、ユーザーが出力したいアイテムを選択可能にします。
    選択されたアイテムのみdfに残します。
    """
    # I/no.の昇順でソートしてからリストに変換
    items_df = df[df['受注番号'] == order_number][['明細_共通項目2', '明細_商品コード', '明細_商品受注名']]\
        .rename(columns={'明細_共通項目2': 'I/no', '明細_商品コード': '商品コード', '明細_商品受注名': '商品名'})\
        .sort_values('I/no')  # I/no.の昇順でソート
    
    items = items_df.to_dict('records')
    
    dialog = SelectOutputItemsDialog(root, items, target_number, search_method)
    root.wait_window(dialog.top)
    selected_items = dialog.selected_items
    selected_ino = []
    excluded_ino = []
    if selected_items:
        search_type = "受注番号" if search_method == "order_number" else "見積管理番号"
        confirm_message = f"【{search_type}: {target_number}】\n以下の商品を出力します。よろしいですか？\n\n"
        for item in selected_items:
            confirm_message += f"[{item['I/no']}][{item['商品コード']}][{item['商品名']}]\n"
        if messagebox.askyesno("確認", confirm_message):
            selected_ino = [item['I/no'] for item in selected_items]
            all_ino = [item['I/no'] for item in items]
            excluded_ino = [i_no for i_no in all_ino if i_no not in selected_ino]
            df = df[~(df['受注番号'] == order_number) | (df['明細_共通項目2'].isin(selected_ino))]
    return df, excluded_ino

def get_save_path_for_packing_person(packing_person):
    """
    選択された梱包担当者に応じた候補フォルダのうち、存在するパスを返します。
    """
    candidates = get_packing_person_path_candidates(packing_person)
    for path in candidates:
        if os.path.exists(path):
            return path
    raise FileNotFoundError("該当の保存先フォルダが見つかりません。候補: " + ", ".join(candidates))


def get_packing_person_path_candidates(packing_person, filename=None):
    """
    梱包担当者に応じた保存先候補パス（3パターン）を返します。
    filename を指定した場合は各候補フォルダ配下のファイルパスを返します。
    """
    username = getuser()
    suffix = ""
    if packing_person == "11_細田":
        suffix = "11_細田"
    elif packing_person == "12_平松":
        suffix = "12_平松"
    elif packing_person == "13_坂上":
        suffix = "13_坂上"
    elif packing_person == "16_土田":
        suffix = "16_土田"
    else:
        raise ValueError("無効な梱包担当者です。")

    base_candidates = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\事前梱包依頼書\\{suffix}",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\事前梱包依頼書\\{suffix}",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\事前梱包依頼書\\{suffix}"
    ]

    if filename:
        return [os.path.join(base, filename) for base in base_candidates]
    return base_candidates


def resolve_output_file_path_for_current_pc(output_file_path, packing_person):
    """
    output_file_path の実体が現PCにない場合、梱包担当者の3候補パスで存在する実体を解決します。
    """
    if output_file_path and os.path.exists(output_file_path):
        return output_file_path

    filename = os.path.basename(output_file_path) if output_file_path else ""
    if not filename:
        return output_file_path

    try:
        candidate_paths = get_packing_person_path_candidates(packing_person, filename=filename)
    except Exception:
        return output_file_path

    for candidate in candidate_paths:
        if os.path.exists(candidate):
            return candidate

    return candidate_paths[0] if candidate_paths else output_file_path


def process_single_packing_request(row_data, use_batch_connection=False):
    """
    単一の梱包依頼を処理
    use_batch_connection: Trueの場合、BatchDBConnectionを使用
    """
    try:
        username = getuser()

        ### (1) マージ済みデータをキャッシュから取得 ###
        df4 = get_merged_data_for_packing().copy()
        excluded_codes = ['888888-88888','777777-77777']

        # 在庫一覧データをキャッシュから取得
        df_inventory = get_cached_inventory_data()

        ### (2) 行データから設定を取得 ###
        search_method = "order_number" if row_data['検索方法'] == '受注番号' else "estimate_no"
        input_value = clean_input(row_data['番号'])
        packaging_note = row_data['梱包依頼摘要']
        packing_person = row_data['梱包担当者']
        show_exclude = row_data['出力除外選択'] == 'はい'
        show_zero = row_data['受注残0出力'] == 'はい'
        change_quantity = row_data['梱包可能数変更'] == 'はい'
        packing_detail_requested = row_data.get('梱包明細', '') == '梱包明細を依頼する案件'
        packaging_deadline = datetime.strptime(row_data['梱包期限日'], '%Y/%m/%d').date()
        
        if not input_value:
            raise ValueError("検索条件が入力されていません。")
        
        # 【変更2】見積管理番号での検索条件を「取込伝票番号」に変更
        selected_order_number = None
        if search_method == "order_number":
            order_df = df4[df4['受注番号'] == input_value]
            if order_df.empty:
                raise ValueError(f"指定された受注番号 {input_value} のデータが見つかりません。")
            selected_order_number = input_value
        else:
            order_df = df4[df4['取込伝票番号'] == input_value]
            if order_df.empty:
                raise ValueError(f"指定された見積管理番号 {input_value} のデータが見つかりません。")
            
            order_numbers = order_df['受注番号'].unique()
            if len(order_numbers) > 1:
                selected_order_number = select_order_number_dialog(root, order_numbers, order_df)
                if not selected_order_number:
                    return False
                order_df = order_df[order_df['受注番号'] == selected_order_number]
            else:
                selected_order_number = order_numbers[0]
                order_df = order_df[order_df['受注番号'] == selected_order_number]

        is_os_order = str(selected_order_number).upper().startswith("OS")
        if is_os_order:
            packing_detail_requested = True
            row_data['梱包明細'] = '梱包明細を依頼する案件'
        
        # 出力対象アイテムの選択
        excluded_ino = []
        if show_exclude:
            df4_filtered, excluded_ino = select_items_to_output(df4, selected_order_number, input_value, search_method)
            if search_method == "order_number":
                order_df = df4_filtered[df4_filtered['受注番号'] == selected_order_number]
            else:
                order_df = df4_filtered[(df4_filtered['受注番号'] == selected_order_number) & (df4_filtered['取込伝票番号'] == input_value)]
        
        header_row = order_df.loc[
            order_df[~order_df['明細_商品コード'].isin(excluded_codes)]['明細_共通項目2'].astype(int).idxmin()
        ]
        order_df['梱包可能数'] = order_df.apply(
            lambda row: (row['明細_自社在庫引当数量'] - row['明細_出荷売上数量'])
                        if row['明細_倉庫コード'] in ["TYT01", "TYT02"]
                        else (row['明細_発注引当仕入数量'] - row['明細_出荷売上数量']),
            axis=1
        )
        
        # 受注金額の集計
        valid_packing_df = order_df[order_df['梱包可能数'] > 0]
        total_order_amount = valid_packing_df['明細_受注金額'].sum()
        print(f"集計対象レコード数: {len(valid_packing_df)}")
        print(f"合計受注金額: ¥{total_order_amount:,.0f}")
        
        # 梱包可能数を変更する
        if change_quantity:
            packing_dialog = PackingQuantityDialog(root, order_df, input_value, search_method)
            root.wait_window(packing_dialog.dialog)
            
            if packing_dialog.result is not None:
                for i_no, new_quantity in packing_dialog.result.items():
                    order_df.loc[order_df['明細_共通項目2'] == i_no, '梱包可能数'] = new_quantity
                
                valid_packing_df = order_df[order_df['梱包可能数'] > 0]
                total_order_amount = valid_packing_df['明細_受注金額'].sum()
                print(f"梱包可能数変更後の集計対象レコード数: {len(valid_packing_df)}")
                print(f"梱包可能数変更後の合計受注金額: ¥{total_order_amount:,.0f}")
            else:
                return False
        
        zero_packing_ino = []
        if not show_zero:
            zero_packing_df = order_df[(order_df['梱包可能数'] <= 0) & (order_df['受注残数'] <= 0)]
            zero_packing_ino = zero_packing_df['明細_共通項目2'].tolist()
            order_df = order_df[(order_df['梱包可能数'] > 0) | (order_df['受注残数'] > 0)]
        detail_df = order_df.sort_values('明細_共通項目2')

        ### (3) Excelファイルを作成 ###
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "事前：梱包依頼書"
        ws['A1'] = "事前：梱包依頼書"
        ws['A1'].font = Font(name='メイリオ', size=22, bold=True)
        ws['A1'].alignment = Alignment(horizontal='left')
        packing_detail_text = "必要" if packing_detail_requested else "なし"
        ws.merge_cells('J1:L2')
        ws['J1'] = f"【梱包明細：{packing_detail_text}】"
        ws['J1'].font = Font(name='メイリオ', size=16, bold=True, underline='single')
        ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
        black_emphasis_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        for row_idx in range(1, 3):
            for col_idx in range(10, 13):
                ws.cell(row=row_idx, column=col_idx).border = black_emphasis_border
        unique_number = generate_unique_number(use_batch_connection=use_batch_connection)
        ws['A2'] = f"事前梱包依頼番号: {unique_number}"
        ws['A3'] = f"梱包期限日: {packaging_deadline.strftime('%Y/%m/%d')}"
        
        # 【変更3】Excelヘッダーで「取込伝票番号」を参照
        ws['A4'] = f"見積管理番号: {header_row['取込伝票番号']}"
        ws['A5'] = f"船名: {header_row['受注件名']}"
        ws['A6'] = f"受注番号: {header_row['受注番号']}"
        ws['A7'] = f"客注番号: {header_row['客注番号']}"
        order_numbers = order_df['発注番号'].unique()
        order_numbers = [o for o in order_numbers if o]
        if len(order_numbers) > 1:
            order_numbers_str = ', '.join(order_numbers)
        else:
            order_numbers_str = order_numbers[0] if order_numbers else "発注番号なし"
        ws['A8'] = f"発注番号: {order_numbers_str}"
        ws['A9'] = f"得意先名: {header_row['得意先名']}"
        ws['A10'] = f"受渡場所名: {header_row['受渡場所名']}"
        ws['A11'] = f"営業担当者: {header_row['社員名']}"
        ws['A12'] = f"梱包依頼摘要: {packaging_note}"
        if excluded_ino:
            ws['A13'] = f"出力除外I/no.：{','.join(map(str, excluded_ino))}"
            ws['A13'].font = Font(name='メイリオ', size=12, bold=True)
            ws['A13'].alignment = Alignment(horizontal='left')
        for row_cells in ws['A1:A13']:
            for c in row_cells:
                c.alignment = Alignment(horizontal='left')
                if c.row != 1:
                    c.font = Font(name='メイリオ', size=12, bold=True)
        headers = [
            "QR_L", "I/no._L", "倉庫", "受注数量", "商品コード",
            "商品受注名", "受注残数", "梱包可能数", "ロット番号",
            "QR_R", "I/no._R", "備考"
        ]
        header_fill = PatternFill(patternType="solid", fgColor="D9D9D9")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col_idx, head_text in enumerate(headers, start=1):
            cell = ws.cell(row=15, column=col_idx, value=head_text)
            cell.font = Font(name='メイリオ', bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            cell.fill = header_fill
        def px_to_excel_col_width(px): 
            return px / 7.0
        col_width_map = {
            'A': 60,
            'B': 72,
            'C': 62,
            'D': 78,
            'E': 136,
            'F': 308,
            'G': 78,
            'H': 95,
            'I': 110,
            'J': 60,
            'K': 72,
            'L': 113
        }
        for col_letter, px_val in col_width_map.items():
            ws.column_dimensions[col_letter].width = px_to_excel_col_width(px_val)
        row_start = 16
        for row_index, data in enumerate(detail_df.itertuples(index=False), start=row_start):
            warehouse_code = data.明細_倉庫コード
            if warehouse_code in ["TYT01", "TYT02"]:
                matched_inv = df_inventory[df_inventory['商品コード'] == data.明細_商品コード]
                if not matched_inv.empty:
                    inv_row = matched_inv.iloc[0]
                    lot_number = inv_row['ロット番号']
                    remark_text = f"棚番{inv_row['棚番１']}"
                else:
                    lot_number = ""
                    remark_text = "在庫なし"
            else:
                lot_number = data.明細_ロット番号
                remark_text = ""
            if data.受注残数 != 0 and data.受注残数 > data.梱包可能数:
                remark_text = f"未入荷数：{data.受注残数 - data.梱包可能数}"
            elif (warehouse_code in ["TYT01", "TYT02"] and data.明細_自社在庫引当数量 == 0) or \
                 (warehouse_code not in ["TYT01", "TYT02"] and data.明細_発注引当仕入数量 == 0):
                remark_text = "未入荷"

            i_no_raw = data.明細_共通項目2
            try:
                i_no_padded = f"{int(i_no_raw):03d}"
            except:
                i_no_padded = str(i_no_raw)
            qr_data = f"{lot_number}_{int(data.梱包可能数):03d}" if lot_number else ""
            if data.梱包可能数 == 0:
                lot_number = ""
                qr_data = ""
            is_even = (row_index % 2 == 0)
            cell_qr_left = ws.cell(row=row_index, column=1)
            cell_qr_left.alignment = Alignment(horizontal='center', vertical='center')
            cell_qr_right = ws.cell(row=row_index, column=10)
            cell_qr_right.alignment = Alignment(horizontal='center', vertical='center')
            if is_even:
                if qr_data:
                    qr_left_bytes = generate_qr_code(qr_data)
                    if qr_left_bytes:
                        qr_buffer = BytesIO(qr_left_bytes)
                        qr_buffer.seek(0)  # Read position reset for Excel compatibility
                        img_left = Image(qr_buffer)
                        img_left.width = 50
                        img_left.height = 50
                        img_left.anchor = cell_qr_left.coordinate
                        ws.add_image(img_left)
                    else:
                        cell_qr_left.value = "QRエラー"
                else:
                    cell_qr_left.value = ""
                cell_qr_right.value = ""
            else:
                if qr_data:
                    qr_right_bytes = generate_qr_code(qr_data)
                    if qr_right_bytes:
                        qr_buffer = BytesIO(qr_right_bytes)
                        qr_buffer.seek(0)  # Read position reset for Excel compatibility
                        img_right = Image(qr_buffer)
                        img_right.width = 50
                        img_right.height = 50
                        img_right.anchor = cell_qr_right.coordinate
                        ws.add_image(img_right)
                    else:
                        cell_qr_right.value = "QRエラー"
                else:
                    cell_qr_right.value = ""
                cell_qr_left.value = ""
            ws.cell(row=row_index, column=2, value=i_no_raw)
            ws.cell(row=row_index, column=3, value=warehouse_code)
            ws.cell(row=row_index, column=4, value=data.明細_受注数量)
            ws.cell(row=row_index, column=5, value=data.明細_商品コード)
            ws.cell(row=row_index, column=6, value=data.明細_商品受注名)
            ws.cell(row=row_index, column=7, value=data.受注残数)
            ws.cell(row=row_index, column=8, value=data.梱包可能数)
            ws.cell(row=row_index, column=9, value=lot_number)
            ws.cell(row=row_index, column=11, value=i_no_raw)
            ws.cell(row=row_index, column=12, value=remark_text)
            ws.row_dimensions[row_index].height = 40
        max_row_used = ws.max_row
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row_idx in range(16, max_row_used + 1):
            for col_idx in range(1, 13):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.font = Font(name='メイリオ', size=12, bold=True)
                if col_idx in [2, 5, 6, 11]:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            cell_h = ws.cell(row=row_idx, column=8)
            cell_h.font = Font(name='メイリオ', size=22, bold=True, underline='single')
        
        # 保存先は、選択された梱包担当者に応じたフォルダへ振り分けます。
        save_path = get_save_path_for_packing_person(packing_person)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        unique_number_str = unique_number
        
        # ファイル名に見積管理番号を追加（検索方法が見積管理番号の場合）
        if search_method == "estimate_no":
            filename = f"事前梱包依頼書_{selected_order_number}_{header_row['取込伝票番号']}_{unique_number_str}_{timestamp}.xlsx"
        else:
            filename = f"事前梱包依頼書_{selected_order_number}_{unique_number_str}_{timestamp}.xlsx"

        form_output_path = os.path.join(save_path, filename)

        # OneDrive同期競合を避けるため、一時フォルダに保存してからコピー
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, filename)
        wb.save(temp_path)

        # 一時ファイルをOneDriveフォルダにコピー
        shutil.copy2(temp_path, form_output_path)
        resolved_output_path = resolve_output_file_path_for_current_pc(form_output_path, packing_person)

        # 一時ファイルを削除
        try:
            os.remove(temp_path)
        except:
            pass  # Ignore cleanup errors
        clipboard_text = f"{unique_number_str}にて事前梱包依頼済"
        pyperclip.copy(clipboard_text)
        
        # 完了メッセージの表示
        completion_message = f"事前梱包依頼書が作成されました。\n保存先: {form_output_path}\n\nクリップボードにコピーされた文字列:\n{clipboard_text}"
        root.after(0, lambda: messagebox.showinfo("完了", completion_message))
        
        # ★追加：ファイルを開くかどうか確認★
        def ask_and_open_file():
            if messagebox.askyesno("確認", "作成したExcelファイルを開きますか？"):
                try:
                    subprocess.Popen(["start", "", form_output_path], shell=True)
                except Exception as e:
                    messagebox.showerror("エラー", f"ファイルを開く際にエラーが発生しました:\n{str(e)}")
        
        root.after(0, ask_and_open_file)
        # ★追加ここまで★
        
        # 【変更4】Teams通知用のヘッダー情報で「取込伝票番号」を参照
        header_info = {
            "title": "事前：梱包依頼書",
            "unique_number": unique_number,
            "deadline": packaging_deadline.strftime('%Y/%m/%d'),
            "estimate_no": header_row['取込伝票番号'],
            "ship_name": header_row['受注件名'],
            "order_no": header_row['受注番号'],
            "customer_order_no": header_row['客注番号'],
            "order_numbers": order_numbers_str,
            "customer_name": header_row['得意先名'],
            "customer_code": header_row.get('得意先', ''),
            "delivery_location": header_row['受渡場所名'],
            "salesperson": header_row['社員名'],
            "packing_person": packing_person,
            "packaging_note": packaging_note,
            "order_amount": total_order_amount,
            "packing_detail": 1 if packing_detail_requested else 0,
            "output_file_path": resolved_output_path
        }
        if excluded_ino:
            header_info["exclude_inos"] = excluded_ino
        
        # DBに保存（進行状況表示付き）
        if save_packing_request_with_detailed_feedback(header_info, detail_df, use_batch_connection=use_batch_connection):
            send_adaptive_card_to_teams(header_info)
        else:
            root.after(0, lambda: messagebox.showwarning("警告", "データベースへの保存に失敗しましたが、Excelファイルは正常に作成されました。"))
            return False
        
        if not show_zero and zero_packing_ino:
            root.after(0, lambda: messagebox.showinfo("通知", f"以下のI/no.製品は売上計上済かつ受注残がないため出力しません。\n{', '.join(map(str, zero_packing_ino))}"))
        
        return True
        
    except Exception as e:
        error_message = f"処理中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
        root.after(0, lambda: messagebox.showerror("エラー", error_message))
        with open("error_log.txt", "w", encoding='utf-8') as f:
            f.write(error_message)
        print("エラーログが error_log.txt に保存されました。")
        return False

# ファイル検索と開く機能
def find_and_open_packing_file():
   """
   入力された事前梱包依頼番号と選択された梱包担当者に基づいて、
   対応するExcelファイルを検索して開きます。
   """
   try:
       request_number = request_number_entry.get().strip()
       packing_person = file_packing_person_var.get()
       
       if not request_number:
           messagebox.showerror("エラー", "事前梱包依頼番号を入力してください。")
           return
           
       username = getuser()
       candidate_paths = [
           f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\事前梱包依頼書\\{packing_person}",
           f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\事前梱包依頼書\\{packing_person}",
           f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\事前梱包依頼書\\{packing_person}"
       ]
       
       # それぞれのパスで事前梱包依頼番号を含むファイルを検索
       found_files = []
       for base_path in candidate_paths:
           if os.path.exists(base_path):
               # ワイルドカード検索パターン
               search_pattern = os.path.join(base_path, f"*{request_number}*.xlsx")
               matching_files = glob.glob(search_pattern)
               found_files.extend(matching_files)
       
       if not found_files:
           messagebox.showerror("エラー", f"指定された事前梱包依頼番号 [{request_number}] を含むファイルが見つかりませんでした。\n\n検索パス:\n" + "\n".join(candidate_paths))
           return
           
       if len(found_files) > 1:
           # 複数のファイルが見つかった場合は選択ダイアログを表示
           selected_file = select_file_dialog(found_files)
           if not selected_file:
               return  # ユーザーがキャンセルした場合
       else:
           selected_file = found_files[0]
           
       # ファイルを開く
       try:
           subprocess.Popen(["start", "", selected_file], shell=True)
       except Exception as e:
           messagebox.showerror("エラー", f"ファイルを開く際にエラーが発生しました:\n{str(e)}")
           
   except Exception as e:
       error_message = f"処理中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
       messagebox.showerror("エラー", error_message)
       with open("error_log.txt", "w", encoding='utf-8') as f:
           f.write(error_message)
       print("エラーログが error_log.txt に保存されました。")

def select_file_dialog(file_list):
   """
   複数のファイルが見つかった場合に、どのファイルを開くか選択するダイアログを表示します。
   """
   dialog = Toplevel(root)
   dialog.title("ファイル選択")
   dialog.geometry("800x400")
   
   label = tk.Label(dialog, text="開くファイルを選択してください:")
   label.pack(pady=10)
   
   # ファイルリストをリストボックスに表示
   listbox = Listbox(dialog, width=90, height=15)
   for file_path in file_list:
       # ファイル名だけを表示すると分かりやすい
       file_name = os.path.basename(file_path)
       listbox.insert(tk.END, file_name)
   listbox.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)
   
   # ファイルパスマッピング
   file_mapping = {os.path.basename(path): path for path in file_list}
   
   selected_file = [None]  # リスト内に変数を格納して参照を共有
   
   def on_select():
       selected_indices = listbox.curselection()
       if not selected_indices:
           messagebox.showwarning("警告", "ファイルを選択してください。")
           return
       
       selected_name = listbox.get(selected_indices[0])
       selected_file[0] = file_mapping[selected_name]
       dialog.destroy()
   
   def on_cancel():
       dialog.destroy()
   
   # ボタンフレーム
   btn_frame = tk.Frame(dialog)
   btn_frame.pack(pady=10)
   
   select_btn = Button(btn_frame, text="選択", command=on_select)
   select_btn.pack(side=tk.LEFT, padx=10)
   
   cancel_btn = Button(btn_frame, text="キャンセル", command=on_cancel)
   cancel_btn.pack(side=tk.LEFT, padx=10)
   
   # モーダルダイアログとして表示
   dialog.transient(root)
   dialog.grab_set()
   root.wait_window(dialog)
   
   return selected_file[0]

# 欠品QR表示機能
def generate_missing_item_qr():
    """
    見積管理番号/発注番号/受注番号と行番号、数量からQRコードを表示します。
    商品コードと商品略名も合わせて表示します。
    """
    try:
        # 入力値を取得
        estimate_no = missing_estimate_no_entry.get().strip()
        order_no = missing_order_no_entry.get().strip()
        order_number = missing_order_number_entry.get().strip()
        line_no = missing_line_no_entry.get().strip()
        quantity_str = missing_quantity_entry.get().strip()
        
        # 入力チェック1: 検索キーの排他チェック
        filled_key_count = sum(1 for v in [estimate_no, order_no, order_number] if v)
        if filled_key_count == 0:
            messagebox.showerror("エラー", "見積管理番号・発注番号・受注番号のいずれか1つを入力してください。")
            return
        
        if filled_key_count > 1:
            messagebox.showerror("エラー", "見積管理番号・発注番号・受注番号は同時入力できません。\nどれか1つだけ入力してください。")
            return
        
        # 入力チェック2: 行番号
        if not line_no:
            messagebox.showerror("エラー", "行番号を入力してください。")
            return
        
        # 入力チェック3: 数量（空欄チェック）
        if not quantity_str:
            messagebox.showwarning("警告", "数量を入力してください。")
            return
        
        # 全角数字を半角に変換
        quantity_str = quantity_str.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
        
        # 数量の数値変換と検証
        try:
            quantity = int(quantity_str)
        except ValueError:
            messagebox.showerror("エラー", "数量は数値で入力してください。")
            return
        
        # 入力チェック4: 数量範囲（999超の警告）
        if quantity > 999:
            messagebox.showwarning("警告", "数量が999を超えています。\n999以下の値を入力してください。")
            return
        
        if quantity < 1:
            messagebox.showerror("エラー", "数量は1以上の値を入力してください。")
            return
        
        # 入荷データをキャッシュから取得
        df_arrival = get_cached_arrival_data()
        
        # 見積管理番号・発注番号・受注番号で照合
        if estimate_no:
            # 見積管理番号で検索
            matching_records = df_arrival[(df_arrival['明細_共通項目3'] == estimate_no) & 
                                        (df_arrival['明細_共通項目2'] == line_no)]
            search_type = "見積管理番号"
            search_key = estimate_no
        elif order_no:
            # 発注番号で検索
            matching_records = df_arrival[(df_arrival['発注番号'] == order_no) & 
                                        (df_arrival['明細_共通項目2'] == line_no)]
            search_type = "発注番号"
            search_key = order_no
        else:
            # 受注番号で検索（purchase_order_dataから発注番号を解決）
            df_purchase = get_cached_purchase_order_data()
            purchase_records = df_purchase[(df_purchase['受注番号'] == order_number) &
                                           (df_purchase['明細_共通項目2'] == line_no)]
            if purchase_records.empty:
                messagebox.showerror(
                    "エラー",
                    f"受注番号 [{order_number}] と行番号 [{line_no}] に一致する発注データが見つかりません。"
                )
                return

            linked_order_nos = purchase_records['発注番号'].dropna().astype(str).unique().tolist()
            linked_product_codes = purchase_records['明細_商品コード'].dropna().astype(str).unique().tolist()

            matching_records = df_arrival[(df_arrival['発注番号'].isin(linked_order_nos)) &
                                          (df_arrival['明細_共通項目2'] == line_no)]
            if linked_product_codes:
                matching_records = matching_records[matching_records['明細_商品コード'].isin(linked_product_codes)]

            search_type = "受注番号"
            search_key = order_number
        
        if matching_records.empty:
            messagebox.showerror("エラー", 
                               f"{search_type} [{search_key}] と行番号 [{line_no}] に一致するデータが見つかりません。")
            return
        
        # 最初の一致レコードから必要な情報を取得
        record = matching_records.iloc[0]
        lot_number = record['明細_ロット番号']
        product_code = record['明細_商品コード']
        product_name = record['明細_商品略名']
        
        if not lot_number:
            messagebox.showerror("エラー", "該当するロット番号がありません。")
            return
        
        # QRコード用のデータを生成（ロット番号_数量3桁ゼロ埋め）
        qr_data = f"{lot_number}_{quantity:03d}"
        
        # QRコードを生成（サイズを半分に）
        qr_bytes = generate_qr_code(qr_data, box_size=5)  # ← box_size=5を追加
        if not qr_bytes:
            messagebox.showerror("エラー", "QRコードの生成に失敗しました。")
            return
        
        # 結果を表示
        qr_image = PIL.Image.open(BytesIO(qr_bytes))
        qr_photo = PIL.ImageTk.PhotoImage(qr_image)
        
        # 既存の画像があれば削除
        missing_qr_display_label.config(image='')
        missing_qr_display_label.image = None
        
        # 新しい画像を表示
        missing_qr_display_label.config(image=qr_photo)
        missing_qr_display_label.image = qr_photo  # リファレンスを保持（ガベージコレクションを防ぐ）
        
        # 情報を表示（ロット番号、商品コード、商品略名、数量、QRデータ）
        info_text = f"ロット番号: {lot_number}\n商品コード: {product_code}\n商品略名: {product_name}\n数量: {quantity}\nQRデータ: {qr_data}"
        missing_info_var.set(info_text)
        
        # 枠付きのフレームにするために背景色と枠線色を設定
        missing_info_label.config(bg="white", fg="red", relief=tk.GROOVE, borderwidth=2, padx=5, pady=5)
        
    except Exception as e:
        error_message = f"処理中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
        messagebox.showerror("エラー", error_message)
        with open("error_log.txt", "w", encoding='utf-8') as f:
            f.write(error_message)
        print("エラーログが error_log.txt に保存されました。")

def date_to_serial(date_obj):
    """
    PythonのdateオブジェクをExcelシリアル値に変換
    """
    try:
        from datetime import date
        # Excelのベース日付は1900年1月1日（シリアル値1）
        # ただし、Excelは1900年を閏年として扱う（実際は平年）ため、
        # 1900年3月1日以降は+1する必要がある
        base_date = date(1899, 12, 30)  # Excelシリアル値1に対応
        delta = date_obj - base_date
        return delta.days
    except Exception as e:
        print(f"日付シリアル値変換エラー: {e}")
        return None

def get_shipment_label_file_path():
    """
    出荷指示ラベル用データファイルのパスを取得
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷指示ラベル用データ\\出荷指示ラベル用データ.csv",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷指示ラベル用データ\\出荷指示ラベル用データ.csv",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\出荷指示ラベル用データ\\出荷指示ラベル用データ.csv"
    ]
    
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    
    raise FileNotFoundError("出荷指示ラベル用データファイルが見つかりません。候補: " + ", ".join(candidate_paths))

def get_delivery_location_file_path():
    """
    受渡場所.csvファイルのパスを取得
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\受渡場所一覧\\受渡場所.csv",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\受渡場所一覧\\受渡場所.csv",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\受渡場所一覧\\受渡場所.csv"
    ]
    
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    
    raise FileNotFoundError("受渡場所.csvファイルが見つかりません。候補: " + ", ".join(candidate_paths))

def get_weight_master_db_path():
    """
    Weight_master.dbファイルのパスを取得
    """
    username = getuser()
    candidate_paths = [
        f"C:\\Users\\{username}\\OneDrive - 東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\重量関連\\Weight_master.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\重量関連\\Weight_master.db",
        f"C:\\Users\\{username}\\東邦ヤンマーテック株式会社\\CR推進本部 - CR推進本部フォルダ\\06_社内管理資料\\miraimiru移行関連\\フォルダ共有テスト\\現場用\\DB\\重量関連\\Weight_master.db"
    ]
    
    for path in candidate_paths:
        if os.path.exists(path):
            return path
    
    raise FileNotFoundError("Weight_master.dbファイルが見つかりません。候補: " + ", ".join(candidate_paths))

def get_product_weight(product_code):
    """
    製品コードから重量を取得
    """
    try:
        db_path = get_weight_master_db_path()
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        cursor.execute("""
        SELECT 重量 FROM Weight_master
        WHERE 製品コード = ?
        """, (product_code,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row and row[0] is not None:
            return float(row[0])
        else:
            return 0.0
            
    except Exception as e:
        print(f"重量取得エラー（製品コード: {product_code}）: {e}")
        return 0.0

def calculate_total_weight(detail_df):
    """
    明細データから総重量を計算
    各商品コードの重量 × 自社在庫引当数量の合計を返す
    """
    total_weight = 0.0
    
    for _, row in detail_df.iterrows():
        product_code = row['明細_商品コード']
        quantity = row['明細_自社在庫引当数量']
        
        # 重量マスタから重量を取得
        weight = get_product_weight(product_code)
        
        # 重量 × 数量を加算
        total_weight += weight * quantity
    
    return total_weight

def load_delivery_location_data():
    """
    受渡場所.csvファイルを読み込む
    """
    try:
        file_path = get_delivery_location_file_path()
        df = pd.read_csv(file_path, encoding='cp932', dtype=str, keep_default_na=False)
        return df
    except Exception as e:
        print(f"受渡場所データの読み込みでエラー: {e}")
        return None

def get_delivery_locations_by_customer(customer_code):
    """
    得意先コードに対応する受渡場所リストを取得
    """
    try:
        df = load_delivery_location_data()
        if df is None:
            print("受渡場所データの読み込みに失敗しました")
            return []
        
        # デバッグ用：CSVファイルの列名を表示
        print(f"受渡場所.csvの列名: {df.columns.tolist()}")
        
        # デバッグ用：検索対象の列の最初の数行を表示
        if '得意先コード-得意先枝番' in df.columns:
            print(f"得意先コード-得意先枝番列の最初の5行: {df['得意先コード-得意先枝番'].head().tolist()}")
        
        # 得意先コード-得意先枝番列で検索
        matching_rows = df[df['得意先コード-得意先枝番'] == customer_code]
        
        print(f"キー「{customer_code}」でのマッチング行数: {len(matching_rows)}")
        
        if matching_rows.empty:
            return []
        
        # 受渡場所名のリストを返す（重複除去）
        delivery_locations = matching_rows['受渡場所名'].unique().tolist()
        
        # 空文字や空白のみの値を除去
        delivery_locations = [loc for loc in delivery_locations if loc and str(loc).strip()]
        
        print(f"最終的な受渡場所リスト: {delivery_locations}")
        
        return delivery_locations
        
    except Exception as e:
        print(f"受渡場所取得でエラー: {e}")
        return []

def search_shipment_request_numbers():
    """
    受注番号と出荷予定日から出荷依頼番号を検索してQRコード表示
    """
    try:
        # 入力値を取得
        order_number = shipment_order_number_entry.get().strip()
        shipment_date = shipment_date_cal.get_date()  # DateEntryから日付を取得
        
        if not order_number:
            messagebox.showerror("エラー", "受注番号を入力してください。")
            return
        
        # 日付をシリアル値に変換
        serial_date = date_to_serial(shipment_date)
        if serial_date is None:
            messagebox.showerror("エラー", "日付の変換に失敗しました。")
            return
        
        # 検索キーを作成（受注番号 + シリアル値）
        search_key = f"{order_number}{serial_date}"
        print(f"検索キー: {search_key} (受注番号: {order_number}, 出荷予定日: {shipment_date}, シリアル値: {serial_date})")
        
        # CSVファイルを読み込み
        csv_file_path = get_shipment_label_file_path()
        df = pd.read_csv(csv_file_path, encoding='cp932', dtype=str, keep_default_na=False)
        
        # 必要な列が存在するかチェック
        required_columns = ['受注番号', '出荷予定日', '出荷依頼番号']
        optional_columns = ['受注件名', '受渡場所', '配送便']  # 追加情報用
        
        for col in required_columns:
            if col not in df.columns:
                raise ValueError(f"必要な列 '{col}' がCSVファイルに存在しません。")
        
        # CSVファイル内で検索キーと一致する行を検索
        matching_rows = []
        for _, row in df.iterrows():
            csv_order_number = str(row['受注番号']).strip()
            csv_shipment_date = str(row['出荷予定日']).strip()
            
            # CSVの出荷予定日がシリアル値として保存されている場合とテキスト形式の場合を両方考慮
            try:
                # シリアル値として保存されている場合
                csv_serial_date = int(float(csv_shipment_date))
            except ValueError:
                # テキスト形式で保存されている場合は変換
                try:
                    from datetime import datetime
                    csv_date_obj = datetime.strptime(csv_shipment_date, "%Y/%m/%d").date()
                    csv_serial_date = date_to_serial(csv_date_obj)
                except ValueError:
                    try:
                        csv_date_obj = datetime.strptime(csv_shipment_date, "%Y-%m-%d").date()
                        csv_serial_date = date_to_serial(csv_date_obj)
                    except ValueError:
                        continue  # 変換できない行はスキップ
            
            # 検索キーと照合
            csv_search_key = f"{csv_order_number}{csv_serial_date}"
            if csv_search_key == search_key:
                # 基本情報
                match_data = {
                    'order_number': csv_order_number,
                    'shipment_date': csv_shipment_date,
                    'serial_date': csv_serial_date,
                    'shipment_request_number': str(row['出荷依頼番号']).strip()
                }
                
                # 追加情報（存在する場合のみ）
                for col in optional_columns:
                    if col in df.columns:
                        match_data[col] = str(row[col]).strip()
                    else:
                        match_data[col] = ""
                
                matching_rows.append(match_data)
        
        if not matching_rows:
            messagebox.showinfo("検索結果", f"受注番号 '{order_number}' と出荷予定日 '{shipment_date}' に一致するデータが見つかりませんでした。")
            # 既存のQRコード表示をクリア
            clear_shipment_qr_display()
            return
        
        # 一致した出荷依頼番号のQRコードを生成・表示
        display_shipment_qr_codes(matching_rows)
        
        print(f"{len(matching_rows)}件の一致する出荷依頼番号が見つかりました。")
        
    except Exception as e:
        error_message = f"検索処理中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
        messagebox.showerror("エラー", error_message)
        with open("error_log.txt", "w", encoding='utf-8') as f:
            f.write(error_message)
        print("エラーログが error_log.txt に保存されました。")

def clear_shipment_qr_display():
    """
    出荷QRコード表示エリアをクリア
    """
    for widget in shipment_qr_display_frame.winfo_children():
        widget.destroy()

def display_shipment_qr_codes(matching_rows):
    """
    複数の出荷依頼番号のQRコードを表示（サイズ小、詳細情報拡張）
    """
    try:
        # 既存の表示をクリア
        clear_shipment_qr_display()
        
        for i, row_data in enumerate(matching_rows):
            shipment_request_number = row_data['shipment_request_number']
            
            # QRコードを生成
            qr_bytes = generate_qr_code(shipment_request_number)
            if qr_bytes:
                # QRコード画像を表示（サイズを小さく調整）
                qr_image = PIL.Image.open(BytesIO(qr_bytes))
                # QRコードサイズを縮小（元の60%に）
                new_size = (int(qr_image.width * 0.6), int(qr_image.height * 0.6))
                qr_image = qr_image.resize(new_size, PIL.Image.Resampling.LANCZOS)
                qr_photo = PIL.ImageTk.PhotoImage(qr_image)
                
                # 各QRコード用のフレーム
                qr_frame = tk.Frame(shipment_qr_display_frame, relief=tk.GROOVE, borderwidth=2, padx=15, pady=10)
                qr_frame.pack(pady=5, padx=10, fill=tk.X)
                
                # 左側：QRコード
                left_frame = tk.Frame(qr_frame)
                left_frame.pack(side=tk.LEFT, padx=10)
                
                # 出荷依頼番号ラベル
                number_label = tk.Label(left_frame, text=f"出荷依頼番号: {shipment_request_number}", 
                                      font=("Arial", 11, "bold"), fg="blue")
                number_label.pack()
                
                # QRコード画像
                qr_label = tk.Label(left_frame, image=qr_photo)
                qr_label.image = qr_photo  # ガベージコレクション防止
                qr_label.pack(pady=3)
                
                # 右側：詳細情報
                right_frame = tk.Frame(qr_frame)
                right_frame.pack(side=tk.LEFT, padx=20, fill=tk.X, expand=True)
                
                # 詳細情報の構築
                detail_lines = [
                    f"受注番号: {row_data['order_number']}",
                    f"出荷予定日: {row_data['shipment_date']}"
                ]
                
                # 追加情報（存在する場合のみ表示）
                if row_data.get('受注件名'):
                    detail_lines.append(f"受注件名: {row_data['受注件名']}")
                if row_data.get('受渡場所'):
                    detail_lines.append(f"受渡場所: {row_data['受渡場所']}")
                if row_data.get('配送便'):
                    detail_lines.append(f"配送便: {row_data['配送便']}")
                
                # 詳細情報を表示
                for line in detail_lines:
                    detail_label = tk.Label(right_frame, text=line, font=("Arial", 9), anchor='w')
                    detail_label.pack(anchor='w', pady=1)
                    
            else:
                # QR生成失敗時
                error_frame = tk.Frame(shipment_qr_display_frame, relief=tk.GROOVE, borderwidth=2, padx=10, pady=10)
                error_frame.pack(pady=10)
                
                error_label = tk.Label(error_frame, text=f"出荷依頼番号: {shipment_request_number}\nQRコード生成エラー", 
                                     font=("Arial", 10), fg="red")
                error_label.pack()
    
    except Exception as e:
        error_message = f"QRコード表示中にエラーが発生しました：{str(e)}"
        messagebox.showerror("エラー", error_message)

def bind_mousewheel_to_frame(widget):
    """ウィジェットとその子要素にマウスホイールイベントをバインド"""
    def _bind_to_mousewheel(event):
        widget.bind_all("<MouseWheel>", _on_mousewheel)

    def _unbind_from_mousewheel(event):
        widget.unbind_all("<MouseWheel>")

    def _on_mousewheel(event):
        if hasattr(widget, 'yview'):
            widget.yview_scroll(int(-1*(event.delta/120)), "units")

    widget.bind('<Enter>', _bind_to_mousewheel)
    widget.bind('<Leave>', _unbind_from_mousewheel)

class ShipmentStatusDialog:
    """
    依頼状況確認ダイアログクラス（複数ページ対応・梱包明細対応・PDF対応・受渡場所変更対応・統合明細記憶対応）
    """
    def __init__(self, parent, search_data_list, search_keys, search_method, shipment_request_type):
        self.parent = parent
        self.search_data_list = search_data_list  # 複数のデータセットのリスト
        self.search_keys = search_keys  # 複数の検索キーのリスト
        self.search_method = search_method
        self.shipment_request_type = shipment_request_type
        self.current_page = 0  # 現在のページ番号
        self.total_pages = len(search_data_list)
        self.result = None
        self.edited_quantities = {}  # 編集された数量を保存
        self.selected_shipment_numbers = {}  # ページごとの選択されたshipment_number
        self.selected_delivery_locations = {}  # ページごとの選択された受渡場所
        
        # ★新規追加：統合・結合明細を一時的に記憶するための辞書★
        self.consolidated_packing_data_by_page = {}  # ページごとの統合明細データ
        
        # ダイアログウィンドウを作成
        self.dialog = Toplevel(parent)
        self.dialog.title("依頼状況確認")
        self.dialog.geometry("1500x900")  # 高さを増やす
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # メインフレーム
        self.main_frame = tk.Frame(self.dialog)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # ページ遷移フレーム
        self.page_frame = tk.Frame(self.main_frame)
        self.page_frame.pack(fill=tk.X, pady=(0, 10))
        
        # ページ遷移ボタンとページ情報
        self.prev_button = tk.Button(self.page_frame, text="前ページ", command=self.prev_page)
        self.prev_button.pack(side=tk.LEFT, padx=5)
        
        self.page_label = tk.Label(self.page_frame, text="", font=("Arial", 12, "bold"))
        self.page_label.pack(side=tk.LEFT, padx=20)
        
        self.next_button = tk.Button(self.page_frame, text="次ページ", command=self.next_page)
        self.next_button.pack(side=tk.LEFT, padx=5)
        
        # タイトル
        self.title_label = tk.Label(self.main_frame, text="", font=("Arial", 14, "bold"))
        self.title_label.pack(pady=(0, 10))
        
        # ★新規追加：ヘッダー情報表示エリア★
        self.header_info_frame = tk.Frame(self.main_frame, relief=tk.GROOVE, borderwidth=2, bg="lightgray")
        self.header_info_frame.pack(fill=tk.X, pady=(0, 10), padx=5)
        
        # ヘッダー情報のラベル（動的に更新される）
        self.customer_name_label = tk.Label(self.header_info_frame, text="", font=("Arial", 11, "bold"), bg="lightgray")
        self.customer_name_label.pack(anchor='w', padx=10, pady=2)
        
        self.ship_name_label = tk.Label(self.header_info_frame, text="", font=("Arial", 11, "bold"), bg="lightgray")
        self.ship_name_label.pack(anchor='w', padx=10, pady=2)
        
        self.original_delivery_location_label = tk.Label(self.header_info_frame, text="", font=("Arial", 11, "bold"), bg="lightgray")
        self.original_delivery_location_label.pack(anchor='w', padx=10, pady=2)
        
        # ★新規追加：受渡場所変更セクション★
        self.delivery_change_frame = tk.Frame(self.main_frame)
        self.delivery_change_frame.pack(fill=tk.X, pady=(0, 10))
        
        # 受渡場所を変更するチェックボックス
        self.change_delivery_location = IntVar()
        self.change_delivery_checkbox = Checkbutton(
            self.delivery_change_frame, 
            text="受渡場所を変更する", 
            variable=self.change_delivery_location,
            command=self.on_delivery_change_toggled,
            font=("Arial", 11, "bold")
        )
        self.change_delivery_checkbox.pack(side=tk.LEFT, padx=5)
        
        # 受渡場所選択ドロップダウン（初期は非表示）
        self.delivery_location_label = tk.Label(self.delivery_change_frame, text="受渡場所:", font=("Arial", 11))
        self.delivery_location_var = StringVar(value="")
        self.delivery_location_combo = ttk.Combobox(
            self.delivery_change_frame, 
            textvariable=self.delivery_location_var, 
            width=30, 
            state="readonly"
        )
        self.delivery_location_combo.bind("<<ComboboxSelected>>", self.on_delivery_location_changed)
        
        # 出荷日選択セクション
        self.shipment_date_frame = tk.Frame(self.main_frame)
        self.shipment_date_frame.pack(pady=(0, 10))
        
        shipment_date_label = tk.Label(self.shipment_date_frame, text="出荷日:", font=("Arial", 12, "bold"))
        shipment_date_label.pack(side=tk.LEFT, padx=5)
        
        # 出荷日選択用のDateEntry
        self.shipment_date_cal = DateEntry(self.shipment_date_frame, width=12, background='darkblue', 
                                         foreground='white', borderwidth=2, locale='ja_JP', font=("Arial", 10))
        self.shipment_date_cal.pack(side=tk.LEFT, padx=5)
        
        # 梱包明細選択セクション
        self.packing_section = tk.Frame(self.main_frame)
        self.packing_section.pack(fill=tk.X, pady=(0, 10))
        
        packing_label = tk.Label(self.packing_section, text="梱包明細:", font=("Arial", 12, "bold"))
        packing_label.pack(side=tk.LEFT, padx=5)
        
        # shipment_number選択用プルダウン
        self.shipment_var = StringVar(value="")
        self.shipment_combo = ttk.Combobox(self.packing_section, textvariable=self.shipment_var, 
                                          width=20, state="readonly")
        self.shipment_combo.pack(side=tk.LEFT, padx=5)
        self.shipment_combo.bind("<<ComboboxSelected>>", self.on_shipment_selection_changed)
        
        # プレビューボタン
        preview_button = tk.Button(self.packing_section, text="プレビュー", command=self.show_packing_preview)
        preview_button.pack(side=tk.LEFT, padx=10)
        
        # コンテンツエリア（動的に更新される）
        self.content_frame = tk.Frame(self.main_frame)
        self.content_frame.pack(fill=tk.BOTH, expand=True)
        
        # ボタンフレーム
        button_frame = tk.Frame(self.main_frame)
        button_frame.pack(pady=10)

        # ★修正：数量比較ボタンを事前梱包依頼の場合のみ表示★
        if self.shipment_request_type == "advanced_packing":
            compare_button = tk.Button(button_frame, text="引当済と依頼済数量比較", command=self.compare_quantities)
            compare_button.pack(side=tk.LEFT, padx=10)

        # 梱包明細統合設定ボタン
        consolidation_button = tk.Button(button_frame, text="梱包明細統合設定", command=self.show_consolidation_dialog, bg="yellow")
        consolidation_button.pack(side=tk.LEFT, padx=10)

        # Excel出荷報告書出力ボタン
        excel_output_button = tk.Button(button_frame, text="Excel出荷報告書出力", command=self.output_shipment_report, bg="lightgreen")
        excel_output_button.pack(side=tk.LEFT, padx=10)
        
        # PDF出荷報告書出力ボタン（単体）
        pdf_output_button = tk.Button(button_frame, text="PDF出荷報告書出力", command=self.output_shipment_report_pdf, bg="lightblue")
        pdf_output_button.pack(side=tk.LEFT, padx=10)
        
        # PDF出荷報告書出力ボタン（一括）- 複数ページがある場合のみ表示
        if self.total_pages > 1:
            pdf_output_all_button = tk.Button(button_frame, text="PDF出荷報告書出力(一括)", command=self.output_all_shipment_reports_pdf, bg="orange")
            pdf_output_all_button.pack(side=tk.LEFT, padx=10)
        
        close_button = tk.Button(button_frame, text="閉じる", command=self.on_close)
        close_button.pack(side=tk.LEFT, padx=10)
        
        # 初期ページを表示
        self.update_display()
        
        # ダイアログを中央に配置
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
    
    def prev_page(self):
        """前のページに移動"""
        if self.current_page > 0:
            self.current_page -= 1
            self.update_display()
    
    def next_page(self):
        """次のページに移動"""
        if self.current_page < self.total_pages - 1:
            self.current_page += 1
            self.update_display()
    
    def get_current_search_data(self):
        """現在のページの検索データを取得"""
        return self.search_data_list[self.current_page]
    
    def get_current_search_key(self):
        """現在のページの検索キーを取得"""
        return self.search_keys[self.current_page]
    
    def update_display(self):
        """現在のページの表示を更新"""
        # ページ情報を更新
        self.page_label.config(text=f"{self.current_page + 1} / {self.total_pages}")
        
        # 前/次ボタンの状態を更新
        self.prev_button.config(state=tk.NORMAL if self.current_page > 0 else tk.DISABLED)
        self.next_button.config(state=tk.NORMAL if self.current_page < self.total_pages - 1 else tk.DISABLED)
        
        # タイトルを更新（見積管理番号の場合は受注番号も表示）
        current_key = self.get_current_search_key()
        current_data = self.get_current_search_data()
        
        if self.search_method == "estimate_no":
            # 見積管理番号の場合は対応する受注番号を取得
            if not current_data.empty:
                order_number = current_data.iloc[0]['受注番号']
                title_text = f"依頼状況確認 - {current_key}（受注番号：{order_number}）"
            else:
                title_text = f"依頼状況確認 - {current_key}"
        else:
            title_text = f"依頼状況確認 - {current_key}"
        
        self.title_label.config(text=title_text)
        
        # ヘッダー情報を更新
        self.update_header_info()
        
        # 受渡場所変更機能を更新
        self.update_delivery_location_options()
        
        # shipment_numberのプルダウンを更新
        self.update_shipment_combo()
        
        # コンテンツエリアをクリア
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        # 現在のページのコンテンツを作成
        self.create_content()
    
    def update_header_info(self):
        """ヘッダー情報を更新"""
        current_data = self.get_current_search_data()
        if not current_data.empty:
            sample_row = current_data.iloc[0]
            
            # 得意先名、受注件名、受渡場所名を表示
            customer_name = sample_row['得意先名']
            ship_name = sample_row['受注件名']
            delivery_location = sample_row['受渡場所名']
            
            self.customer_name_label.config(text=f"得意先名: {customer_name}")
            self.ship_name_label.config(text=f"受注件名: {ship_name}")
            self.original_delivery_location_label.config(text=f"受渡場所名: {delivery_location}")
        else:
            self.customer_name_label.config(text="得意先名: -")
            self.ship_name_label.config(text="受注件名: -")
            self.original_delivery_location_label.config(text="受渡場所名: -")
    
    def update_delivery_location_options(self):
        """受渡場所選択肢を更新"""
        current_data = self.get_current_search_data()
        if not current_data.empty:
            sample_row = current_data.iloc[0]
            
            # 受注.csvの「得意先」列の値をキーとして使用
            customer_key = sample_row.get('得意先', '')
            
            # デバッグ用：取得したキー値を表示
            print(f"受渡場所検索キー（得意先列の値）: {customer_key}")
            
            # 受渡場所リストを取得
            delivery_locations = get_delivery_locations_by_customer(customer_key)
            
            # デバッグ用：取得した受渡場所リストを表示
            print(f"取得した受渡場所リスト: {delivery_locations}")
            
            # 元の受渡場所を先頭に追加（重複を避ける）
            original_delivery = sample_row['受渡場所名']
            if original_delivery and original_delivery not in delivery_locations:
                delivery_locations.insert(0, original_delivery)
            
            # プルダウンの選択肢を更新
            self.delivery_location_combo['values'] = delivery_locations
            
            # 現在のページで選択されていた値を復元、またはデフォルト値を設定
            if self.current_page in self.selected_delivery_locations:
                selected_location = self.selected_delivery_locations[self.current_page]
                # 選択されていた値が現在の選択肢に含まれているかチェック
                if selected_location in delivery_locations:
                    self.delivery_location_var.set(selected_location)
                else:
                    # 含まれていない場合はデフォルト値を設定
                    self.delivery_location_var.set(original_delivery)
                    self.selected_delivery_locations[self.current_page] = original_delivery
            else:
                # デフォルトは元の受渡場所
                self.delivery_location_var.set(original_delivery)
                self.selected_delivery_locations[self.current_page] = original_delivery
                
            # 取得した受渡場所が0件の場合の処理
            if not delivery_locations:
                messagebox.showinfo("情報", f"得意先コード-得意先枝番「{customer_key}」に対応する受渡場所が見つかりませんでした。\n元の受渡場所のみ表示されます。")
        else:
            self.delivery_location_combo['values'] = []
            self.delivery_location_var.set("")
    
    def on_delivery_change_toggled(self):
        """受渡場所変更チェックボックスの状態変更時の処理"""
        if self.change_delivery_location.get():
            # チェックボックスがONの場合：ドロップダウンを表示
            self.delivery_location_label.pack(side=tk.LEFT, padx=10)
            self.delivery_location_combo.pack(side=tk.LEFT, padx=5)
        else:
            # チェックボックスがOFFの場合：ドロップダウンを非表示
            self.delivery_location_label.pack_forget()
            self.delivery_location_combo.pack_forget()
    
    def on_delivery_location_changed(self, event=None):
        """受渡場所選択変更時の処理"""
        selected_delivery = self.delivery_location_var.get()
        self.selected_delivery_locations[self.current_page] = selected_delivery
    
    def get_effective_delivery_location(self):
        """有効な受渡場所名を取得（変更されている場合は変更後、そうでなければ元の値）"""
        if self.change_delivery_location.get():
            # 受渡場所変更がONの場合は選択された値を使用
            return self.delivery_location_var.get()
        else:
            # OFFの場合は元の受渡場所を使用
            current_data = self.get_current_search_data()
            if not current_data.empty:
                return current_data.iloc[0]['受渡場所名']
            return ""
    
    def update_shipment_combo(self):
        """shipment_numberのプルダウンを更新"""
        current_data = self.get_current_search_data()
        if not current_data.empty:
            order_number = current_data.iloc[0]['受注番号']
            shipment_numbers = get_shipment_numbers_by_order(order_number)
            
            # プルダウンの選択肢を更新
            self.shipment_combo['values'] = [""] + shipment_numbers
            
            # 現在のページで選択されていた値を復元、または最大値をデフォルト設定
            if self.current_page in self.selected_shipment_numbers:
                self.shipment_var.set(self.selected_shipment_numbers[self.current_page])
            else:
                # 出荷依頼形式に応じて優先番号を決定
                if shipment_numbers:
                    default_shipment = self.get_default_shipment_number(shipment_numbers)
                    self.shipment_var.set(default_shipment)
                    # 選択された値を保存
                    self.selected_shipment_numbers[self.current_page] = default_shipment
                else:
                    self.shipment_var.set("")
        else:
            self.shipment_combo['values'] = [""]
            self.shipment_var.set("")
    
    def get_default_shipment_number(self, shipment_numbers):
        """出荷依頼形式に応じてデフォルトのshipment_numberを取得"""
        if self.shipment_request_type == "advanced_packing":
            # 事前梱包依頼の場合：「JK」から始まる番号を優先
            jk_numbers = [num for num in shipment_numbers if num.startswith("JK")]
            if jk_numbers:
                # JK番号の中で最大値を返す（昇順ソート済みなので最後の要素）
                return jk_numbers[-1]
        else:
            # 通常依頼の場合：「E」から始まる番号を優先
            e_numbers = [num for num in shipment_numbers if num.startswith("E")]
            if e_numbers:
                # E番号の中で最大値を返す（昇順ソート済みなので最後の要素）
                return e_numbers[-1]
        
        # 優先番号がない場合は全体の中で最大値（従来の動作）
        return shipment_numbers[-1]
    
    def on_shipment_selection_changed(self, event=None):
        """shipment_number選択変更時の処理"""
        selected_shipment = self.shipment_var.get()
        self.selected_shipment_numbers[self.current_page] = selected_shipment
    
    def show_packing_preview(self):
        """選択されたshipment_numberの梱包明細をプレビュー表示"""
        # ★修正：統合明細がある場合はそれを優先表示★
        if self.current_page in self.consolidated_packing_data_by_page:
            packing_data = self.consolidated_packing_data_by_page[self.current_page]
            selected_shipment = "統合明細"
        else:
            selected_shipment = self.shipment_var.get()
            if not selected_shipment:
                messagebox.showwarning("警告", "shipment_numberを選択してください。")
                return
            
            packing_data = get_packing_details_by_shipment(selected_shipment)
        
        if not packing_data:
            messagebox.showinfo("情報", "選択されたshipment_numberに対応する梱包明細が見つかりません。")
            return
        
        # プレビューダイアログを作成
        preview_dialog = Toplevel(self.dialog)
        preview_dialog.title(f"梱包明細プレビュー - {selected_shipment}")
        preview_dialog.geometry("1000x600")
        preview_dialog.transient(self.dialog)
        preview_dialog.grab_set()
        
        # スクロール可能なフレーム
        main_frame = tk.Frame(preview_dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        canvas = tk.Canvas(main_frame)
        scrollbar = tk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # ★修正：統合明細の場合のヘッダー表示★
        if selected_shipment == "統合明細":
            # 統合明細の場合は関連出荷指示番号を表示
            if packing_data and 'source_shipments' in packing_data[0]:
                info_label = tk.Label(scrollable_frame, 
                                    text=f"統合明細 - 関連出荷指示番号: {packing_data[0]['source_shipments']}", 
                                    font=("Arial", 12, "bold"), fg="red")
                info_label.pack(pady=(0, 10))
        
        # ヘッダー
        headers = ["ケース番号", "縦(cm)", "横(cm)", "高(cm)", "重量(kg)", "アイテム明細", "荷姿"]
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        col_widths = [12, 8, 8, 8, 10, 30, 15]
        for i, (header, width) in enumerate(zip(headers, col_widths)):
            tk.Label(header_frame, text=header, width=width, relief=tk.RIDGE, 
                    font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        
        # データ行
        for data in packing_data:
            data_frame = tk.Frame(scrollable_frame)
            data_frame.pack(fill=tk.X, pady=1)
            
            values = [
                data['case_number'],
                data['length'],
                data['width'],
                data['height'],
                data['weight'],
                data['item_details'],
                data['packing_style']
            ]
            
            for value, width in zip(values, col_widths):
                tk.Label(data_frame, text=str(value), width=width, relief=tk.RIDGE, 
                        font=("Arial", 10)).pack(side=tk.LEFT)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # ★追加：マウスホイールスクロール対応★
        bind_mousewheel_to_frame(canvas)
        
        # 閉じるボタン
        close_button = tk.Button(preview_dialog, text="閉じる", command=preview_dialog.destroy)
        close_button.pack(pady=10)
    
    def create_content(self):
        """現在のページのコンテンツを作成"""
        current_data = self.get_current_search_data()
        
        # スクロール可能なフレーム
        canvas = tk.Canvas(self.content_frame)
        scrollbar = tk.Scrollbar(self.content_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # ★追加：マウスホイールスクロール対応★
        bind_mousewheel_to_frame(canvas)
        
        # 修正：列幅を出荷依頼形式に応じて調整
        if self.shipment_request_type == "advanced_packing":
            # 事前梱包依頼の場合（従来通り）
            col_widths = {
                'ino': 8,
                'product_code': 15,
                'product_name': 25,
                'order_qty': 10,
                'self_stock_qty': 18,
                'previous_qty': 20,
                'remaining_qty': 15,
                'shipment_qty': 12
            }
        else:
            # 通常依頼の場合（事前梱包依頼済数量列を削除）
            col_widths = {
                'ino': 8,
                'product_code': 15,
                'product_name': 30,  # 少し幅を広げる
                'order_qty': 12,
                'self_stock_qty': 20,
                'remaining_qty': 18,
                'shipment_qty': 15
            }
        
        # 修正：ヘッダーを出荷依頼形式に応じて変更
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(header_frame, text="I/no.", width=col_widths['ino'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="商品コード", width=col_widths['product_code'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="商品名", width=col_widths['product_name'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="受注数量", width=col_widths['order_qty'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="入荷・在庫引当済", width=col_widths['self_stock_qty'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        
        # 条件分岐：事前梱包依頼の場合のみ「事前梱包依頼済数量」列を表示
        if self.shipment_request_type == "advanced_packing":
            tk.Label(header_frame, text="事前梱包依頼済数量", width=col_widths['previous_qty'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        
        tk.Label(header_frame, text="受注残数", width=col_widths['remaining_qty'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="今回出荷数", width=col_widths['shipment_qty'], relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        
        # 現在のページの入力エントリの辞書を初期化
        if not hasattr(self, 'entries_by_page'):
            self.entries_by_page = {}
        
        self.entries_by_page[self.current_page] = {}
        
        # データをI/no.順でソート
        sorted_data = current_data.sort_values('明細_共通項目2')
        
        # 各アイテムの入力行を作成
        for _, item in sorted_data.iterrows():
            item_frame = tk.Frame(scrollable_frame)
            item_frame.pack(fill=tk.X, pady=1)
            
            i_no = item['明細_共通項目2']
            order_number = item['受注番号']
            
            # 各種数量を取得
            order_quantity = item['明細_受注数量']
            self_stock_quantity = item.get('明細_自社在庫引当数量', 0)
            self_shipment_quantity = item.get('明細_自社出荷数量', 0)
            
            # 受注残数を計算
            remaining_quantity = order_quantity - self_shipment_quantity
            
            # 修正：今回出荷数のデフォルト値を出荷依頼形式に応じて変更
            if self.shipment_request_type == "advanced_packing":
                # 事前梱包依頼の場合（従来通り）
                previous_quantity = get_previous_packing_quantities(order_number, i_no)
                default_shipment = min(previous_quantity, remaining_quantity)
                if default_shipment < 0:
                    default_shipment = 0
            else:
                # 通常依頼の場合：入荷・在庫引当済の値（受注残数を上限とする）
                previous_quantity = 0  # 表示しないため0で初期化
                default_shipment = min(self_stock_quantity, remaining_quantity)
                if default_shipment < 0:
                    default_shipment = 0
            
            # 修正：列の表示を出荷依頼形式に応じて変更
            tk.Label(item_frame, text=str(i_no), width=col_widths['ino'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            tk.Label(item_frame, text=item['明細_商品コード'], width=col_widths['product_code'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            tk.Label(item_frame, text=item['明細_商品受注名'][:25 if self.shipment_request_type == "normal_request" else 23], width=col_widths['product_name'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            tk.Label(item_frame, text=str(order_quantity), width=col_widths['order_qty'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            tk.Label(item_frame, text=str(self_stock_quantity), width=col_widths['self_stock_qty'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            
            # 条件分岐：事前梱包依頼の場合のみ「事前梱包依頼済数量」列を表示
            if self.shipment_request_type == "advanced_packing":
                tk.Label(item_frame, text=str(previous_quantity), width=col_widths['previous_qty'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            
            tk.Label(item_frame, text=str(remaining_quantity), width=col_widths['remaining_qty'], relief=tk.RIDGE, font=("Arial", 10)).pack(side=tk.LEFT)
            
            # 今回出荷数の入力エントリ
            entry = tk.Entry(item_frame, width=col_widths['shipment_qty'], font=("Arial", 10))
            entry.pack(side=tk.LEFT, padx=5)
            entry.insert(0, str(default_shipment))
            
            self.entries_by_page[self.current_page][i_no] = {
                'entry': entry,
                'item_data': item,
                'previous_quantity': previous_quantity,
                'order_quantity': order_quantity,
                'self_stock_quantity': self_stock_quantity,
                'remaining_quantity': remaining_quantity
            }
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
    
    def get_shipment_quantities(self, page_index=None):
        """指定されたページの今回出荷数の値を取得（page_indexがNoneの場合は現在のページ）"""
        if page_index is None:
            page_index = self.current_page
            
        quantities = {}
        try:
            current_entries = self.entries_by_page.get(page_index, {})
            for i_no, data in current_entries.items():
                value = data['entry'].get().strip()
                entered_qty = int(value) if value else 0
                remaining_qty = data['remaining_quantity']
                
                # 受注残数を上限として制限
                final_qty = min(entered_qty, remaining_qty) if remaining_qty > 0 else 0
                quantities[i_no] = max(0, final_qty)
                
            return quantities
        except ValueError:
            messagebox.showerror("エラー", "数値を正しく入力してください。")
            return None
    
    def compare_quantities(self):
        """現在のページの自社在庫引当数量と過去依頼済数量を比較"""
        try:
            differences = []
            current_entries = self.entries_by_page.get(self.current_page, {})
            
            for i_no, data in current_entries.items():
                self_stock_qty = data['self_stock_quantity']
                previous_qty = data['previous_quantity']
                
                if self_stock_qty != previous_qty:
                    product_code = data['item_data']['明細_商品コード']
                    product_name = data['item_data']['明細_商品受注名']
                    
                    differences.append({
                        'i_no': i_no,
                        'product_code': product_code,
                        'product_name': product_name,
                        'self_stock_qty': self_stock_qty,
                        'previous_qty': previous_qty,
                        'difference': self_stock_qty - previous_qty
                    })
            
            if not differences:
                messagebox.showinfo("比較結果", "入荷・在庫引当済数量と事前梱包依頼済数量に差があるアイテムはありません。")
                return
            
            # 差異があるアイテムを表示するダイアログを作成
            self.show_differences_dialog(differences)
            
        except Exception as e:
            error_message = f"数量比較中にエラーが発生しました：{str(e)}"
            messagebox.showerror("エラー", error_message)
    
    def show_differences_dialog(self, differences):
        """差異があるアイテムを表示するダイアログ"""
        diff_dialog = Toplevel(self.dialog)
        diff_dialog.title("数量差異確認")
        diff_dialog.geometry("1100x500")
        diff_dialog.transient(self.dialog)
        diff_dialog.grab_set()
        
        # メインフレーム
        main_frame = tk.Frame(diff_dialog)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # ヘッダー部
        header_section = tk.Frame(main_frame)
        header_section.pack(fill=tk.X, pady=(0, 10))
        
        # タイトル
        title_label = tk.Label(header_section, text="入荷・在庫引当済数量と事前梱包依頼済数量に差があるアイテム", font=("Arial", 12, "bold"))
        title_label.pack()
        
        # 色の説明
        info_label = tk.Label(header_section, text="青色：入荷・在庫引当済数量が多い　赤色：事前梱包依頼済数量が多い", font=("Arial", 10))
        info_label.pack(pady=(5, 0))
        
        # 明細部（スクロール可能）
        detail_section = tk.Frame(main_frame)
        detail_section.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # スクロール可能なフレーム
        canvas = tk.Canvas(detail_section)
        scrollbar = tk.Scrollbar(detail_section, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # ★追加：マウスホイールスクロール対応★
        bind_mousewheel_to_frame(canvas)
        
        # 明細ヘッダー
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(header_frame, text="I/no.", width=8, relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="商品コード", width=15, relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="商品名", width=25, relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="入荷・在庫引当済", width=18, relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="事前梱包依頼済数量", width=20, relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        tk.Label(header_frame, text="差異", width=8, relief=tk.RIDGE, font=("Arial", 10, "bold")).pack(side=tk.LEFT)
        
        # 差異データの表示
        for diff in differences:
            item_frame = tk.Frame(scrollable_frame)
            item_frame.pack(fill=tk.X, pady=1)
            
            # 差異がプラスかマイナスかで色を変更
            bg_color = "lightblue" if diff['difference'] > 0 else "lightcoral"
            
            tk.Label(item_frame, text=str(diff['i_no']), width=8, relief=tk.RIDGE, font=("Arial", 10), bg=bg_color).pack(side=tk.LEFT)
            tk.Label(item_frame, text=diff['product_code'], width=15, relief=tk.RIDGE, font=("Arial", 10), bg=bg_color).pack(side=tk.LEFT)
            tk.Label(item_frame, text=diff['product_name'][:23], width=25, relief=tk.RIDGE, font=("Arial", 10), bg=bg_color).pack(side=tk.LEFT)
            tk.Label(item_frame, text=str(diff['self_stock_qty']), width=18, relief=tk.RIDGE, font=("Arial", 10), bg=bg_color).pack(side=tk.LEFT)
            tk.Label(item_frame, text=str(diff['previous_qty']), width=20, relief=tk.RIDGE, font=("Arial", 10), bg=bg_color).pack(side=tk.LEFT)
            tk.Label(item_frame, text=f"{diff['difference']:+d}", width=8, relief=tk.RIDGE, font=("Arial", 10, "bold"), bg=bg_color).pack(side=tk.LEFT)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # フッター部（ボタン）
        footer_section = tk.Frame(main_frame)
        footer_section.pack(fill=tk.X)
        
        # 閉じるボタン
        close_button = tk.Button(footer_section, text="閉じる", command=diff_dialog.destroy)
        close_button.pack(pady=5)
        
        # ダイアログを中央に配置
        diff_dialog.update_idletasks()
        x = (diff_dialog.winfo_screenwidth() // 2) - (diff_dialog.winfo_width() // 2)
        y = (diff_dialog.winfo_screenheight() // 2) - (diff_dialog.winfo_height() // 2)
        diff_dialog.geometry(f"+{x}+{y}")

    def output_shipment_report(self):
        """現在のページの出荷報告書をExcel形式で出力（梱包明細付き・受渡場所変更対応）"""
        try:
            # 今回出荷数を取得
            shipment_quantities = self.get_shipment_quantities()
            if shipment_quantities is None:
                return
            
            # 選択された出荷日を取得
            shipment_date = self.shipment_date_cal.get_date()
            
            # 現在のページのデータを取得
            current_data = self.get_current_search_data()
            current_key = self.get_current_search_key()
            
            # ★修正：統合された梱包明細があるかチェック★
            if self.current_page in self.consolidated_packing_data_by_page:
                # 統合設定がある場合は統合された明細を使用
                packing_data = self.consolidated_packing_data_by_page[self.current_page]
                selected_shipment = "統合明細"
                print(f"統合明細を使用: {len(packing_data)}件のケース")
            else:
                # 従来通りの処理
                selected_shipment = self.shipment_var.get()
                packing_data = []
                if selected_shipment:
                    packing_data = get_packing_details_by_shipment(selected_shipment)
                print(f"通常明細を使用: selected_shipment={selected_shipment}, packing_data件数={len(packing_data)}")
            
            # Excelファイルを作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "出荷報告書"
            
            # ヘッダー部分
            ws['A1'] = "出荷報告書"
            ws['A1'].font = Font(name='メイリオ', size=22, bold=True)
            ws['A1'].alignment = Alignment(horizontal='left')
            
            # G2セルに「東邦ヤンマーテック株式会社」を右揃えで配置
            ws['G2'] = "東邦ヤンマーテック株式会社"
            ws['G2'].font = Font(name='メイリオ', size=12, bold=True)
            ws['G2'].alignment = Alignment(horizontal='right')
            
            # G3セルに「出荷日:YYYY/MM/DD」を右揃えで配置
            ws['G3'] = f"出荷日:{shipment_date.strftime('%Y/%m/%d')}"
            ws['G3'].font = Font(name='メイリオ', size=12, bold=True)
            ws['G3'].alignment = Alignment(horizontal='right')
            
            # サンプルデータから情報を取得
            sample_row = current_data.iloc[0]
            
            # 修正：有効な受渡場所名を使用
            delivery_location = self.get_effective_delivery_location()
            if not delivery_location or str(delivery_location).strip() == '' or pd.isna(delivery_location):
                delivery_location = "同上"
            
            # ヘッダー項目を配置
            ws['A2'] = f"得意先名: {sample_row['得意先名']}様"
            ws['A3'] = f"船名: {sample_row['受注件名']}"
            ws['A4'] = f"客注番号: {sample_row['客注番号']}"
            ws['A5'] = f"受渡場所名: {delivery_location}"  # 修正：選択された受渡場所を使用
            ws['A6'] = f"見積番号: {sample_row['明細_共通項目3']}"
            ws['A7'] = f"受注番号: {sample_row['受注番号']}"
            
            # 発注番号の処理
            order_numbers = current_data['発注番号'].unique()
            order_numbers = [o for o in order_numbers if o]
            if len(order_numbers) > 1:
                order_numbers_str = ', '.join(order_numbers)
            else:
                order_numbers_str = order_numbers[0] if order_numbers else "発注番号なし"
            ws['A8'] = f"発注番号: {order_numbers_str}"
            
            ws['A9'] = f"営業担当者: {sample_row['社員名']}"
            
            # ヘッダーのスタイル設定
            for row_cells in ws['A1:A9']:
                for c in row_cells:
                    c.alignment = Alignment(horizontal='left')
                    if c.row != 1:
                        c.font = Font(name='メイリオ', size=12, bold=True)
            
            # 明細部分のヘッダー
            headers = ["I/no.", "商品コード", "商品受注名", "受注数量", "受注残数", "今回出荷数", "備考"]
            
            header_fill = PatternFill(patternType="solid", fgColor="D9D9D9")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 明細ヘッダーの行番号を11行目に設定
            for col_idx, head_text in enumerate(headers, start=1):
                cell = ws.cell(row=11, column=col_idx, value=head_text)
                cell.font = Font(name='メイリオ', bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                cell.fill = header_fill
            
            # 列幅設定
            col_width_map = {
                'A': 72,   # I/no.
                'B': 136,  # 商品コード
                'C': 308,  # 商品受注名
                'D': 78,   # 受注数量
                'E': 78,   # 受注残数
                'F': 95,   # 今回出荷数
                'G': 113   # 備考
            }
            
            def px_to_excel_col_width(px): 
                return px / 7.0
            
            for col_letter, px_val in col_width_map.items():
                ws.column_dimensions[col_letter].width = px_to_excel_col_width(px_val)
            
            # 明細データの出力
            row_start = 12
            sorted_data = current_data.sort_values('明細_共通項目2')
            
            for row_index, (_, data) in enumerate(sorted_data.iterrows(), start=row_start):
                i_no = data['明細_共通項目2']
                shipment_qty = shipment_quantities.get(i_no, 0)
                
                # 受注残数を計算
                order_quantity = data['明細_受注数量']
                self_shipment_quantity = data.get('明細_自社出荷数量', 0)
                remaining_quantity = order_quantity - self_shipment_quantity
                
                # 備考の設定
                remark_text = ""
                if shipment_qty <= 0:
                    remark_text = "出荷なし"
                
                ws.cell(row=row_index, column=1, value=i_no)
                ws.cell(row=row_index, column=2, value=data['明細_商品コード'])
                ws.cell(row=row_index, column=3, value=data['明細_商品受注名'])
                ws.cell(row=row_index, column=4, value=order_quantity)
                ws.cell(row=row_index, column=5, value=remaining_quantity)
                ws.cell(row=row_index, column=6, value=shipment_qty)
                ws.cell(row=row_index, column=7, value=remark_text)
                
                # 行の高さ設定
                ws.row_dimensions[row_index].height = 25
            
            # 明細部分の罫線とスタイル設定
            max_row_used = ws.max_row
            for row_idx in range(12, max_row_used + 1):
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    cell.font = Font(name='メイリオ', size=12)
                    if col_idx in [2, 3]:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # ここから梱包明細をフッター部分に追加
            if packing_data:
                # 梱包明細開始行を計算（明細の後に3行空ける）
                packing_start_row = max_row_used + 4
                
                # 梱包明細のタイトル
                ws.cell(row=packing_start_row, column=1, value="梱包明細")
                ws.cell(row=packing_start_row, column=1).font = Font(name='メイリオ', size=14, bold=True)
                
                # ★修正：統合明細の場合は関連出荷指示番号を表示★
                if selected_shipment == "統合明細":
                    if packing_data and 'source_shipments' in packing_data[0]:
                        ws.cell(row=packing_start_row + 1, column=1, value=f"関連出荷指示番号: {packing_data[0]['source_shipments']}")
                    else:
                        ws.cell(row=packing_start_row + 1, column=1, value="統合された梱包明細")
                else:
                    ws.cell(row=packing_start_row + 1, column=1, value=f"出荷指示番号: {selected_shipment}")
                
                ws.cell(row=packing_start_row + 1, column=1).font = Font(name='メイリオ', size=12, bold=True)
                
                # 梱包明細のヘッダー情報
                packing_info_start_row = packing_start_row + 3
                
                # 梱包明細用ヘッダー情報
                packing_header_info = [
                    f"得意先名: {sample_row['得意先名']}様",
                    f"船名: {sample_row['受注件名']}",
                    f"客注番号: {sample_row['客注番号']}",
                    f"受渡場所名: {delivery_location}",  # 修正：選択された受渡場所を使用
                    f"見積番号: {sample_row['明細_共通項目3']}",
                    f"受注番号: {sample_row['受注番号']}",
                    f"発注番号: {order_numbers_str}",
                    f"営業担当者: {sample_row['社員名']}"
                ]
                
                # ヘッダー情報を配置
                for i, info in enumerate(packing_header_info):
                    ws.cell(row=packing_info_start_row + i, column=1, value=info)
                    ws.cell(row=packing_info_start_row + i, column=1).font = Font(name='メイリオ', size=10, bold=True)
                
                # 梱包明細のテーブルヘッダー
                packing_headers = ["ケース番号", "縦(cm)", "横(cm)", "高(cm)", "重量(kg)", "アイテム明細", "荷姿"]
                packing_header_row = packing_info_start_row + len(packing_header_info) + 2
                
                for col_idx, header in enumerate(packing_headers, start=1):
                    cell = ws.cell(row=packing_header_row, column=col_idx, value=header)
                    cell.font = Font(name='メイリオ', bold=True, size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    cell.fill = header_fill
                
                # 修正：梱包明細データを自然順序でソート
                import re
                def natural_sort_key(item):
                    """ケース番号を自然順序でソートするためのキー関数"""
                    case_number = str(item['case_number'])
                    # 数値部分を抽出してソートキーとする
                    numbers = re.findall(r'\d+', case_number)
                    if numbers:
                        return int(numbers[0])
                    return 0
                
                # ケース番号で自然順序ソート
                sorted_packing_data = sorted(packing_data, key=natural_sort_key)
                
                # 梱包明細データ
                for data_idx, data in enumerate(sorted_packing_data):
                    data_row = packing_header_row + 1 + data_idx
                    
                    # データを配置
                    ws.cell(row=data_row, column=1, value=data['case_number'])
                    ws.cell(row=data_row, column=2, value=data['length'])
                    ws.cell(row=data_row, column=3, value=data['width'])
                    ws.cell(row=data_row, column=4, value=data['height'])
                    ws.cell(row=data_row, column=5, value=data['weight'])
                    ws.cell(row=data_row, column=6, value=data['item_details'])
                    ws.cell(row=data_row, column=7, value=data['packing_style'])
                    
                    # 行の高さ設定
                    ws.row_dimensions[data_row].height = 20
                
                # 梱包明細部分の罫線とスタイル設定
                packing_end_row = packing_header_row + len(sorted_packing_data)
                for row_idx in range(packing_header_row + 1, packing_end_row + 1):
                    for col_idx in range(1, 8):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.font = Font(name='メイリオ', size=10)
                        if col_idx == 6:  # アイテム明細列は左揃え
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # ドキュメントフォルダに保存
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")

            # 修正：Excelファイル名も[YYMMDD]_[客注番号]_出荷報告書.xlsx形式に変更
            # 出荷日から年月日（YYMMDD）を取得
            shipment_date_str = shipment_date.strftime("%y%m%d")

            # 客注番号を取得
            customer_order_no = current_data.iloc[0]['客注番号']
            # 客注番号から無効な文字を除去（ファイル名に使用できない文字対策）
            safe_customer_order_no = "".join(c for c in customer_order_no if c.isalnum() or c in "-_")

            # ファイル名を生成
            filename = f"{shipment_date_str}_{safe_customer_order_no}_出荷報告書.xlsx"
            output_path = os.path.join(documents_path, filename)

            print(f"Excel出力ファイル名: {filename}")  # デバッグ
            
            wb.save(output_path)
            
            completion_message = f"Excel出荷報告書が出力されました。\n保存先: {output_path}"
            if packing_data:
                if selected_shipment == "統合明細":
                    completion_message += f"\n\n統合された梱包明細も含まれています"
                else:
                    completion_message += f"\n\n梱包明細も含まれています（出荷指示番号: {selected_shipment}）"
            
            messagebox.showinfo("完了", completion_message)
            
            # ファイルを開くかどうか確認
            if messagebox.askyesno("確認", "Excel出荷報告書を開きますか？"):
                try:
                    subprocess.Popen(["start", "", output_path], shell=True)
                except Exception as e:
                    messagebox.showerror("エラー", f"ファイルを開く際にエラーが発生しました:\n{str(e)}")
                    
        except Exception as e:
            error_message = f"Excel出荷報告書出力中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
            messagebox.showerror("エラー", error_message)
            with open("error_log.txt", "w", encoding='utf-8') as f:
                f.write(error_message)

    def register_japanese_font(self):
        """日本語フォントを登録"""
        try:
            # システムに応じたフォントパスを試行
            font_paths = []
            
            system = platform.system()
            if system == "Windows":
                # Windowsの場合
                font_paths = [
                    "C:/Windows/Fonts/msgothic.ttc",  # MS Gothic
                    "C:/Windows/Fonts/msmincho.ttc",  # MS Mincho
                    "C:/Windows/Fonts/NotoSansCJK-Regular.ttc",  # Noto Sans CJK
                    "C:/Windows/Fonts/meiryo.ttc",    # Meiryo
                ]
            elif system == "Darwin":  # macOS
                font_paths = [
                    "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
                    "/System/Library/Fonts/Hiragino Sans GB.ttc",
                    "/Library/Fonts/NotoSansCJK-Regular.ttc",
                ]
            else:  # Linux
                font_paths = [
                    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
                    "/usr/share/fonts/truetype/takao-gothic/TakaoPGothic.ttf",
                    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
                ]
            
            # フォントファイルを探して登録
            font_registered = False
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        # 通常フォント
                        pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
                        # 太字フォント（同じフォントファイルを使用）
                        pdfmetrics.registerFont(TTFont('JapaneseFont-Bold', font_path))
                        font_registered = True
                        print(f"日本語フォントを登録しました: {font_path}")
                        break
                    except Exception as e:
                        print(f"フォント登録失敗 {font_path}: {e}")
                        continue
            
            if not font_registered:
                # フォールバック: 利用可能なフォントを探す
                print("警告: 標準的な日本語フォントが見つかりません。代替フォントを探しています...")
                
                # Windowsで他のフォントも試す
                if system == "Windows":
                    additional_fonts = [
                        "C:/Windows/Fonts/yugothm.ttc",  # Yu Gothic Medium
                        "C:/Windows/Fonts/yugothb.ttc",  # Yu Gothic Bold
                        "C:/Windows/Fonts/malgun.ttf",   # Malgun Gothic
                    ]
                    for font_path in additional_fonts:
                        if os.path.exists(font_path):
                            try:
                                pdfmetrics.registerFont(TTFont('JapaneseFont', font_path))
                                pdfmetrics.registerFont(TTFont('JapaneseFont-Bold', font_path))
                                font_registered = True
                                print(f"代替日本語フォントを登録しました: {font_path}")
                                break
                            except Exception as e:
                                continue
                
                if not font_registered:
                    # 最終手段: Helveticaを日本語フォントとしてエイリアス
                    print("最終手段: Helveticaフォントを使用します（日本語は表示されない可能性があります）")
                    # この場合、フォント名だけ変更して処理を続行
        
        except Exception as e:
            print(f"フォント登録でエラーが発生しました: {e}")

    def create_pdf_report(self, data, search_key, shipment_quantities, shipment_date, selected_shipment="", packing_data=None, custom_delivery_location=None):
        """単一のPDF出荷報告書を作成（日本語対応・梱包明細対応・受渡場所変更対応）"""
        try:
            # 日本語フォントを登録
            self.register_japanese_font()
            
            # メモリ上にPDFを作成
            buffer = io.BytesIO()
            
            # サンプルデータから情報を取得
            sample_row = data.iloc[0]
            
            # 修正：受渡場所名の処理（カスタム受渡場所が指定されている場合はそれを使用）
            if custom_delivery_location:
                delivery_location = custom_delivery_location
            else:
                delivery_location = sample_row['受渡場所名']
            
            if not delivery_location or str(delivery_location).strip() == '' or pd.isna(delivery_location):
                delivery_location = "同上"
            
            # 発注番号の処理
            order_numbers = data['発注番号'].unique()
            order_numbers = [o for o in order_numbers if o]
            if len(order_numbers) > 1:
                order_numbers_str = ', '.join(order_numbers)
            else:
                order_numbers_str = order_numbers[0] if order_numbers else "発注番号なし"
            
            # ドキュメント要素のリスト
            elements = []
            
            # スタイルを設定（日本語フォント対応）
            styles = getSampleStyleSheet()
            
            # タイトルスタイル
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                fontName='JapaneseFont-Bold',
                alignment=1,  # 中央揃え
                spaceAfter=20
            )
            
            # 通常テキストスタイル
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                fontName='JapaneseFont',
                leftIndent=0
            )
            
            # 右揃えスタイル
            right_style = ParagraphStyle(
                'CustomRight',
                parent=styles['Normal'],
                fontSize=10,
                fontName='JapaneseFont-Bold',
                alignment=2  # 右揃え
            )
            
            # タイトル
            title = Paragraph("出荷報告書", title_style)
            elements.append(title)
            
            # 右上の会社名と出荷日
            company_info = [
                ["", "東邦ヤンマーテック株式会社"],
                ["", f"出荷日:{shipment_date.strftime('%Y/%m/%d')}"]
            ]
            
            company_table = Table(company_info, colWidths=[400, 150])
            company_table.setStyle(TableStyle([
                ('ALIGN', (1, 0), (1, 1), 'RIGHT'),
                ('FONTNAME', (1, 0), (1, 1), 'JapaneseFont-Bold'),
                ('FONTSIZE', (1, 0), (1, 1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(company_table)
            elements.append(Spacer(1, 10))
            
            # ヘッダー情報
            header_info = [
                [f"得意先名: {sample_row['得意先名']}様"],
                [f"船名: {sample_row['受注件名']}"],
                [f"客注番号: {sample_row['客注番号']}"],
                [f"受渡場所名: {delivery_location}"],  # 修正：選択された受渡場所を使用
                [f"見積番号: {sample_row['明細_共通項目3']}"],
                [f"受注番号: {sample_row['受注番号']}"],
                [f"発注番号: {order_numbers_str}"],
                [f"営業担当者: {sample_row['社員名']}"]
            ]
            
            header_table = Table(header_info, colWidths=[550])
            header_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'JapaneseFont-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ]))
            elements.append(header_table)
            elements.append(Spacer(1, 15))
            
            # 明細テーブルのヘッダー
            table_headers = ["I/no.", "商品コード", "商品受注名", "受注数量", "受注残数", "今回出荷数", "備考"]
            
            # 明細データの準備
            table_data = [table_headers]
            sorted_data = data.sort_values('明細_共通項目2')
            
            for _, row_data in sorted_data.iterrows():
                i_no = row_data['明細_共通項目2']
                shipment_qty = shipment_quantities.get(i_no, 0)
                
                # 受注残数を計算
                order_quantity = row_data['明細_受注数量']
                self_shipment_quantity = row_data.get('明細_自社出荷数量', 0)
                remaining_quantity = order_quantity - self_shipment_quantity
                
                # 備考の設定
                remark_text = ""
                if shipment_qty == 0:
                    # 今回出荷数が0の場合：出荷なし
                    remark_text = "出荷なし"
                elif order_quantity == shipment_qty:
                    # 受注数量=今回出荷数の場合：空欄
                    remark_text = ""
                elif order_quantity == remaining_quantity and shipment_qty != 0 and shipment_qty < order_quantity:
                    # 受注数量=受注残数 かつ 今回出荷数が0ではない、かつ受注数未満の場合：分納
                    remark_text = "分納"
                elif order_quantity > remaining_quantity and shipment_qty != 0:
                    # 受注数量＞受注残数 かつ 今回出荷数が0ではない場合：欠品残
                    remark_text = "欠品残"
                
                # 商品受注名を適切な長さに調整
                product_name = str(row_data['明細_商品受注名'])[:25]  # 日本語対応のため少し長めに
                
                table_data.append([
                    str(i_no),
                    str(row_data['明細_商品コード']),
                    product_name,
                    str(order_quantity),
                    str(remaining_quantity),
                    str(shipment_qty),
                    remark_text
                ])
            
            # A4に収まる列幅を計算（合計550pt）
            col_widths = [50, 80, 180, 60, 60, 60, 60]
            
            # テーブルを作成
            detail_table = Table(table_data, colWidths=col_widths)
            detail_table.setStyle(TableStyle([
                # ヘッダー行
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'JapaneseFont-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # データ行
                ('FONTNAME', (0, 1), (-1, -1), 'JapaneseFont'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # I/no.
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),    # 商品コード
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # 商品受注名
                ('ALIGN', (3, 1), (-1, -1), 'CENTER'), # 数量関連
                
                # 罫線
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
            ]))
            
            elements.append(detail_table)
            
            # ★デバッグ: 梱包明細データの確認★
            print(f"PDF作成時の梱包明細チェック:")
            print(f"  selected_shipment: {selected_shipment}")
            print(f"  packing_data: {packing_data}")
            print(f"  packing_data件数: {len(packing_data) if packing_data else 0}")

            # 梱包明細セクションを追加（統合表示対応）
            if packing_data and len(packing_data) > 0:
                print(f"梱包明細セクション作成開始: {len(packing_data)}件")  # デバッグ
                
                try:
                    # 梱包明細セクションの要素を準備
                    packing_elements = []
                    
                    # 梱包明細タイトル
                    packing_title_style = ParagraphStyle(
                        'PackingTitle',
                        parent=styles['Heading2'],
                        fontSize=14,
                        fontName='JapaneseFont-Bold',
                        spaceAfter=10
                    )
                    
                    # 修正：統合した場合でもタイトルは「梱包明細」に統一
                    packing_title = Paragraph("梱包明細", packing_title_style)
                    if selected_shipment == "統合明細":
                        print("統合梱包明細タイトル作成（タイトルは「梱包明細」）")  # デバッグ
                    else:
                        print("通常梱包明細タイトル作成")  # デバッグ
                    
                    packing_elements.append(packing_title)
                    
                    # ★修正：統合明細時は関連出荷指示番号を表示★
                    if selected_shipment == "統合明細":
                        # 統合明細の場合は関連出荷指示番号を表示
                        if packing_data and 'source_shipments' in packing_data[0]:
                            shipment_info = Paragraph(f"関連出荷指示番号: {packing_data[0]['source_shipments']}", normal_style)
                            print(f"関連出荷指示番号: {packing_data[0]['source_shipments']}")  # デバッグ
                        else:
                            shipment_info = Paragraph("統合された梱包明細", normal_style)
                            print("統合された梱包明細（関連番号なし）")  # デバッグ
                    else:
                        # 従来通りの表示
                        shipment_info = Paragraph(f"出荷指示番号: {selected_shipment}", normal_style)
                        print(f"出荷指示番号: {selected_shipment}")  # デバッグ
                    
                    packing_elements.append(shipment_info)
                    packing_elements.append(Spacer(1, 10))
                    
                    # 追加：梱包明細のヘッダー情報
                    print("ヘッダー情報作成開始")  # デバッグ
                    
                    # 梱包明細用ヘッダー情報
                    packing_header_info = [
                        [f"得意先名: {sample_row['得意先名']}様"],
                        [f"船名: {sample_row['受注件名']}"],
                        [f"客注番号: {sample_row['客注番号']}"],
                        [f"受渡場所名: {delivery_location}"],
                        [f"見積番号: {sample_row['明細_共通項目3']}"],
                        [f"受注番号: {sample_row['受注番号']}"],
                        [f"発注番号: {order_numbers_str}"],
                        [f"営業担当者: {sample_row['社員名']}"]
                    ]
                    
                    packing_header_table = Table(packing_header_info, colWidths=[550])
                    packing_header_table.setStyle(TableStyle([
                        ('FONTNAME', (0, 0), (-1, -1), 'JapaneseFont-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 10),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ]))
                    packing_elements.append(packing_header_table)
                    packing_elements.append(Spacer(1, 15))
                    
                    print("ヘッダー情報作成完了")  # デバッグ
                    
                    # 梱包明細テーブルのヘッダー（統合/通常問わず「アイテム明細」で固定）
                    packing_headers = ["ケース番号", "縦(cm)", "横(cm)", "高(cm)", "重量(kg)", "アイテム明細", "荷姿"]
                    if selected_shipment == "統合明細":
                        print("統合明細用ヘッダー設定（アイテム明細で固定）")  # デバッグ
                    else:
                        print("通常明細用ヘッダー設定")  # デバッグ
                    
                    # 修正：梱包明細データを自然順序でソート
                    import re
                    def natural_sort_key(item):
                        """ケース番号を自然順序でソートするためのキー関数"""
                        case_number = str(item['case_number'])
                        numbers = re.findall(r'\d+', case_number)
                        if numbers:
                            return int(numbers[0])
                        return 0
                    
                    # ケース番号で自然順序ソート
                    sorted_packing_data = sorted(packing_data, key=natural_sort_key)
                    print(f"ソート後のデータ: {len(sorted_packing_data)}件")  # デバッグ
                    
                    # 梱包明細データの準備
                    packing_table_data = [packing_headers]
                    
                    for i, packing_item in enumerate(sorted_packing_data):
                        print(f"梱包明細データ[{i}]処理: {packing_item}")  # デバッグ
                        
                        # 修正：統合されたアイテム詳細を表示
                        item_details = str(packing_item['item_details'])
                        
                        row_data = [
                            str(packing_item['case_number']),
                            str(packing_item['length']),
                            str(packing_item['width']),
                            str(packing_item['height']),
                            str(packing_item['weight']),
                            item_details,  # 統合された範囲を表示
                            str(packing_item['packing_style'])
                        ]
                        
                        packing_table_data.append(row_data)
                        print(f"テーブル行データ: {row_data}")  # デバッグ
                    
                    print(f"テーブルデータ準備完了: {len(packing_table_data)}行")  # デバッグ
                    
                    # 梱包明細テーブルの列幅（合計550pt）
                    packing_col_widths = [70, 50, 50, 50, 60, 150, 120]
                    
                    # 梱包明細テーブルを作成
                    print("テーブル作成開始")  # デバッグ
                    packing_table = Table(packing_table_data, colWidths=packing_col_widths)
                    packing_table.setStyle(TableStyle([
                        # ヘッダー行
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'JapaneseFont-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        
                        # データ行
                        ('FONTNAME', (0, 1), (-1, -1), 'JapaneseFont'),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('ALIGN', (5, 1), (5, -1), 'LEFT'),    # アイテム明細は左揃え
                        ('ALIGN', (6, 1), (6, -1), 'LEFT'),    # 荷姿は左揃え
                        
                        # 罫線
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.lightyellow, colors.white]),
                    ]))
                    
                    packing_elements.append(packing_table)
                    print("テーブル作成完了")  # デバッグ
                    
                    # 改ページ制御：梱包明細セクション全体をまとめて扱う
                    from reportlab.platypus import KeepTogether
                    
                    # 改ページを追加してから梱包明細セクションを配置
                    elements.append(PageBreak())
                    print("改ページ追加")  # デバッグ
                    
                    # 梱包明細セクション全体をKeepTogetherでラップ
                    packing_section = KeepTogether(packing_elements)
                    elements.append(packing_section)
                    print("梱包明細セクション追加完了")  # デバッグ
                    
                except Exception as e:
                    print(f"梱包明細セクション作成エラー: {str(e)}")  # デバッグ
                    import traceback
                    traceback.print_exc()
                    
            else:
                print("梱包明細なし - 条件不一致")  # デバッグ
                print(f"packing_data: {packing_data}")  # デバッグ
                print(f"packing_data length: {len(packing_data) if packing_data else 'None'}")  # デバッグ
            
            # PDFドキュメントを作成
            print(f"PDF構築開始: elements数 = {len(elements)}")  # デバッグ
            for i, element in enumerate(elements):
                print(f"Element[{i}]: {type(element).__name__}")  # デバッグ

            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=20,
                leftMargin=20,
                topMargin=20,
                bottomMargin=20
            )

            # PDFを構築
            print("PDF構築実行")  # デバッグ
            doc.build(elements)
            print("PDF構築完了")  # デバッグ
            
            # バッファの内容を取得
            pdf_content = buffer.getvalue()
            buffer.close()
            
            return pdf_content
            
        except Exception as e:
            raise Exception(f"PDF作成中にエラーが発生しました: {str(e)}")
    
    def output_shipment_report_pdf(self):
        """現在のページの出荷報告書をPDF形式で出力（梱包明細付き・受渡場所変更対応）"""
        try:
            # 今回出荷数を取得
            shipment_quantities = self.get_shipment_quantities()
            if shipment_quantities is None:
                return
            
            # 選択された出荷日を取得
            shipment_date = self.shipment_date_cal.get_date()
            
            # 現在のページのデータを取得
            current_data = self.get_current_search_data()
            current_key = self.get_current_search_key()
            
            # ★修正：統合された梱包明細があるかチェック★
            if self.current_page in self.consolidated_packing_data_by_page:
                # 統合設定がある場合は統合された明細を使用
                packing_data = self.consolidated_packing_data_by_page[self.current_page]
                selected_shipment = "統合明細"
                print(f"統合明細を使用: {len(packing_data)}件のケース")  # デバッグ
                for item in packing_data:
                    print(f"  ケース: {item['case_number']}, 明細: {item['item_details']}")  # デバッグ
            else:
                # 従来通りの処理
                selected_shipment = self.shipment_var.get()
                packing_data = []
                if selected_shipment:
                    packing_data = get_packing_details_by_shipment(selected_shipment)
                print(f"通常明細を使用: selected_shipment={selected_shipment}, packing_data件数={len(packing_data)}")  # デバッグ
            
            # 修正：有効な受渡場所を取得
            custom_delivery_location = self.get_effective_delivery_location()
            
            # PDFを作成
            pdf_content = self.create_pdf_report(current_data, current_key, shipment_quantities, 
                                               shipment_date, selected_shipment, packing_data, 
                                               custom_delivery_location)
            
            # ドキュメントフォルダに保存
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")

            # 修正：ファイル名を[YYMMDD]_[客注番号]_出荷報告書.pdf形式に変更
            # 出荷日から年月日（YYMMDD）を取得
            shipment_date_str = shipment_date.strftime("%y%m%d")

            # 客注番号を取得
            customer_order_no = current_data.iloc[0]['客注番号']
            # 客注番号から無効な文字を除去（ファイル名に使用できない文字対策）
            safe_customer_order_no = "".join(c for c in customer_order_no if c.isalnum() or c in "-_")

            # ファイル名を生成
            filename = f"{shipment_date_str}_{safe_customer_order_no}_出荷報告書.pdf"
            output_path = os.path.join(documents_path, filename)

            print(f"生成されたファイル名: {filename}")  # デバッグ
            
            with open(output_path, 'wb') as f:
                f.write(pdf_content)
            
            completion_message = f"PDF出荷報告書が出力されました。\n保存先: {output_path}"
            if packing_data:
                if selected_shipment == "統合明細":
                    completion_message += f"\n\n統合された梱包明細も含まれています"
                else:
                    completion_message += f"\n\n梱包明細も含まれています（Shipment Number: {selected_shipment}）"
            
            messagebox.showinfo("完了", completion_message)
            
            # ファイルを開くかどうか確認
            if messagebox.askyesno("確認", "PDF出荷報告書を開きますか？"):
                try:
                    subprocess.Popen(["start", "", output_path], shell=True)
                except Exception as e:
                    messagebox.showerror("エラー", f"ファイルを開く際にエラーが発生しました:\n{str(e)}")
                    
        except Exception as e:
            error_message = f"PDF出荷報告書出力中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
            messagebox.showerror("エラー", error_message)
            with open("error_log.txt", "w", encoding='utf-8') as f:
                f.write(error_message)

    def output_all_shipment_reports_pdf(self):
        """全ページの出荷報告書を一括でPDF出力し、1つのファイルに結合（梱包明細付き・受渡場所変更対応・統合明細対応）"""
        try:
            # 選択された出荷日を取得
            shipment_date = self.shipment_date_cal.get_date()
            
            # 確認メッセージ
            if not messagebox.askyesno("確認", f"{self.total_pages}件のPDF出荷報告書を一括出力しますか？"):
                return
            
            # 各ページのPDFを作成してリストに保存
            pdf_list = []
            
            for page_index in range(self.total_pages):
                try:
                    # 該当ページのデータを取得
                    page_data = self.search_data_list[page_index]
                    page_key = self.search_keys[page_index]
                    
                    # 該当ページの出荷数量を取得（現在のページでない場合はデフォルト値を使用）
                    if page_index == self.current_page:
                        page_shipment_quantities = self.get_shipment_quantities(page_index)
                        if page_shipment_quantities is None:
                            page_shipment_quantities = {}
                    else:
                        # 他のページはデフォルト値を使用
                        page_shipment_quantities = {}
                        sorted_data = page_data.sort_values('明細_共通項目2')
                        for _, item in sorted_data.iterrows():
                            i_no = item['明細_共通項目2']
                            order_number = item['受注番号']
                            
                            # 修正：今回出荷数のデフォルト値を出荷依頼形式に応じて変更
                            if self.shipment_request_type == "advanced_packing":
                                # 事前梱包依頼の場合（従来通り）
                                previous_quantity = get_previous_packing_quantities(order_number, i_no)
                                order_quantity = item['明細_受注数量']
                                self_shipment_quantity = item.get('明細_自社出荷数量', 0)
                                remaining_quantity = order_quantity - self_shipment_quantity
                                default_shipment = min(previous_quantity, remaining_quantity)
                            else:
                                # 通常依頼の場合：入荷・在庫引当済の値（受注残数を上限とする）
                                self_stock_quantity = item.get('明細_自社在庫引当数量', 0)
                                order_quantity = item['明細_受注数量']
                                self_shipment_quantity = item.get('明細_自社出荷数量', 0)
                                remaining_quantity = order_quantity - self_shipment_quantity
                                default_shipment = min(self_stock_quantity, remaining_quantity)
                            
                            page_shipment_quantities[i_no] = max(0, default_shipment)
                    
                    # ★修正：該当ページの梱包明細を取得（統合明細対応）★
                    if page_index in self.consolidated_packing_data_by_page:
                        # 統合明細がある場合
                        packing_data = self.consolidated_packing_data_by_page[page_index]
                        selected_shipment = "統合明細"
                        print(f"ページ{page_index + 1}: 統合明細を使用")  # デバッグ
                    else:
                        # 通常の明細
                        selected_shipment = self.selected_shipment_numbers.get(page_index, "")
                        packing_data = []
                        if selected_shipment:
                            packing_data = get_packing_details_by_shipment(selected_shipment)
                        print(f"ページ{page_index + 1}: 通常明細を使用 ({selected_shipment})")  # デバッグ
                    
                    # 修正：該当ページの受渡場所を取得
                    # 一括出力では受渡場所変更は考慮せず、元の値を使用
                    custom_delivery_location = None
                    if page_index in self.selected_delivery_locations and self.change_delivery_location.get():
                        custom_delivery_location = self.selected_delivery_locations[page_index]
                    
                    # PDFを作成
                    pdf_content = self.create_pdf_report(page_data, page_key, page_shipment_quantities, 
                                                       shipment_date, selected_shipment, packing_data,
                                                       custom_delivery_location)
                    pdf_list.append(pdf_content)
                    
                except Exception as e:
                    print(f"ページ {page_index + 1} のPDF作成でエラー: {e}")
                    continue
            
            if not pdf_list:
                messagebox.showerror("エラー", "PDF作成に失敗しました。")
                return
            
            # PDFを結合
            from PyPDF2 import PdfWriter, PdfReader
            merger = PdfWriter()
            
            for pdf_content in pdf_list:
                pdf_reader = PdfReader(io.BytesIO(pdf_content))
                for page in pdf_reader.pages:
                    merger.add_page(page)
            
            # 結合されたPDFを保存
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")

            # 修正：一括出力時も統一されたファイル名形式を使用
            # 出荷日から年月日（YYMMDD）を取得
            shipment_date_str = shipment_date.strftime("%y%m%d")

            # 一括出力時は複数の客注番号があるため、タイムスタンプを併用
            timestamp = datetime.now().strftime("%H%M%S")
            filename = f"{shipment_date_str}_一括出荷報告書_{timestamp}.pdf"
            output_path = os.path.join(documents_path, filename)

            print(f"一括出力ファイル名: {filename}")  # デバッグ
            
            with open(output_path, 'wb') as output_file:
                merger.write(output_file)
            
            merger.close()
            
            messagebox.showinfo("完了", f"PDF一括出荷報告書が出力されました。\n保存先: {output_path}\n{len(pdf_list)}件のレポートを結合しました。")
            
            # ファイルを開くかどうか確認
            if messagebox.askyesno("確認", "PDF一括出荷報告書を開きますか？"):
                try:
                    subprocess.Popen(["start", "", output_path], shell=True)
                except Exception as e:
                    messagebox.showerror("エラー", f"ファイルを開く際にエラーが発生しました:\n{str(e)}")
                    
        except Exception as e:
            error_message = f"PDF一括出荷報告書出力中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
            messagebox.showerror("エラー", error_message)
            with open("error_log.txt", "w", encoding='utf-8') as f:
                f.write(error_message)

    def show_consolidation_dialog(self):
        """梱包明細統合ダイアログを表示"""
        current_data = self.get_current_search_data()
        if current_data.empty:
            messagebox.showwarning("警告", "データがありません。")
            return
        
        order_number = current_data.iloc[0]['受注番号']
        
        # 統合ダイアログを表示
        consolidation_dialog = PackingConsolidationDialog(self.dialog, order_number)
        self.dialog.wait_window(consolidation_dialog.dialog)
        
        if consolidation_dialog.result is not None:
            # ★修正：統合結果を現在のページに保存★
            self.consolidated_packing_data_by_page[self.current_page] = consolidation_dialog.result
            print(f"ページ{self.current_page + 1}に統合明細設定完了: {len(consolidation_dialog.result)}件")  # デバッグ
            
            # デバッグ: 保存された統合データの内容を確認
            for i, item in enumerate(self.consolidated_packing_data_by_page[self.current_page]):
                print(f"ページ{self.current_page + 1}の統合データ[{i}]: {item}")
            
            messagebox.showinfo("完了", f"現在のページ（{self.current_page + 1}ページ目）の梱包明細統合設定が完了しました。\n統合後のケース数: {len(consolidation_dialog.result)}")
        else:
            print("統合ダイアログがキャンセルされました")  # デバッグ

    def on_close(self):
        """閉じるボタンが押された時の処理"""
        self.dialog.destroy()

def search_shipment_status():
    """
    依頼状況確認の検索処理（複数番号対応）
    """
    try:
        # 入力値を取得（複数行対応）
        search_keys_text = status_search_text.get("1.0", tk.END).strip()
        search_method = status_search_method_var.get()
        shipment_request_type = shipment_request_type_var.get()
        
        if not search_keys_text:
            messagebox.showwarning("警告", "検索キーを入力してください。")
            return
        
        # 改行で分割して複数の検索キーを取得
        search_keys = [key.strip() for key in search_keys_text.split('\n') if key.strip()]
        
        if not search_keys:
            messagebox.showwarning("警告", "有効な検索キーを入力してください。")
            return
        
        # 追加：重複チェックと除去
        original_count = len(search_keys)
        
        # 重複を検出
        seen = set()
        duplicates = set()
        for key in search_keys:
            if key in seen:
                duplicates.add(key)
            else:
                seen.add(key)
        
        # 重複がある場合は警告を表示
        if duplicates:
            duplicate_list = ', '.join(sorted(duplicates))
            search_type_name = "受注番号" if search_method == "order_number" else "見積管理番号"
            messagebox.showwarning(
                "重複検出", 
                f"以下の{search_type_name}が重複して入力されています：\n{duplicate_list}\n\n重複を除去して検索を続行します。"
            )
        
        # 重複を除去（入力順序を保持）
        search_keys = list(dict.fromkeys(search_keys))
        
        # 重複除去後の件数確認メッセージ
        if len(search_keys) != original_count:
            removed_count = original_count - len(search_keys)
            if not messagebox.askyesno("確認", f"重複を除去した結果、{len(search_keys)}件の番号で検索を実行します。\n（{removed_count}件の重複を除去）\n\n検索を続行しますか？"):
                return
        elif len(search_keys) > 1:
            if not messagebox.askyesno("確認", f"{len(search_keys)}件の番号で検索を実行しますか？"):
                return
        
        # 発注データをキャッシュから取得
        df_order = get_cached_purchase_order_data()
        df1 = df_order[['発注番号','明細_商品コード','明細_共通項目2','受注番号']].copy()
        df1['A'] = df1['発注番号'] + df1['明細_商品コード'] + df1['明細_共通項目2']
        df1['B'] = df1['明細_共通項目2']
        df1['C'] = df1['受注番号']
        df1['D'] = df1['受注番号'] + df1['明細_共通項目2']
        df1['E'] = df1['発注番号']
        df1 = df1[['A','B','C','D','E']]

        # 受注データをキャッシュから取得
        df_order_list = get_cached_order_data()
        excluded_codes = ['888888-88888','777777-77777']
        columns_needed = [
            '受注番号','客注番号','得意先名','得意先','受渡場所名','社員名','明細_倉庫コード',  # ★「得意先」を追加
            '明細_共通項目2','明細_商品コード','明細_商品受注名',
            '明細_発注引当仕入数量','明細_受注数量','受注件名','明細_共通項目3',
            '明細_出荷売上数量','明細_自社在庫引当数量','明細_直接売上数量',
            '明細_自社出荷数量'
        ]
        
        # ★修正：得意先コード-得意先枝番の列を追加★
        # 列が存在するかチェック
        available_customer_code_columns = []
        for col in ['得意先コード-得意先枝番', '得意先コード', '得意先枝番']:
            if col in df_order_list.columns:
                available_customer_code_columns.append(col)
        
        # 得意先コード-得意先枝番が直接ある場合
        if '得意先コード-得意先枝番' in df_order_list.columns:
            columns_needed.append('得意先コード-得意先枝番')
        # 得意先コードと得意先枝番が別々にある場合
        elif '得意先コード' in df_order_list.columns and '得意先枝番' in df_order_list.columns:
            columns_needed.extend(['得意先コード', '得意先枝番'])
        # 得意先コードのみある場合
        elif '得意先コード' in df_order_list.columns:
            columns_needed.append('得意先コード')
        
        df4 = df_order_list[~df_order_list['明細_商品コード'].isin(excluded_codes)][columns_needed].copy()
        
        # ★修正：得意先コード-得意先枝番の列を作成★
        if '得意先コード-得意先枝番' not in df4.columns:
            if '得意先コード' in df4.columns and '得意先枝番' in df4.columns:
                # 得意先コードと得意先枝番を結合
                df4['得意先コード-得意先枝番'] = df4['得意先コード'].astype(str) + '-' + df4['得意先枝番'].astype(str)
            elif '得意先コード' in df4.columns:
                # 得意先コードのみの場合はそのまま使用
                df4['得意先コード-得意先枝番'] = df4['得意先コード'].astype(str)
            else:
                # 得意先コード関連の列がない場合は得意先名を使用
                df4['得意先コード-得意先枝番'] = df4['得意先名']
        
        # 発注番号をマージで追加
        df4['key'] = df4['受注番号'] + df4['明細_共通項目2']
        df4 = pd.merge(df4, df1[['D','E']], left_on='key', right_on='D', how='left')
        df4.rename(columns={'E':'発注番号'}, inplace=True)
        df4.drop(['key','D'], axis=1, inplace=True)
        
        # データ型変換
        for col in df4.columns:
            if col in ['明細_共通項目2','明細_発注引当仕入数量','明細_受注数量','明細_出荷売上数量','明細_自社在庫引当数量','明細_直接売上数量','明細_自社出荷数量']:
                df4[col] = pd.to_numeric(df4[col], errors='coerce').fillna(0).astype(int)
            elif col == '発注番号':
                df4[col] = df4[col].fillna('')
            else:
                df4[col] = df4[col].replace('nan','').fillna('')
        
        # 複数の検索キーに対してデータを取得
        search_data_list = []
        valid_search_keys = []
        not_found_keys = []
        
        for search_key in search_keys:
            if search_method == "order_number":
                search_data = df4[df4['受注番号'] == search_key]
            else:
                search_data = df4[df4['明細_共通項目3'] == search_key]
            
            if not search_data.empty:
                search_data_list.append(search_data)
                valid_search_keys.append(search_key)
            else:
                not_found_keys.append(search_key)
                print(f"警告: {search_method}「{search_key}」のデータが見つかりませんでした。")
        
        # 見つからなかった番号がある場合は警告を表示
        if not_found_keys:
            not_found_message = f"以下の番号はデータが見つかりませんでした：\n" + "\n".join(not_found_keys)
            if valid_search_keys:
                not_found_message += f"\n\n見つかった{len(valid_search_keys)}件のデータで検索結果を表示します。"
                messagebox.showwarning("一部データなし", not_found_message)
            else:
                messagebox.showerror("エラー", not_found_message)
                return
        
        if not search_data_list:
            messagebox.showerror("エラー", "指定された検索条件に一致するデータが見つかりません。")
            return
        
        # 成功メッセージ
        if len(valid_search_keys) > 1:
            messagebox.showinfo("検索完了", f"{len(valid_search_keys)}件のデータが見つかりました。\nページ遷移ボタンで各データを確認できます。")
        
        # ダイアログを表示（出荷依頼形式パラメータを追加）
        dialog = ShipmentStatusDialog(root, search_data_list, valid_search_keys, search_method, shipment_request_type)
        root.wait_window(dialog.dialog)
        
    except Exception as e:
        error_message = f"検索処理中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
        messagebox.showerror("エラー", error_message)
        with open("error_log.txt", "w", encoding='utf-8') as f:
            f.write(error_message)
        print("エラーログが error_log.txt に保存されました。")
        
# -------------------------------------------------
# ▼ GUI作成（タブ付きインターフェース）
# -------------------------------------------------
minimize_console_window()
root = tk.Tk()
root.title(f"{APP_NAME} v{APP_VERSION}")

# アプリ起動時にアップデートチェック
check_and_prompt_update(root)

# タブコントロールの作成
tab_control = ttk.Notebook(root)

# タブ1（既存の処理）
tab1 = ttk.Frame(tab_control)
tab_control.add(tab1, text='新規梱包依頼')

# タブ2（ファイルを開く機能）
tab2 = ttk.Frame(tab_control)

# タブ3（欠品QR機能）
tab3 = ttk.Frame(tab_control)

tab_control.pack(expand=1, fill="both")

# タブ1の内容を表形式に変更
frame = tk.Frame(tab1, padx=20, pady=20)
frame.pack(fill='both', expand=True)

label = tk.Label(frame, text="梱包依頼書生成", font=("Arial", 14, "bold"))
label.pack(pady=10)

# 表形式の入力エリア
table_frame = tk.Frame(frame)
table_frame.pack(fill='both', expand=True, pady=10)

# Treeviewを使用した表形式のGUI - 列名を独自の変数にする
packing_columns = ('検索方法', '番号', '梱包期限日', '梱包依頼摘要', '梱包担当者', '出力除外選択', '受注残0出力', '梱包可能数変更', '梱包明細')
packing_tree = ttk.Treeview(table_frame, columns=packing_columns, show='headings', height=12)

# 列の設定
packing_tree.heading('検索方法', text='検索方法')
packing_tree.heading('番号', text='番号')
packing_tree.heading('梱包期限日', text='梱包期限日')
packing_tree.heading('梱包依頼摘要', text='梱包依頼摘要')
packing_tree.heading('梱包担当者', text='梱包担当者')
packing_tree.heading('出力除外選択', text='出力アイテム選択')
packing_tree.heading('受注残0出力', text='受注残0出力')
packing_tree.heading('梱包可能数変更', text='梱包可能数変更')
packing_tree.heading('梱包明細', text='梱包明細')

# 列幅の設定
packing_tree.column('検索方法', width=100, anchor='center')
packing_tree.column('番号', width=150, anchor='center')
packing_tree.column('梱包期限日', width=120, anchor='center')
packing_tree.column('梱包依頼摘要', width=200, anchor='w')
packing_tree.column('梱包担当者', width=100, anchor='center')
packing_tree.column('出力除外選択', width=100, anchor='center')
packing_tree.column('受注残0出力', width=100, anchor='center')
packing_tree.column('梱包可能数変更', width=120, anchor='center')
packing_tree.column('梱包明細', width=200, anchor='center')

# スクロールバー
tree_scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=packing_tree.yview)
packing_tree.configure(yscrollcommand=tree_scrollbar.set)
packing_tree.bind("<Double-1>", on_packing_tree_double_click)

packing_tree.pack(side=tk.LEFT, fill='both', expand=True)
tree_scrollbar.pack(side=tk.RIGHT, fill='y')

# 行操作ボタンフレーム
button_frame = tk.Frame(frame)
button_frame.pack(fill='x', pady=10)

add_row_button = Button(button_frame, text="行追加", command=lambda: add_packing_row())
add_row_button.pack(side=tk.LEFT, padx=5)

edit_row_button = Button(button_frame, text="行編集", command=lambda: edit_packing_row())
edit_row_button.pack(side=tk.LEFT, padx=5)

delete_row_button = Button(button_frame, text="行削除", command=lambda: delete_packing_row())
delete_row_button.pack(side=tk.LEFT, padx=5)

# 実行ボタンフレーム
execute_frame = tk.Frame(frame)
execute_frame.pack(fill='x', pady=10)

execute_selected_button = Button(execute_frame, text="選択行を実行", command=lambda: execute_selected_rows(), 
                                bg="lightgreen", font=("Arial", 11, "bold"))
execute_selected_button.pack(side=tk.LEFT, padx=10)

execute_all_button = Button(execute_frame, text="全行を実行", command=lambda: execute_all_rows(), 
                           bg="orange", font=("Arial", 11, "bold"))
execute_all_button.pack(side=tk.LEFT, padx=10)

# データ保存用のリスト
packing_data_list = []

# タブ2の内容（ファイルを開く機能 - 変更なし）
frame2 = tk.Frame(tab2, padx=20, pady=20)
frame2.pack()
label2 = tk.Label(frame2, text="事前梱包依頼書ファイルを開く")
label2.pack(pady=10)

request_number_label = tk.Label(frame2, text="事前梱包依頼番号:")
request_number_label.pack(pady=5)
request_number_entry = Entry(frame2, width=30)
request_number_entry.pack(pady=5)

file_packing_person_label = tk.Label(frame2, text="梱包担当者:")
file_packing_person_label.pack(pady=5)
file_packing_person_var = StringVar()
file_packing_person_var.set("11_細田")
file_packing_person_option = tk.OptionMenu(frame2, file_packing_person_var, "11_細田", "12_平松", "13_坂上", "16_土田")
file_packing_person_option.pack(pady=5)

open_file_button = Button(frame2, text="ファイルを検索して開く", command=find_and_open_packing_file)
open_file_button.pack(pady=20)

# タブ3の内容（欠品QR機能 - レイアウト改善版）
frame3 = tk.Frame(tab3, padx=20, pady=20)
frame3.pack(fill='both', expand=True)

# タイトル
label3 = tk.Label(frame3, text="見積管理番号/発注番号/受注番号と行番号からQRコードを表示", font=("Arial", 12, "bold"))
label3.pack(pady=10)

# メインコンテンツエリア（左右分割）
main_content = tk.Frame(frame3)
main_content.pack(fill='both', expand=True, pady=10)

# 左側フレーム（入力エリア）
left_frame = tk.Frame(main_content)
left_frame.pack(side=tk.LEFT, fill='both', expand=True, padx=(0, 20))

# 見積管理番号入力
missing_estimate_no_label = tk.Label(left_frame, text="見積管理番号:")
missing_estimate_no_label.pack(pady=5, anchor='w')
missing_estimate_no_entry = Entry(left_frame, width=30)
missing_estimate_no_entry.pack(pady=5, anchor='w')

# 発注番号入力
missing_order_no_label = tk.Label(left_frame, text="発注番号:")
missing_order_no_label.pack(pady=5, anchor='w')
missing_order_no_entry = Entry(left_frame, width=30)
missing_order_no_entry.pack(pady=5, anchor='w')

# 受注番号入力
missing_order_number_label = tk.Label(left_frame, text="受注番号:")
missing_order_number_label.pack(pady=5, anchor='w')
missing_order_number_entry = Entry(left_frame, width=30)
missing_order_number_entry.pack(pady=5, anchor='w')

# 注意書き
note_label = tk.Label(left_frame, text="※見積管理番号・発注番号・受注番号のいずれか1つだけ入力してください", 
                     font=("Arial", 9), fg="blue")
note_label.pack(pady=(0, 10), anchor='w')

# 行番号入力
missing_line_no_label = tk.Label(left_frame, text="行番号:")
missing_line_no_label.pack(pady=5, anchor='w')
missing_line_no_entry = Entry(left_frame, width=30)
missing_line_no_entry.pack(pady=5, anchor='w')

# 数量入力
missing_quantity_label = tk.Label(left_frame, text="数量（1～999）:")
missing_quantity_label.pack(pady=5, anchor='w')
missing_quantity_entry = Entry(left_frame, width=30)
missing_quantity_entry.pack(pady=5, anchor='w')

# QRコード生成ボタン
generate_missing_qr_button = Button(left_frame, text="QRコードを生成", command=generate_missing_item_qr, 
                                   font=("Arial", 11), bg="lightgreen", padx=20)
generate_missing_qr_button.pack(pady=15, anchor='w')

# 結果表示用のラベル（ロット番号、商品コード、商品略名）
missing_info_var = StringVar()
missing_info_var.set("")
missing_info_label = tk.Label(left_frame, textvariable=missing_info_var, font=("Arial", 10, "bold"), 
                             justify=tk.LEFT, anchor='w')
missing_info_label.pack(pady=10, fill=tk.X, anchor='w')

# 右側フレーム（QR表示エリア）
right_frame = tk.Frame(main_content, relief=tk.RIDGE, borderwidth=2, bg="white")
right_frame.pack(side=tk.RIGHT, fill='both', padx=(20, 0))

# QRコード表示用のラベル
missing_qr_display_label = tk.Label(right_frame, bg="white")
missing_qr_display_label.pack(padx=20, pady=20)

# 使用方法の説明
help_text_tab3 = """
【使用方法】
1. 見積管理番号・発注番号・受注番号のいずれか1つを入力
2. 行番号を入力
3. 数量を入力（1～99の範囲）
4. 「QRコードを生成」ボタンを押す

【注意事項】
・見積管理番号・発注番号・受注番号は同時入力しないでください
・数量は必須です（空欄不可）
"""

help_label_tab3 = tk.Label(frame3, text=help_text_tab3, font=("Arial", 9), justify=tk.LEFT, 
                          fg="darkgreen", relief=tk.GROOVE, borderwidth=1, padx=10, pady=10)
help_label_tab3.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

# タブ4（梱包依頼閲覧）
tab4 = ttk.Frame(tab_control)
tab_control.add(tab4, text='梱包依頼閲覧＆削除')

frame4 = tk.Frame(tab4, padx=20, pady=20)
frame4.pack(fill='both', expand=True)

# 検索方法選択用ラジオボタン
search_method_view_var = StringVar(value="order_number")
view_search_frame = tk.Frame(frame4)
view_search_frame.pack(anchor='w', pady=(0,10))
tk.Label(view_search_frame, text="検索方法:").pack(side=tk.LEFT)
Radiobutton(view_search_frame, text="受注番号", variable=search_method_view_var, value="order_number").pack(side=tk.LEFT, padx=5)
Radiobutton(view_search_frame, text="見積管理番号", variable=search_method_view_var, value="estimate_no").pack(side=tk.LEFT)

# 入力欄＋検索ボタン
view_input_label_var = StringVar(value="受注番号:")
view_input_frame = tk.Frame(frame4)
view_input_frame.pack(anchor='w', fill='x', pady=(0, 10))
view_input_label = Label(view_input_frame, textvariable=view_input_label_var)
view_input_label.pack(side=tk.LEFT)
view_input_entry = Entry(view_input_frame, width=30)
view_input_entry.pack(side=tk.LEFT, padx=5)

def update_view_input_label(*args):
    if search_method_view_var.get() == "order_number":
        view_input_label_var.set("受注番号:")
    else:
        view_input_label_var.set("見積管理番号:")
search_method_view_var.trace_add("write", update_view_input_label)

# 検索ボタン
view_button = Button(view_input_frame, text="検索", command=lambda: view_packing_requests())
view_button.pack(side=tk.LEFT, padx=8)

# 削除対象選択
view_result_records = []
view_target_unique_var = StringVar()
view_action_frame = tk.Frame(frame4)
view_action_frame.pack(anchor='w', pady=(0, 10))
tk.Label(view_action_frame, text="削除対象（事前梱包依頼番号）:").pack(side=tk.LEFT)
view_target_combo = ttk.Combobox(view_action_frame, textvariable=view_target_unique_var, state="readonly", width=40)
view_target_combo.pack(side=tk.LEFT, padx=5)

# 結果表示用 Treeview
columns = ('項目','値')
tree = ttk.Treeview(frame4, columns=columns, show='headings', height=12)
tree.heading('項目', text='項目')
tree.heading('値', text='値')
tree.column('項目', width=150, anchor='w')
tree.column('値', width=400, anchor='w')
tree.pack(fill='both', expand=True)

def view_packing_requests():
    """
    梱包依頼を検索して表示する関数
    """
    key = view_input_entry.get().strip()
    if not key:
        messagebox.showwarning("警告", "検索キーを入力してください。")
        return

    try:
        # DBからヘッダ情報のみ取得
        db_path = get_db_path()
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        if search_method_view_var.get() == "order_number":
            cursor.execute("""
                SELECT unique_number, order_number, deadline, estimate_no,
                       ship_name, customer_order_no, order_numbers,
                       customer_name, delivery_location, salesperson,
                       packing_person, packaging_note, item_count, order_amount, packing_detail
                FROM packing_requests
                WHERE order_number = ?
                  AND IFNULL(is_deleted, 0) = 0
                ORDER BY created_at DESC
            """, (key,))
        else:
            cursor.execute("""
                SELECT unique_number, order_number, deadline, estimate_no,
                       ship_name, customer_order_no, order_numbers,
                       customer_name, delivery_location, salesperson,
                       packing_person, packaging_note, item_count, order_amount, packing_detail
                FROM packing_requests
                WHERE estimate_no = ?
                  AND IFNULL(is_deleted, 0) = 0
                ORDER BY created_at DESC
            """, (key,))
        
        rows = cursor.fetchall()
        conn.close()

        # Treeviewクリア
        for iid in tree.get_children():
            tree.delete(iid)
        view_target_combo['values'] = []
        view_target_unique_var.set("")
        view_result_records.clear()

        if not rows:
            messagebox.showinfo("情報", f"該当する梱包依頼が見つかりません：{key}")
            return

        # 削除対象候補を更新
        unique_numbers = [str(r[0]) for r in rows if r and r[0]]
        view_result_records.extend(rows)
        view_target_combo['values'] = unique_numbers
        if unique_numbers:
            view_target_unique_var.set(unique_numbers[0])

        # 項目名リスト（受注金額を追加）
        display_fields = [
            "事前梱包依頼番号", "受注番号", "梱包期限日", "見積管理番号",
            "船名", "客注番号", "発注番号", "得意先名",
            "受渡場所名", "営業担当者", "梱包担当者", "梱包依頼摘要", 
            "アイテム数", "受注金額"
        ]

        # 複数ヒット時は続けて表示
        for row in rows:
            for idx, field in enumerate(display_fields):
                # アイテム数の表示処理
                if field == "アイテム数":
                    item_count = row[idx] if row[idx] is not None else "未設定"
                    tree.insert('', 'end', values=(field, f"{item_count}件"))
                # 受注金額の表示処理
                elif field == "受注金額":
                    order_amount = row[idx] if row[idx] is not None else 0.0
                    tree.insert('', 'end', values=(field, f"¥{order_amount:,.0f}"))  # ★追加：金額表示
                else:
                    tree.insert('', 'end', values=(field, row[idx]))
            packing_detail_value = row[-1] if row[-1] is not None else 0
            detail_text = "梱包明細を依頼する案件" if int(packing_detail_value) == 1 else "通常"
            tree.insert('', 'end', values=("梱包明細", detail_text))
            tree.insert('', 'end', values=('', ''))  # レコード間の空行
            
    except Exception as e:
        error_message = f"検索処理中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
        messagebox.showerror("エラー", error_message)
        with open("error_log.txt", "w", encoding='utf-8') as f:
            f.write(error_message)
        print("エラーログが error_log.txt に保存されました。")

def delete_selected_packing_request():
    """梱包依頼閲覧タブで選択されたレコードを論理削除する"""
    unique_number = view_target_unique_var.get().strip()
    if not unique_number:
        messagebox.showwarning("警告", "削除対象の事前梱包依頼番号を選択してください。")
        return

    confirm_message = (
        f"以下のレコードを削除します。\n\n"
        f"事前梱包依頼番号: {unique_number}\n\n"
        f"関連Excelファイルは「削除済み」フォルダへ退避します。続行しますか？"
    )
    if not messagebox.askyesno("削除確認", confirm_message):
        return

    success, message = soft_delete_packing_request(unique_number, "")
    if success:
        messagebox.showinfo("完了", message)
        view_packing_requests()
    else:
        messagebox.showerror("エラー", message)

delete_button = Button(
    view_action_frame,
    text="選択レコードを削除",
    command=delete_selected_packing_request,
    bg="lightcoral"
)
delete_button.pack(side=tk.LEFT, padx=5)

# タブ3をこの位置で追加（並び順: 新規梱包依頼 → 梱包依頼閲覧＆削除 → 欠品QR）
tab_control.add(tab3, text='欠品QR')

# タブ5（依頼状況確認）
tab5 = ttk.Frame(tab_control)
tab_control.add(tab5, text='依頼状況確認')

frame5 = tk.Frame(tab5, padx=20, pady=20)
frame5.pack()

label5 = tk.Label(frame5, text="依頼状況確認")
label5.pack(pady=10)

# 検索方法選択用のラジオボタン
status_search_method_frame = tk.Frame(frame5)
status_search_method_frame.pack(pady=5)
status_search_method_var = StringVar()
status_search_method_var.set("order_number")  # デフォルトは受注番号検索

status_search_method_label = tk.Label(status_search_method_frame, text="検索方法:")
status_search_method_label.pack(side=tk.LEFT, padx=5)
status_order_number_radio = Radiobutton(status_search_method_frame, text="受注番号", variable=status_search_method_var, value="order_number")
status_order_number_radio.pack(side=tk.LEFT, padx=5)
status_estimate_no_radio = Radiobutton(status_search_method_frame, text="見積管理番号", variable=status_search_method_var, value="estimate_no")
status_estimate_no_radio.pack(side=tk.LEFT, padx=5)

# ★新規追加：出荷依頼形式選択用のラジオボタン★
shipment_request_type_frame = tk.Frame(frame5)
shipment_request_type_frame.pack(pady=5)
shipment_request_type_var = StringVar()
shipment_request_type_var.set("advanced_packing")  # デフォルトは事前梱包

shipment_request_type_label = tk.Label(shipment_request_type_frame, text="出荷依頼形式:")
shipment_request_type_label.pack(side=tk.LEFT, padx=5)
advanced_packing_radio = Radiobutton(shipment_request_type_frame, text="事前梱包依頼", variable=shipment_request_type_var, value="advanced_packing")
advanced_packing_radio.pack(side=tk.LEFT, padx=5)
normal_request_radio = Radiobutton(shipment_request_type_frame, text="通常依頼", variable=shipment_request_type_var, value="normal_request")
normal_request_radio.pack(side=tk.LEFT, padx=5)

# 入力フィールドのラベルを動的に変更
status_input_label_var = StringVar()
status_input_label_var.set("受注番号（複数可、改行区切り）:")
status_input_label = tk.Label(frame5, textvariable=status_input_label_var)
status_input_label.pack(pady=5)

# ラジオボタン切り替え時のラベル更新関数
def update_status_input_label(*args):
    if status_search_method_var.get() == "order_number":
        status_input_label_var.set("受注番号（複数可、改行区切り）:")
    else:
        status_input_label_var.set("見積管理番号（複数可、改行区切り）:")

# ラジオボタン変更時にラベルを更新するよう設定
status_search_method_var.trace("w", update_status_input_label)

# EntryをTextウィジェットに変更（複数行入力対応）
status_search_text = Text(frame5, width=40, height=5, font=("Arial", 10))
status_search_text.pack(pady=5)

# 使用方法の説明ラベルを追加
usage_info_label = tk.Label(frame5, text="※複数の番号を検索する場合は、1行に1つずつ入力してください", 
                           font=("Arial", 9), fg="blue")
usage_info_label.pack(pady=(0, 10))

status_search_button = Button(frame5, text="検索", command=search_shipment_status, 
                             font=("Arial", 11), bg="lightgreen", padx=20)
status_search_button.pack(pady=20)

# クリアボタンの追加
clear_button = Button(frame5, text="入力クリア", command=lambda: status_search_text.delete("1.0", tk.END), 
                     font=("Arial", 10), bg="lightcoral", padx=15)
clear_button.pack(pady=5)

# 使用方法の詳細説明
help_text = """
【使用方法】
1. 検索方法を選択してください（受注番号 または 見積管理番号）
2. 検索したい番号を入力してください
   - 単一番号の場合：そのまま入力
   - 複数番号の場合：1行に1つずつ入力
3. 「検索」ボタンを押してください
4. 複数件見つかった場合は「前ページ」「次ページ」で遷移できます
5. 見積管理番号で検索した場合、対応する受注番号も表示されます

【機能】
・引当済と依頼済数量比較：数量に差異があるアイテムを確認
・出荷報告書出力：現在表示中のページの出荷報告書を作成
・出荷日設定：報告書に記載する出荷日を指定可能
"""

help_label = tk.Label(frame5, text=help_text, font=("Arial", 9), justify=tk.LEFT, 
                     fg="darkgreen", relief=tk.GROOVE, borderwidth=1, padx=10, pady=10)
help_label.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

# タブ6（出荷依頼番号QR）
tab6 = ttk.Frame(tab_control)
tab_control.add(tab6, text='出荷依頼番号QR')

# タブ2（ファイルを開く機能）を最後に配置
tab_control.add(tab2, text='ファイルを開く')

frame6 = tk.Frame(tab6, padx=20, pady=20)
frame6.pack(fill='both', expand=True)

label6 = tk.Label(frame6, text="出荷依頼番号QR生成", font=("Arial", 14, "bold"))
label6.pack(pady=10)

# 入力セクション
input_section = tk.Frame(frame6)
input_section.pack(pady=20)

# 受注番号入力
order_number_frame = tk.Frame(input_section)
order_number_frame.pack(pady=8)
tk.Label(order_number_frame, text="受注番号:", font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
shipment_order_number_entry = Entry(order_number_frame, width=20, font=("Arial", 10))
shipment_order_number_entry.pack(side=tk.LEFT, padx=5)

# 出荷予定日入力（カレンダー形式）
date_frame = tk.Frame(input_section)
date_frame.pack(pady=8)
tk.Label(date_frame, text="出荷予定日:", font=("Arial", 10)).pack(side=tk.LEFT, padx=5)
# tkcalendarのDateEntryを使用
shipment_date_cal = DateEntry(date_frame, width=12, background='darkblue', foreground='white', 
                             borderwidth=2, locale='ja_JP', font=("Arial", 10))
shipment_date_cal.pack(side=tk.LEFT, padx=5)

# 検索ボタン
search_shipment_button = Button(input_section, text="QRコード生成", command=search_shipment_request_numbers, 
                               font=("Arial", 11), bg="lightblue", padx=20)
search_shipment_button.pack(pady=15)

# QRコード表示エリア
qr_section = tk.Frame(frame6)
qr_section.pack(fill='both', expand=True, pady=10)

# QRコード表示用のスクロール可能フレーム
qr_canvas = tk.Canvas(qr_section, bg="white")
qr_scrollbar = tk.Scrollbar(qr_section, orient="vertical", command=qr_canvas.yview)
shipment_qr_display_frame = tk.Frame(qr_canvas, bg="white")

shipment_qr_display_frame.bind(
    "<Configure>",
    lambda e: qr_canvas.configure(scrollregion=qr_canvas.bbox("all"))
)

qr_canvas.create_window((0, 0), window=shipment_qr_display_frame, anchor="nw")
qr_canvas.configure(yscrollcommand=qr_scrollbar.set)

qr_canvas.pack(side="left", fill="both", expand=True)
qr_scrollbar.pack(side="right", fill="y")

# 使用方法の説明
help_text = """
【使用方法】
1. 受注番号を入力してください
2. 出荷予定日をカレンダーから選択してください
3. 「QRコード生成」ボタンを押してください
4. 該当する出荷依頼番号のQRコードが詳細情報と共に表示されます

※複数の出荷依頼番号が該当する場合は、全て表示されます
※QRコードと共に受注件名、受渡場所、配送便の情報も表示されます
"""

help_label = tk.Label(frame6, text=help_text, font=("Arial", 9), justify=tk.LEFT, 
                     fg="darkgreen", relief=tk.GROOVE, borderwidth=1, padx=10, pady=10)
help_label.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

# データベースの初期化
try:
    init_database()
    print("データベースの初期化が完了しました。")
except Exception as e:
    print(f"データベースの初期化でエラーが発生しました: {e}")
    messagebox.showerror("エラー", f"データベースの初期化に失敗しました:\n{str(e)}")

# メインループ開始
if __name__ == "__main__":
    try:
        root.mainloop()
    except Exception as e:
        error_message = f"アプリケーション実行中にエラーが発生しました：{str(e)}\n\n詳細なエラー情報:\n{traceback.format_exc()}"
        print(error_message)
        with open("error_log.txt", "w", encoding='utf-8') as f:
            f.write(error_message)
        messagebox.showerror("エラー", "予期しないエラーが発生しました。詳細はerror_log.txtを確認してください。")
